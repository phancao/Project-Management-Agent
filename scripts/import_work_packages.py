#!/usr/bin/env python3
"""
Import work packages from an Excel workbook into OpenProject.

Workflow:
1. Prompt for the OpenProject server URL and API token.
2. Read the Excel workbook (default path points to the intranet work
   packages file).
3. Validate that every user and task type referenced in the workbook
   exists in OpenProject, automatically creating any missing users.
4. Optionally create missing projects.
5. Create work packages for each row in the workbook,
   assigning them to the matching user and task type.

Requirements:
    pip install openpyxl requests

This script is intentionally cautious. It will abort before creating anything
if referenced task types are missing. Use the --dry-run flag to see what would
happen without making any API changes.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import math
import os
import re
import secrets
import subprocess
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, time, timedelta, timezone
from getpass import getpass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

import requests
import time as time_module


API_PREFIX = "/api/v3"
CURRENT_OP_VERSION: Optional[str] = None
DEFAULT_CREATION_TIME = time(9, 0, tzinfo=timezone.utc)
DEFAULT_LOG_TIME = time(9, 0, tzinfo=timezone.utc)
OPENPROJECT_CONTAINER = os.environ.get(
    "OPENPROJECT_CONTAINER",
    "project-management-agent-openproject-1",
)
OPENPROJECT_DB_HOST = os.environ.get(
    "OPENPROJECT_DB_HOST",
    "127.0.0.1",
)
OPENPROJECT_DB_USER = os.environ.get(
    "OPENPROJECT_DB_USER",
    "openproject",
)
OPENPROJECT_DB_PASSWORD = os.environ.get(
    "OPENPROJECT_DB_PASSWORD",
    "openproject",
)
OPENPROJECT_DB_NAME = os.environ.get(
    "OPENPROJECT_DB_NAME",
    "openproject",
)

# Global toggle for verification source (Step 20)
VERIFICATION_SOURCE: str = "auto"
# Default path for time entry cache (persists created entries between runs)
DEFAULT_TIME_ENTRY_CACHE = Path("/tmp/op_time_entry_cache.json")
# Global debug log file (optional; if set, detailed logs are duplicated to this file)
DEBUG_LOG_PATH: Optional[Path] = None

def _write_debug_log(level: str, message: str) -> None:
    """Append a line to the debug log if configured."""
    global DEBUG_LOG_PATH
    if not DEBUG_LOG_PATH:
        return
    try:
        stamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Strip ANSI colors if any
        ansi_escape = re.compile(r"\x1B\[[0-?]*[ -/]*[@-~]")
        plain = ansi_escape.sub("", message)
        with DEBUG_LOG_PATH.open("a", encoding="utf-8") as fh:
            fh.write(f"[{stamp}] [{level.upper()}] {plain}\n")
    except Exception:
        # Never fail the main flow due to logging
        pass

def _fetch_counts_from_db(base_url: Optional[str], ui: Optional[ConsoleUI] = None) -> Dict[str, int]:
    container_name = _get_openproject_container(base_url)
    sql = """
        SELECT
          (SELECT COUNT(*) FROM projects) AS projects,
          (SELECT COUNT(*) FROM users) AS users,
          (SELECT COUNT(*) FROM work_packages) AS work_packages;
    """
    if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
        command = [
            "docker", "exec", "-i", container_name, "bash", "-c",
            f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F '|' <<'SQL'\n{sql}\nSQL\nEOSQL",
        ]
    else:
        command = [
            "docker", "exec", "-i", container_name, "bash", "-lc",
            (
                f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                f"-h {OPENPROJECT_DB_HOST} -t -A -F '|' -c \"{sql}\""
            ),
        ]
    try:
        result = subprocess.run(command, check=True, text=True, capture_output=True, timeout=30)
        line = result.stdout.strip().splitlines()[-1] if result.stdout else ""
        parts = line.split("|") if line else []
        projects = int(parts[0]) if len(parts) >= 1 and parts[0].isdigit() else 0
        users = int(parts[1]) if len(parts) >= 2 and parts[1].isdigit() else 0
        work_packages = int(parts[2]) if len(parts) >= 3 and parts[2].isdigit() else 0
        return {"projects": projects, "users": users, "work_packages": work_packages}
    except Exception as exc:
        if ui:
            ui.info(f"  Failed DB counts, will fallback to API: {exc}")
        return {}

def _fetch_counts_from_api(client: "OpenProjectClient") -> Dict[str, int]:
    projects = len(client.list_projects())
    users = len(client.list_users())
    # List all work packages using the generic collection
    try:
        work_packages = sum(1 for _ in client._get_collection("/work_packages"))
    except Exception:
        work_packages = 0
    return {"projects": projects, "users": users, "work_packages": work_packages}

def get_openproject_counts(client: "OpenProjectClient", ui: Optional[ConsoleUI] = None) -> Dict[str, int]:
    # Prefer DB on v13
    if client.version == "v13":
        counts = _fetch_counts_from_db(getattr(client, "base_url", None), ui)
        if counts:
            return counts
    return _fetch_counts_from_api(client)


def _get_openproject_container(base_url: Optional[str] = None) -> str:
    """Detect the correct OpenProject container name based on URL or running containers.
    
    Args:
        base_url: Optional OpenProject base URL (e.g., http://localhost:8081)
    
    Returns:
        Container name to use for database operations
    """
    # If base_url is provided and contains port 8081, use v13 container
    if base_url and "8081" in base_url:
        return "project-management-agent-openproject_v13-1"
    
    # Try to detect from running containers
    try:
        result = subprocess.run(
            ["docker", "ps", "--format", "{{.Names}}"],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            containers = result.stdout.strip().split("\n")
            # Prefer v13 container if it exists
            v13_container = next(
                (c for c in containers if "openproject_v13" in c),
                None
            )
            if v13_container:
                return v13_container
            # Fall back to v16 container
            v16_container = next(
                (c for c in containers if "openproject" in c and "db" not in c),
                None
            )
            if v16_container:
                return v16_container
    except Exception:
        pass
    
    # Default fallback
    return OPENPROJECT_CONTAINER


class OpenProjectError(Exception):
    """Raised when the OpenProject API returns an unexpected response."""


def slugify(value: str) -> str:
    """Convert a project name into a lowercase identifier acceptable to
    OpenProject."""
    normalized = unicodedata.normalize("NFKD", value)
    normalized_ascii = normalized.encode("ascii", "ignore").decode("ascii")
    pattern = r"[^a-zA-Z0-9]+"
    slug = re.sub(pattern, "-", normalized_ascii).strip("-").lower()
    return slug or "project"


def iso_date(value) -> Optional[str]:
    """Convert a cell value to YYYY-MM-DD if it looks like a date."""
    if value is None:
        return None
    if isinstance(value, dt.date):
        return value.isoformat()
    # Attempt to parse common string formats
    if isinstance(value, str):
        # Try various date formats, including 2-digit year formats
        # Note: strptime requires exact format match, so we need separate
        # formats for single-digit vs double-digit months/days
        formats = [
            "%d/%m/%Y",      # DD/MM/YYYY (e.g., "23/04/2024")
            "%d/%m/%y",      # DD/MM/YY (e.g., "23/04/24")
            "%d/%-m/%y",     # DD/M/YY (e.g., "23/4/24") - Note: %-m is platform-specific
            "%Y-%m-%d",      # YYYY-MM-DD (ISO format)
            "%m/%d/%Y",      # MM/DD/YYYY (US format)
            "%m/%d/%y",      # MM/DD/YY (US format with 2-digit year)
        ]
        # Try standard formats first
        for fmt in formats:
            try:
                parsed = dt.datetime.strptime(value, fmt).date()
                # If year is 2-digit, assume 2000-2099 range
                if parsed.year < 100:
                    parsed = parsed.replace(year=2000 + parsed.year)
                return parsed.isoformat()
            except ValueError:
                continue
        # Try flexible parsing for DD/M/YY format (single digit month/day)
        # This handles cases like "23/4/24" where month/day can be 1-2 digits
        match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", value.strip())
        if match:
            day, month, year = map(int, match.groups())
            if year < 100:
                year = 2000 + year
            try:
                parsed = dt.date(year, month, day)
                return parsed.isoformat()
            except ValueError:
                pass
    return None


def iso_duration_from_hours(hours: Optional[float]) -> Optional[str]:
    """Convert a numeric hour value to ISO 8601 duration (PTxH)."""
    if hours is None:
        return None
    try:
        value = float(hours)
    except (TypeError, ValueError):
        return None
    if value <= 0:
        return None
    # Allow fractional hours (e.g., 1.5 -> PT1H30M)
    total_minutes = int(round(value * 60))
    minutes = total_minutes % 60
    hours_part = total_minutes // 60
    parts = []
    if hours_part:
        parts.append(f"{hours_part}H")
    if minutes:
        parts.append(f"{minutes}M")
    if not parts:
        parts.append("0H")
    return f"PT{''.join(parts)}"


def minutes_from_iso_duration(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    match = re.fullmatch(
        r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+(?:\.\d+)?)S)?",
        value,
    )
    if not match:
        return None
    hours = int(match.group(1)) if match.group(1) else 0
    minutes = int(match.group(2)) if match.group(2) else 0
    seconds = float(match.group(3)) if match.group(3) else 0.0
    total_seconds = hours * 3600 + minutes * 60 + seconds
    return int(round(total_seconds / 60))


@dataclass
class WorkbookRow:
    date_spent: Optional[str]
    user_name: str
    activity: str
    work_package: str
    project_name: str
    units: Optional[float]


@dataclass
class AggregatedTask:
    project_name: str
    assignee_name: str
    raw_title: str
    subject: str
    type_name: str
    issue_id: Optional[str]
    start_date: Optional[str]
    due_date: Optional[str]
    total_hours: float
    logs: List[WorkbookRow]


@dataclass
class AggregationBucket:
    project_name: str
    assignee_name: str
    raw_title: str
    subject: str
    type_name: str
    issue_id: Optional[str]
    total_hours: float = 0.0
    start_date: Optional[dt.date] = None
    logs: List[WorkbookRow] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.logs is None:
            self.logs = []


@dataclass
class ImportedTimeEntry:
    id: int
    spent_on: Optional[str]
    units: Optional[float]
    order_index: int
    user_id: Optional[int] = None  # User who logged the time (for updating logged_by)


@dataclass
class ImportedWorkPackage:
    task: AggregatedTask
    project_id: int
    work_package_id: int
    time_entries: List[ImportedTimeEntry]


@dataclass
class StagedUser:
    full_name: str
    first_name: str
    last_name: str
    normalized_name: str
    openproject_id: Optional[int] = None


@dataclass
class StagedProject:
    name: str
    slug: str
    openproject_id: Optional[int] = None
    allowed_type_ids: set[int] = None  # type: ignore[assignment]
    lock_version: Optional[int] = None

    def __post_init__(self) -> None:
        if self.allowed_type_ids is None:
            self.allowed_type_ids = set()


@dataclass
class StagedTask:
    key: str
    project_name: str
    type_name: str
    subject: str
    assignee_name: str
    issue_id: Optional[str]
    aggregated: AggregatedTask
    openproject_id: Optional[int] = None
    existing_match_details: Optional[str] = None
    resolved_type_id: Optional[int] = None


@dataclass
class WorkbookStaging:
    users: Dict[str, StagedUser]
    projects: Dict[str, StagedProject]
    tasks: Dict[str, StagedTask]


class Colors:
    """ANSI color codes for terminal output."""
    RESET = "\033[0m"
    BOLD = "\033[1m"
    DIM = "\033[2m"
    
    # Text colors
    BLACK = "\033[30m"
    RED = "\033[31m"
    GREEN = "\033[32m"
    YELLOW = "\033[33m"
    BLUE = "\033[34m"
    MAGENTA = "\033[35m"
    CYAN = "\033[36m"
    WHITE = "\033[37m"
    
    # Background colors
    BG_BLACK = "\033[40m"
    BG_RED = "\033[41m"
    BG_GREEN = "\033[42m"
    BG_YELLOW = "\033[43m"
    BG_BLUE = "\033[44m"
    BG_MAGENTA = "\033[45m"
    BG_CYAN = "\033[46m"
    BG_WHITE = "\033[47m"
    
    @staticmethod
    def disable() -> None:
        """Disable colors (for non-terminal output)."""
        Colors.RESET = ""
        Colors.BOLD = ""
        Colors.DIM = ""
        Colors.BLACK = Colors.RED = Colors.GREEN = Colors.YELLOW = ""
        Colors.BLUE = Colors.MAGENTA = Colors.CYAN = Colors.WHITE = ""
        Colors.BG_BLACK = Colors.BG_RED = Colors.BG_GREEN = Colors.BG_YELLOW = ""
        Colors.BG_BLUE = Colors.BG_MAGENTA = Colors.BG_CYAN = Colors.BG_WHITE = ""


class ConsoleUI:
    """Enhanced console UI with beautiful formatting and colors."""
    
    def __init__(self, auto_confirm: bool = False, use_colors: bool = True) -> None:
        self.auto_confirm = auto_confirm
        self.step_index = 0
        self.use_colors = use_colors and sys.stdout.isatty()
        if not self.use_colors:
            Colors.disable()
        # Step explanations to display once at step start
        self.step_explanations: Dict[str, str] = {
            "Clean caches and logs": "Remove local caches and previous logs to ensure a clean run.",
            "Validate workbook": "Sanity-check the Excel file structure and content before import.",
            "Load workbook data": "Parse Excel rows into normalized staging structures (users, projects, tasks).",
            "Load staging cache": "Load previously saved work package mappings to speed up matching.",
            "Fetch users from OpenProject": "Fetch users from API and map staged users to existing OpenProject users.",
            "Create missing users": "Create users that are referenced in Excel but do not exist in OpenProject.",
            "Fetch projects from OpenProject": "Fetch projects and map staged projects to existing OpenProject projects.",
            "Create missing projects": "Create projects that are referenced in Excel but do not exist in OpenProject.",
            "Fetch work package types": "Fetch type catalog to resolve type IDs required for work package creation.",
            "Inspect project configuration": "Load per-project settings like allowed type IDs and lock versions.",
            "Validate time entry activities": "Ensure time entry activities exist; create missing ones if needed.",
            "Ensure project type permissions": "Enable required work package types per project when missing.",
            "Ensure project memberships": "Add users to projects so they can be assigned and log time.",
            "Check existing work packages": "Match staged tasks to existing work packages (by issue ID or subject).",
            "Create new work packages": "Create any work packages that were not matched to existing ones.",
            "Log time entries": "Create missing time entries per Excel rows, skipping those already present.",
            "Update time entry logged_by field": "Correct the 'logged by' user for time entries via DB batch update.",
            "Adjust activity history": "Align created_at/updated_at for work packages and time entries in journals.",
            "Update project creation dates": "Set project created_at to earliest related work or log date for accuracy.",
            "Verify import": "Compare Excel totals with OpenProject totals and report differences.",
        }
    
    def _colorize(self, text: str, color: str) -> str:
        """Apply color to text if colors are enabled."""
        if self.use_colors:
            return f"{color}{text}{Colors.RESET}"
        return text
    
    def start_step(self, title: str) -> None:
        """Start a new step with beautiful formatting."""
        self.step_index += 1
        stamp = dt.datetime.now().strftime("%H:%M:%S")
        step_num = self._colorize(f"Step {self.step_index}", Colors.CYAN + Colors.BOLD)
        timestamp = self._colorize(f"[{stamp}]", Colors.DIM)
        title_colored = self._colorize(title, Colors.BOLD)
        print(f"\n{timestamp} {step_num}: {title_colored}", flush=True)
        # Print a brief explanation for this step if available
        explanation = self.step_explanations.get(title)
        if explanation:
            self.info(f"  {explanation}", "info")
    
    def info(self, message: str, level: str = "info") -> None:
        """Print an info message with appropriate formatting."""
        icons = {
            "info": "ℹ",
            "success": "✓",
            "warning": "⚠",
            "error": "✗",
            "debug": "🔍",
        }
        colors_map = {
            "info": Colors.BLUE,
            "success": Colors.GREEN,
            "warning": Colors.YELLOW,
            "error": Colors.RED,
            "debug": Colors.DIM,
        }
        icon = icons.get(level, "•")
        color = colors_map.get(level, "")
        icon_colored = self._colorize(icon, color)
        print(f"  {icon_colored} {message}", flush=True)
        # Duplicate detailed line to debug log file if configured
        _write_debug_log(level, message)
    
    def success(self, message: str) -> None:
        """Print a success message."""
        self.info(message, "success")
    
    def warning(self, message: str) -> None:
        """Print a warning message."""
        self.info(message, "warning")
    
    def error(self, message: str) -> None:
        """Print an error message."""
        self.info(message, "error")
    
    def debug(self, message: str) -> None:
        """Print a debug message."""
        self.info(message, "debug")
    
    def progress(self, current: int, total: int, label: str = "") -> None:
        """Show progress with a beautiful progress bar."""
        percent = (current / total) * 100 if total else 100.0
        prefix = f"{label} " if label else ""
        
        # Create a simple progress bar
        bar_width = 30
        filled = int(bar_width * current / total) if total > 0 else bar_width
        bar_colored = (
            self._colorize("█" * filled, Colors.GREEN) +
            self._colorize("░" * (bar_width - filled), Colors.DIM)
        )
        
        percent_str = self._colorize(f"{percent:.1f}%", Colors.CYAN)
        count_str = self._colorize(f"{current}/{total}", Colors.DIM)
        
        print(f"  {prefix}{bar_colored} {percent_str} {count_str}", flush=True)
    
    def complete_step(self, message: str) -> None:
        """Mark a step as complete with success styling."""
        checkmark = self._colorize("✔", Colors.GREEN + Colors.BOLD)
        message_colored = self._colorize(message, Colors.GREEN)
        print(f"  {checkmark} {message_colored}", flush=True)
    
    def skip_step(self, message: str) -> None:
        """Mark a step as skipped with appropriate styling."""
        arrow = self._colorize("↷", Colors.YELLOW)
        message_colored = self._colorize(message, Colors.DIM)
        print(f"  {arrow} {message_colored}", flush=True)
    
    def ask_confirmation(self, prompt: str) -> bool:
        """Ask for user confirmation with styled prompt."""
        arrow = self._colorize("→", Colors.CYAN)
        prompt_colored = self._colorize(prompt, Colors.BOLD)
        
        if self.auto_confirm:
            auto_yes = self._colorize("[auto-yes]", Colors.GREEN + Colors.DIM)
            print(f"  {arrow} {prompt_colored} {auto_yes}", flush=True)
            return True
        
        while True:
            response = input(f"  {arrow} {prompt_colored} {self._colorize('[y/N]', Colors.DIM)}: ").strip().lower()
            if response in {"y", "yes"}:
                return True
            if response in {"", "n", "no"}:
                return False
            self.warning("Please answer 'y' or 'n'.")
    
    def section(self, title: str) -> None:
        """Print a section header."""
        title_colored = self._colorize(title, Colors.BOLD + Colors.CYAN)
        print(f"\n  {title_colored}", flush=True)
        print(f"  {self._colorize('─' * (len(title) + 2), Colors.DIM)}", flush=True)
    
    def table_header(self, *columns: str) -> None:
        """Print a table header."""
        header = "  ".join(columns)
        header_colored = self._colorize(header, Colors.BOLD)
        print(f"  {header_colored}", flush=True)
        separator = "  ".join("-" * len(col) for col in columns)
        separator_colored = self._colorize(separator, Colors.DIM)
        print(f"  {separator_colored}", flush=True)
    
    def table_row(self, *columns: str) -> None:
        """Print a table row."""
        row = "  ".join(columns)
        print(f"  {row}", flush=True)
    
    def progress_summary(
        self,
        existing: int,
        missing: int,
        created: int = 0,
        entity_name: str = "entries",
    ) -> None:
        """Print a progress summary table showing existing, missing, and created counts.
        
        Args:
            existing: Number of entries that already exist and are mapped
            missing: Number of entries that don't exist and need to be created
            created: Number of entries that have been created (default: 0)
            entity_name: Name of the entity type (e.g., "users", "work packages")
        """
        total = existing + missing
        if total == 0:
            return
        
        # Fixed-width columns to keep alignment consistent
        status_w = 18
        count_w = 7
        pct_w = 10
        self.table_header("Status".ljust(status_w), "Count".rjust(count_w), "Percentage".rjust(pct_w))
        
        if existing > 0:
            existing_pct = (existing / total * 100) if total > 0 else 0
            self.table_row(
                self._colorize("✓ Existing/Mapped", Colors.GREEN).ljust(status_w),
                f"{existing:>{count_w},d}",
                f"{existing_pct:>{pct_w}.1f}%"
            )
        
        if missing > 0:
            missing_pct = (missing / total * 100) if total > 0 else 0
            self.table_row(
                self._colorize("→ Need to Create", Colors.YELLOW).ljust(status_w),
                f"{missing:>{count_w},d}",
                f"{missing_pct:>{pct_w}.1f}%"
            )
        
        if created > 0:
            created_pct = (created / total * 100) if total > 0 else 0
            self.table_row(
                self._colorize("✓ Created", Colors.CYAN).ljust(status_w),
                f"{created:>{count_w},d}",
                f"{created_pct:>{pct_w}.1f}%"
            )
        
        self.table_row(
            self._colorize("Total", Colors.BOLD).ljust(status_w),
            f"{total:>{count_w},d}",
            f"{100.0:>{pct_w}.1f}%"
        )
        print()  # Empty line after table


def _format_duration(seconds: float) -> str:
    """Format duration in seconds as H:MM:SS."""
    seconds_int = int(round(seconds))
    hours = seconds_int // 3600
    minutes = (seconds_int % 3600) // 60
    secs = seconds_int % 60
    return f"{hours:d}:{minutes:02d}:{secs:02d}"


def _print_total_duration(ui: ConsoleUI, start_ts: float) -> None:
    """Print the total elapsed time since start_ts."""
    try:
        elapsed = max(0.0, time_module.time() - start_ts)
        hhmmss = _format_duration(elapsed)
        minutes = elapsed / 60.0
        ui.section("⏱ Total Duration")
        ui.info(f"Elapsed: {hhmmss} ({minutes:.1f} min)")
    except Exception:
        # Never fail on reporting duration
        pass


def _try_fix_user_assignment_via_api(
    client: "OpenProjectClient",
    user_id: int,
    project_id: int,
    ui: ConsoleUI,
) -> bool:
    """Try to fix user assignment issues via API.
    
    Attempts to:
    - Activate the user if inactive
    - Unlock the user if locked
    
    Returns True if fix was attempted and might have worked, False otherwise.
    """
    try:
        # Get user details
        resp = client._request("GET", f"/users/{user_id}")
        user_data = resp.json()
        
        status = user_data.get("status", "")
        is_locked = user_data.get("locked", False)
        
        fixed = False
        
        # Try to activate user if inactive
        if status != "active":
            ui.warning(f"  Attempting to activate user (status: {status})...")
            try:
                payload = {"status": "active"}
                client._request("PATCH", f"/users/{user_id}", data=json.dumps(payload))
                ui.info("  ✅ User activated via API")
                fixed = True
            except Exception as e:
                ui.warning(f"  ⚠️  Could not activate user via API: {e}")
        
        # Try to unlock user if locked
        if is_locked:
            ui.warning("  Attempting to unlock user...")
            try:
                payload = {"locked": False}
                client._request("PATCH", f"/users/{user_id}", data=json.dumps(payload))
                ui.info("  ✅ User unlocked via API")
                fixed = True
            except Exception as e:
                ui.warning(f"  ⚠️  Could not unlock user via API: {e}")
        
        return fixed
    except Exception as e:
        ui.warning(f"  ⚠️  Could not check/fix user via API: {e}")
        return False


def _try_fix_user_assignment_via_db(
    user_id: int,
    project_id: int,
    ui: ConsoleUI,
    *,
    base_url: Optional[str] = None,
) -> bool:
    """Try to fix user assignment issues via database.
    
    Attempts to:
    - Activate the user if inactive (status = 1)
    - Unlock the user if locked
    - Ensure user type is correct
    
    Returns True if fix was attempted, False otherwise.
    """
    try:
        ui.warning("  Attempting to fix user via database...")
        
        # Check current user status
        check_sql = f"""
            SELECT id, login, status, type, locked
            FROM users
            WHERE id = {user_id};
        """
        
        # Detect correct container (supports v13 embedded and v16 external DB)
        container_name = _get_openproject_container(base_url)
        if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
            # Embedded DB inside app container (v13): use su - postgres
            command = [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-c",
                f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F '|' <<'SQL'\n{check_sql}\nSQL\nEOSQL",
            ]
        else:
            # External DB (v16+): connect via psql using env credentials
            command = [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-lc",
                (
                    f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                    f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                    f"-h {OPENPROJECT_DB_HOST} -t -A -F '|' -c \"{check_sql}\""
                ),
            ]
        result = subprocess.run(command, capture_output=True, text=True, timeout=20)
        
        if result.returncode != 0:
            ui.warning(f"  ⚠️  Could not query user: {result.stderr}")
            return False
        
        if not result.stdout.strip():
            ui.warning(f"  ⚠️  User {user_id} not found in database")
            return False
        
        parts = result.stdout.strip().split("|")
        if len(parts) < 5:
            ui.warning("  ⚠️  Unexpected user data format")
            return False
        
        current_status = parts[2] if len(parts) > 2 else ""
        current_type = parts[3] if len(parts) > 3 else ""
        current_locked = parts[4] if len(parts) > 4 else ""
        
        fixes = []
        
        # Fix status if not active
        if current_status != "1":
            fixes.append("status = 1")
            ui.info(f"  → Will set status to active (currently: {current_status})")
        
        # Fix locked if true
        if current_locked.lower() in ("t", "true", "1"):
            fixes.append("locked = false")
            ui.info("  → Will unlock user")
        
        # Fix type if not User
        if current_type and current_type != "User":
            fixes.append(f"type = 'User'")
            ui.info(f"  → Will set type to User (currently: {current_type})")
        
        if not fixes:
            ui.info("  → User appears to be in correct state")
            return False
        
        # Apply fixes
        update_sql = f"""
            UPDATE users
            SET {', '.join(fixes)}, updated_at = NOW()
            WHERE id = {user_id};
        """
        
        if "container_name" not in locals():
            container_name = _get_openproject_container(base_url)
        if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
            command = [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-c",
                f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -c \"{update_sql}\"\nEOSQL",
            ]
        else:
            command = [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-lc",
                (
                    f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                    f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                    f"-h {OPENPROJECT_DB_HOST} -c \"{update_sql}\""
                ),
            ]
        result = subprocess.run(command, capture_output=True, text=True, timeout=20)
        
        if result.returncode != 0:
            ui.warning(f"  ⚠️  Could not update user: {result.stderr}")
            return False
        
        ui.success("  ✅ User fixed via database")
        return True
        
    except Exception as e:
        ui.warning(f"  ⚠️  Could not fix user via database: {e}")
        return False


def _try_fix_required_custom_fields_via_api(
    client: "OpenProjectClient",
    project_id: int,
    type_id: int,
    error_msg: str,
    ui: ConsoleUI,
) -> bool:
    """Try to fix required custom fields via API.
    
    Attempts to make custom fields optional or set default values.
    Returns True if fix was attempted, False otherwise.
    """
    # OpenProject API doesn't allow changing custom field requirements via API
    # This would need to be done via web UI or database
    return False


def _try_fix_required_custom_fields_via_db(
    project_id: int,
    error_msg: str,
    ui: ConsoleUI,
) -> bool:
    """Try to fix required custom fields via database.
    
    Attempts to:
    - Make custom fields optional (is_required = false)
    - Or set default values
    
    Returns True if fix was attempted, False otherwise.
    """
    try:
        ui.warning("  Attempting to fix required custom fields via database...")
        
        fixes = []
        
        # Check which custom fields need to be fixed
        if "customField15" in error_msg:
            fixes.append(15)
            ui.info("  → Will make customField15 (QC Activity) optional")
        
        if "customField38" in error_msg:
            fixes.append(38)
            ui.info("  → Will make customField38 (Finished Date) optional")
        
        if not fixes:
            ui.warning("  ⚠️  Could not identify which custom fields to fix")
            return False
        
        # Detect the correct container (try v13 first, then default)
        container_name = OPENPROJECT_CONTAINER
        # Try to find v13 container
        result = subprocess.run(
            ["docker", "ps", "--format", "{{.Names}}"],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            containers = result.stdout.strip().split("\n")
            v13_container = next(
                (c for c in containers if "openproject_v13" in c or "openproject_v13" in c),
                None
            )
            if v13_container:
                container_name = v13_container
                ui.debug(f"  Using container: {container_name}")
        
        # Update custom fields to be optional
        # For v13, database might be external, so try both direct psql and via container
        field_ids_str = ",".join(str(fid) for fid in fixes)
        update_sql = f"""
            UPDATE custom_fields
            SET is_required = false
            WHERE id IN ({field_ids_str});
        """
        
        # Try via OpenProject container (embedded DB) or external DB container
        db_container = container_name.replace("openproject_v13", "openproject_db_v13").replace("openproject-1", "openproject_db-1")
        
        # Try OpenProject container first (embedded DB)
        result = subprocess.run(
            [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-lc",
                (
                    f"su - postgres -c 'psql -d {OPENPROJECT_DB_NAME} -c \"{update_sql}\"'"
                    if OPENPROJECT_DB_HOST == "127.0.0.1"
                    else (
                        f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                        f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                        f"-h {OPENPROJECT_DB_HOST} -c \"{update_sql}\""
                    )
                ),
            ],
            capture_output=True,
            text=True,
            timeout=10,
        )
        
        # If that failed, try external DB container
        if result.returncode != 0:
            result = subprocess.run(
                [
                    "docker",
                    "exec",
                    "-i",
                    db_container,
                    "bash",
                    "-lc",
                    (
                        f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                        f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                        f"-c \"{update_sql}\""
                    ),
                ],
                capture_output=True,
                text=True,
                timeout=10,
            )
        
        if result.returncode != 0:
            ui.warning(
                f"  ⚠️  Could not update custom fields: {result.stderr}"
            )
            return False
        
        # Verify the update worked
        verify_sql = f"SELECT id, is_required FROM custom_fields WHERE id IN ({field_ids_str});"
        verify_result = subprocess.run(
            [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-lc",
                (
                    f"su - postgres -c 'psql -d {OPENPROJECT_DB_NAME} -c \"{verify_sql}\"'"
                    if OPENPROJECT_DB_HOST == "127.0.0.1"
                    else (
                        f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                        f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                        f"-h {OPENPROJECT_DB_HOST} -c \"{verify_sql}\""
                    )
                ),
            ],
            capture_output=True,
            text=True,
            timeout=10,
        )
        
        if verify_result.returncode == 0 and "f" in verify_result.stdout:
            ui.success("  ✅ Required custom fields made optional via database")
            return True
        else:
            ui.warning("  ⚠️  Update may not have persisted (OpenProject may need restart)")
            return False
        
    except Exception as e:
        ui.warning(f"  ⚠️  Could not fix custom fields via database: {e}")
        return False


def _run_psql(sql: str, suppress_output: bool = True) -> None:
    """Run SQL statements via psql, optionally suppressing verbose output."""
    container_name = _get_openproject_container()
    # Prefer embedded Postgres invocation for v13 (or when using 127.0.0.1 inside app container)
    if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-c",
            (
                f"su - postgres <<'EOSQL'\n"
                f"psql -v ON_ERROR_STOP=1 -d {OPENPROJECT_DB_NAME} <<'SQL'\n"
                f"{sql}\n"
                f"SQL\n"
                f"EOSQL"
            ),
        ]
    else:
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-lc",
            (
                f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                f"psql -v ON_ERROR_STOP=1 -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                f"-h {OPENPROJECT_DB_HOST} <<'SQL'\n{sql}\nSQL"
            ),
        ]
    result = subprocess.run(
        command,
        check=True,
        text=True,
        capture_output=suppress_output,
    )
    if not suppress_output and result.stdout:
        # Filter out verbose PostgreSQL output like "UPDATE 1", "BEGIN", "COMMIT"
        lines = result.stdout.split('\n')
        filtered = [
            line for line in lines
            if line.strip() and
            line.strip() not in ('BEGIN', 'COMMIT', 'UPDATE 1') and
            not line.strip().startswith('UPDATE 1')
        ]
        if filtered:
            print('\n'.join(filtered))


def _run_sql_statements(
    statements: List[str],
    chunk_size: int = 200,
    ui: Optional[ConsoleUI] = None,
    description: str = "Updating database",
) -> None:
    """Run SQL statements in chunks with progress reporting."""
    if not statements:
        return
    total = len(statements)
    chunks = (total + chunk_size - 1) // chunk_size
    processed = 0
    for i in range(0, total, chunk_size):
        chunk = statements[i:i + chunk_size]
        chunk_num = (i // chunk_size) + 1
        sql = "BEGIN;\n" + "\n".join(chunk) + "\nCOMMIT;\n"
        _run_psql(sql, suppress_output=True)
        processed += len(chunk)
        if ui:
            # Single-line progress update including chunk position
            ui.progress(
                min(processed, total),
                total,
                f"{description} {chunk_num}/{chunks}"
            )


def update_time_entry_logged_by_batch(
    updates: List[Tuple[int, int]],
    ui: Optional[ConsoleUI] = None,
    dry_run: bool = False,
) -> None:
    """
    Update the logged_by_id field for multiple time entries.
    
    Args:
        updates: List of (time_entry_id, user_id) tuples
        ui: Optional ConsoleUI for progress reporting
        dry_run: If True, only report what would be done
    """
    if not updates:
        if ui:
            ui.info("No time entries to update.", "info")
        return
    
    if dry_run:
        if ui:
            ui.info(
                f"[DRY-RUN] Would update logged_by for {len(updates)} time entry(ies)",
                "info"
            )
        return
    
    if ui:
        ui.info(
            f"Updating logged_by for {len(updates)} time entry(ies)...",
            "info"
        )
    
    # Build SQL update statements.
    # Update whichever column represents the "logged by" user in this OP version.
    # Strategy:
    # - If time_entries.logged_by_id exists, set it to user_id
    # - If time_entries.created_by_id exists, set it to user_id (author of the time entry)
    # - Always ensure time_entries.user_id is set to user_id (person who logged the time)
    # - Also update journals.user_id for the corresponding TimeEntry journal rows
    update_statements: List[str] = []
    for entry_id, user_id in updates:
        stmt = f"""
DO $$
BEGIN
  IF EXISTS (
    SELECT 1 FROM information_schema.columns
    WHERE table_name='time_entries' AND column_name='logged_by_id'
  ) THEN
    UPDATE time_entries SET logged_by_id = {user_id} WHERE id = {entry_id};
  END IF;
  IF EXISTS (
    SELECT 1 FROM information_schema.columns
    WHERE table_name='time_entries' AND column_name='created_by_id'
  ) THEN
    UPDATE time_entries SET created_by_id = {user_id} WHERE id = {entry_id};
  END IF;
  -- Ensure the time entry is attributed to the correct user
  IF EXISTS (
    SELECT 1 FROM information_schema.columns
    WHERE table_name='time_entries' AND column_name='user_id'
  ) THEN
    UPDATE time_entries SET user_id = {user_id} WHERE id = {entry_id};
  END IF;
  -- Update journal author for this time entry
  IF EXISTS (
    SELECT 1 FROM information_schema.columns
    WHERE table_name='journals' AND column_name='user_id'
  ) THEN
    UPDATE journals
      SET user_id = {user_id}
      WHERE journable_type='TimeEntry' AND journable_id={entry_id};
  END IF;
END $$;
"""
        update_statements.append(stmt)

    if update_statements:
        _run_sql_statements(
            update_statements,
            chunk_size=50,  # smaller chunks due to DO blocks
            ui=ui,
            description="Updating logged_by/author/user for time entries",
        )
        if ui:
            ui.complete_step(
                f"Updated logged_by/author/user for {len(updates)} time entry(ies)."
            )


def update_time_entry_logged_by_from_excel(
    workbook_path: Path,
    client: OpenProjectClient,
    staging: WorkbookStaging,
    ui: ConsoleUI,
    dry_run: bool = False,
    source: str = "auto",
) -> int:
    """
    Update logged_by field for time entries by matching Excel data with OpenProject entries.
    
    This function:
    1. Builds a map of (wp_id, date, hours, activity_id) -> user_id from Excel
    2. Fetches all time entries from OpenProject
    3. Matches entries and updates logged_by_id for entries that need it
    
    Args:
        workbook_path: Path to the Excel workbook
        client: OpenProject client
        staging: Workbook staging data
        ui: ConsoleUI for progress reporting
        dry_run: If True, only report what would be done
    
    Returns:
        Number of entries updated
    """
    import re
    
    # Load workbook rows
    workbook_rows = parse_workbook(workbook_path)
    
    # Fetch users to build user_id_map
    user_records = client.list_users()
    user_id_map: Dict[str, Optional[int]] = {}
    for user_name, record in user_records.items():
        href = record["_links"]["self"]["href"]
        user_id_map[normalize_person_name(user_name)] = int(href.split("/")[-1])
    
    # Build normalized activity lookup map
    activity_lookup: Dict[str, int] = {}
    try:
        api_activities = client.list_time_entry_activities()
        for normalized_name, activity_data in api_activities.items():
            activity_lookup[normalized_name] = int(activity_data["id"])
    except Exception:
        pass
    
    try:
        db_activities = fetch_time_entry_activities_from_db()
        for raw_name, activity_id in db_activities.items():
            normalized_name = normalize_activity_name(raw_name)
            if normalized_name not in activity_lookup:
                activity_lookup[normalized_name] = activity_id
    except Exception:
        pass
    
    # Build mapping: (work_package_id, spent_on, hours, activity_id) -> user_id from Excel
    excel_time_entry_map: Dict[Tuple[int, str, float, int], int] = {}
    
    for row in workbook_rows:
        if not row.units or row.units <= 0:
            continue
        if not row.date_spent:
            continue
        
        # Find work package ID by matching task name
        row_task_subject = row.work_package.strip()
        # Try to extract subject if it has "Type #ID: Subject" format
        match = re.match(r"^[^#]*#\d+:\s*(.+)$", row_task_subject)
        if match:
            row_task_subject = match.group(1).strip()
        
        staged_task = None
        for task in staging.tasks.values():
            if (task.project_name == row.project_name and
                task.subject == row_task_subject):
                staged_task = task
                break
        
        if not staged_task or not staged_task.openproject_id:
            continue
        
        wp_id = staged_task.openproject_id
        
        # Find activity ID
        normalized_activity = normalize_activity_name(row.activity or "")
        activity_id = activity_lookup.get(normalized_activity)
        if not activity_id:
            continue
        
        # Find user ID
        normalized_user = normalize_person_name(row.user_name or "")
        user_id = user_id_map.get(normalized_user)
        if not user_id:
            continue
        
        # Build key: (wp_id, spent_on, hours, activity_id)
        hours = float(row.units)
        key = (wp_id, row.date_spent, hours, activity_id)
        excel_time_entry_map[key] = user_id
    
    if not excel_time_entry_map:
        ui.info("No time entries found in Excel to match.", "info")
        return 0
    
    # Decide source (db/api/auto) and fetch entries scoped to staging WPs
    if source == "auto":
        use_db = (client.version == "v13")
    else:
        use_db = (source == "db")
    target_wp_ids: Set[int] = {
        int(task.openproject_id)
        for task in staging.tasks.values()
        if task.openproject_id is not None
    }
    ui.info("  Fetching time entries from OpenProject...")
    all_time_entries: List[dict] = []
    if use_db:
        total_wps = len(target_wp_ids)
        if total_wps:
            ui.progress(0, total_wps, "Fetching time entries (DB)")
        interval = max(1, total_wps // 100) if total_wps > 0 else None
        for idx_wp, wp_id in enumerate(sorted(target_wp_ids), start=1):
            try:
                entries = fetch_time_entries_for_work_package_from_db(wp_id, ui=None, base_url=client.base_url if hasattr(client, "base_url") else None)
                for e in entries:
                    if "_links" not in e:
                        e["_links"] = {}
                    e["_links"].setdefault("workPackage", {"href": f"/api/v3/work_packages/{wp_id}"})
                all_time_entries.extend(entries)
            except Exception as exc:
                _write_debug_log("debug", f"[logged_by] DB fetch failed for WP {wp_id}: {exc}")
            if interval and (idx_wp % interval == 0 or idx_wp == total_wps):
                ui.progress(idx_wp, total_wps, "Fetching time entries (DB)")
        if not all_time_entries:
            ui.info("  [fallback] No DB entries found; falling back to API listing")
            use_db = False
    if not use_db:
        page_size = 200
        offset = 1
        seen_count = 0
        total_entries_api: Optional[int] = None
        ui.progress(0, 1, "Fetching time entries (API)")
        while True:
            params = {"offset": offset, "pageSize": page_size}
            resp = client._request("GET", "/time_entries", params=params)
            payload = resp.json()
            if total_entries_api is None:
                try:
                    total_entries_api = int(payload.get("total")) if payload.get("total") is not None else None
                except Exception:
                    total_entries_api = None
            elements = payload.get("_embedded", {}).get("elements", []) or []
            if target_wp_ids:
                for e in elements:
                    wp_href = (e.get("_links", {}).get("workPackage", {}) or {}).get("href")
                    if not wp_href:
                        continue
                    try:
                        wp_id = int(wp_href.split("/")[-1])
                    except Exception:
                        continue
                    if wp_id in target_wp_ids:
                        all_time_entries.append(e)
            else:
                all_time_entries.extend(elements)
            seen_count += len(elements)
            if total_entries_api and total_entries_api > 0:
                ui.progress(min(seen_count, total_entries_api), total_entries_api, "Fetching time entries (API)")
            else:
                ui.progress(min(offset, 100), 100, "Fetching time entries (API)")
            next_link = payload.get("_links", {}).get("nextByOffset")
            if not elements or not next_link:
                break
            offset += 1
    total_entries = len(all_time_entries)
    ui.info(f"  Found {total_entries:,} time entry(ies) to check")
    
    # Match OpenProject time entries with Excel rows
    ui.info("  Matching time entries with Excel data...")
    logged_by_updates: List[Tuple[int, int]] = []
    
    # Show progress while matching
    if total_entries:
        ui.progress(0, total_entries, "Matching time entries")
    progress_interval = max(1, total_entries // 100) if total_entries > 0 else None
    for idx, entry in enumerate(all_time_entries, start=1):
        if progress_interval and (idx % progress_interval == 0 or idx == total_entries):
            ui.progress(idx, total_entries, "Matching time entries")
        wp_id = entry.get("_links", {}).get("workPackage", {}).get("href", "")
        if not wp_id:
            continue
        try:
            wp_id_int = int(wp_id.split("/")[-1])
        except (ValueError, AttributeError):
            continue
        
        spent_on = entry.get("spentOn", "")
        hours_val = entry.get("hours", "")
        if not spent_on or not hours:
            continue
        
        # Parse hours (ISO 8601 duration to float)
        hours_float: Optional[float] = None
        try:
            if isinstance(hours_val, (int, float)):
                hours_float = float(hours_val)
            elif isinstance(hours_val, str):
                # Try ISO-8601 first
                mins = minutes_from_iso_duration(hours_val)
                if mins is not None:
                    hours_float = mins / 60.0
                else:
                    # Fallback: parse numeric string
                    hours_float = float(hours_val)
        except Exception:
            hours_float = None
        if hours_float is None:
            continue
        
        activity_link = entry.get("_links", {}).get("activity", {}).get("href", "")
        if not activity_link:
            continue
        try:
            activity_id = int(activity_link.split("/")[-1])
        except (ValueError, AttributeError):
            continue
        
        # Build key and look up in Excel map
        key = (wp_id_int, spent_on, hours_float, activity_id)
        excel_user_id = excel_time_entry_map.get(key)
        
        if excel_user_id:
            entry_id = int(entry["id"])
            current_logged_by = entry.get("_links", {}).get("loggedBy", {}).get("href", "")
            
            # Check if update is needed
            needs_update = False
            
            if current_logged_by:
                try:
                    current_user_id = int(current_logged_by.split("/")[-1])
                    if current_user_id != excel_user_id:
                        needs_update = True
                except (ValueError, AttributeError):
                    # Can't extract current user ID, assume needs update
                    needs_update = True
            else:
                # No loggedBy link in API response, update via database
                needs_update = True
            
            if needs_update:
                logged_by_updates.append((entry_id, excel_user_id))
    
    if logged_by_updates:
        ui.info(f"  Found {len(logged_by_updates):,} time entry(ies) that need logged_by update")
    
    # Update logged_by field
    if logged_by_updates:
        update_time_entry_logged_by_batch(
            updates=logged_by_updates,
            ui=ui,
            dry_run=dry_run,
        )
        return len(logged_by_updates)
    else:
        ui.info("No time entries need logged_by update.", "info")
        return 0


def fetch_time_entry_activities_from_db(base_url: Optional[str] = None) -> Dict[str, int]:
    container_name = _get_openproject_container(base_url)
    sql = (
        "SELECT id, name FROM enumerations "
        "WHERE type = 'TimeEntryActivity' AND active = TRUE;"
    )
    
    # For v13, database is embedded, use su - postgres
    if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-c",
            f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F ',' <<'SQL'\n{sql}\nSQL\nEOSQL",
        ]
    else:
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-lc",
            (
                f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                f"-h {OPENPROJECT_DB_HOST} -t -A -F ',' -c \"{sql}\""
            ),
        ]
    
    try:
        result = subprocess.run(
            command,
            check=True,
            text=True,
            capture_output=True,
            timeout=30,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired):
        return {}
    activities: Dict[str, int] = {}
    for line in result.stdout.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        try:
            id_str, name = stripped.split(",", 1)
            activities[name] = int(id_str)
        except ValueError:
            continue
    return activities


def create_time_entry_activities_in_db(names: List[str], base_url: Optional[str] = None) -> None:
    if not names:
        return
    container_name = _get_openproject_container(base_url)
    for index, raw_name in enumerate(names):
        sanitized_name = raw_name.replace("'", "''")
        sql = (
            "WITH next_pos AS ("
            "    SELECT COALESCE(MAX(position), 0) + 1 AS pos "
            "    FROM enumerations "
            "    WHERE type = 'TimeEntryActivity'"
            "), inserted AS ("
            "    INSERT INTO enumerations ("
            "        name, type, active, is_default, position, "
            "        created_at, updated_at"
            "    ) "
            "    SELECT "
            f"        '{sanitized_name}', 'TimeEntryActivity', TRUE, FALSE, pos, "
            "        NOW(), NOW() "
            "    FROM next_pos "
            "    WHERE NOT EXISTS ("
            "        SELECT 1 FROM enumerations "
            "        WHERE type = 'TimeEntryActivity' "
            f"          AND LOWER(name) = LOWER('{sanitized_name}')"
            "    ) "
            "    RETURNING id"
            ") "
            "SELECT id FROM inserted;"
        )
        # For v13, database is embedded, use su - postgres
        if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
            # Use heredoc to avoid quote escaping issues
            command = [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-c",
                f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F ',' <<'SQL'\n{sql}\nSQL\nEOSQL",
            ]
        else:
            command = [
                "docker",
                "exec",
                "-i",
                container_name,
                "bash",
                "-lc",
                (
                    f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                    f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                    f"-h {OPENPROJECT_DB_HOST} -t -A -F ',' -c \"{sql}\""
                ),
            ]
        try:
            result = subprocess.run(
                command,
                check=True,
                text=True,
                capture_output=True,
                timeout=30,
            )
        except subprocess.CalledProcessError as exc:
            raise OpenProjectError(
                f"Failed to insert time entry activity '{raw_name}' via database: {exc}"
            ) from exc
        except subprocess.TimeoutExpired:
            raise OpenProjectError(
                f"Timeout inserting time entry activity '{raw_name}' via database"
            )


def fetch_time_entries_for_work_package_from_db(
    work_package_id: int,
    ui: Optional[ConsoleUI] = None,
    base_url: Optional[str] = None,
) -> List[dict]:
    """Fetch time entries for a work package from database in API-like format.
    
    Returns a list of dicts with the same structure as API responses for matching.
    """
    container_name = _get_openproject_container(base_url)
    # On v13 embedded DB, time_entries are linked by work_package_id only.
    # Some versions may have entity_type/entity_id, but v13 lacks te.entity_id.
    if CURRENT_OP_VERSION == "v13":
        where_clause = f"te.work_package_id = {int(work_package_id)}"
    else:
        # For v16 and newer schemas that may include entity_type/entity_id
        where_clause = (
            f"(te.work_package_id = {int(work_package_id)} "
            f"OR (COALESCE(te.entity_type,'') = 'WorkPackage' AND COALESCE(te.entity_id,0) = {int(work_package_id)}))"
        )
    sql = (
        "SELECT "
        "    te.id, "
        "    te.spent_on, "
        "    te.hours, "
        "    te.user_id, "
        "    te.activity_id "
        "FROM time_entries te "
        f"WHERE {where_clause} "
        "ORDER BY te.spent_on ASC NULLS LAST, te.id ASC;"
    )
    
    # For v13, database is embedded, use su - postgres
    if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-c",
            f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F '|' <<'SQL'\n{sql}\nSQL\nEOSQL",
        ]
    else:
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-lc",
            (
                f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                f"-h {OPENPROJECT_DB_HOST} -t -A -F '|' -c \"{sql}\""
            ),
        ]
    
    try:
        result = subprocess.run(
            command,
            check=True,
            text=True,
            capture_output=True,
            timeout=30,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        if ui:
            ui.debug(
                f"    [history] Failed to query time entries for WP {work_package_id}: {exc}"
            )
        return []
    
    entries: List[dict] = []
    for line in result.stdout.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        parts = stripped.split("|")
        if len(parts) < 5:
            continue
        try:
            entry_id = int(parts[0])
            spent_on = parts[1] if parts[1] else None
            hours = parts[2] if parts[2] else None
            user_id = int(parts[3]) if parts[3] and parts[3].isdigit() else None
            activity_id = int(parts[4]) if parts[4] and parts[4].isdigit() else None
        except (ValueError, IndexError):
            continue
        
        # Convert to API-like format for matching
        entry = {
            "id": entry_id,
            "spentOn": spent_on,
            "hours": hours,
            "_links": {
                "user": {"href": f"{API_PREFIX}/users/{user_id}"} if user_id else {},
                "activity": {"href": f"{API_PREFIX}/time_entries/activities/{activity_id}"} if activity_id else {},
            }
        }
        entries.append(entry)
    
    return entries


def create_type_in_db(name: str) -> int:
    """Create a work package type directly in the OpenProject database.

    This is required because the public API does not expose a `POST /types`
    endpoint. When invoked, the function inserts a new row into `types` and
    returns the generated id.
    """
    sanitized_name = name.replace("'", "''")
    sql = (
        "WITH next_pos AS ("
        "    SELECT COALESCE(MAX(position), 0) + 1 AS pos FROM types"
        "), inserted AS ("
        "    INSERT INTO types ("
        "        name, position, is_in_roadmap, is_milestone, is_default, "
        "        color_id, created_at, updated_at, is_standard, attribute_groups, "
        "        description, patterns, pdf_export_templates_config"
        "    ) "
        "    SELECT "
        f"        '{sanitized_name}', pos, TRUE, FALSE, FALSE, NULL, "
        "        NOW(), NOW(), FALSE, NULL, NULL, NULL, '{}'::jsonb "
        "    FROM next_pos "
        "    RETURNING id"
        ") "
        "SELECT id FROM inserted;"
    )
    command = [
        "docker",
        "exec",
        "-i",
        OPENPROJECT_CONTAINER,
        "bash",
        "-lc",
        (
            f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
            f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
            f"-h {OPENPROJECT_DB_HOST} -t -A -F ',' -c \"{sql}\""
        ),
    ]
    try:
        result = subprocess.run(
            command,
            check=True,
            text=True,
            capture_output=True,
        )
    except subprocess.CalledProcessError as exc:
        raise OpenProjectError(
            f"Failed to insert work package type '{name}' via database: {exc}"
        ) from exc
    output = result.stdout.strip()
    if not output:
        raise OpenProjectError(
            f"Creating work package type '{name}' via database returned no id."
        )
    try:
        return int(output.split(",")[0])
    except ValueError as exc:
        raise OpenProjectError(
            f"Creating work package type '{name}' via database returned "
            f"an invalid id: {output!r}"
        ) from exc


def assign_type_to_project_in_db(project_id: int, type_id: int) -> None:
    project_id_int = int(project_id)
    type_id_int = int(type_id)
    sql = (
        "INSERT INTO projects_types (project_id, type_id) "
        f"SELECT {project_id_int}, {type_id_int} "
        "WHERE NOT EXISTS ("
        "    SELECT 1 FROM projects_types "
        f"    WHERE project_id = {project_id_int} "
        f"      AND type_id = {type_id_int}"
        ");"
    )
    try:
        _run_psql(sql)
    except (subprocess.CalledProcessError, ValueError) as exc:
        raise OpenProjectError(
            "Failed to assign type {type_id} to project {project_id} via "
            "database: {error}".format(
                type_id=type_id_int,
                project_id=project_id_int,
                error=exc,
            )
        ) from exc


def build_basic_auth_header(token: str) -> str:
    """Return HTTP Basic auth header for an OpenProject API token.

    Accepts either a raw OpenProject token or an already base64-encoded value.
    """
    token = token.strip()
    if not token:
        raise ValueError("API token is empty.")

    # Detect if token already looks like base64(apikey:token)
    if ":" not in token and all(c.isalnum() or c in "+/=" for c in token):
        try:
            decoded = base64.b64decode(token, validate=True).decode()
        except (ValueError, UnicodeDecodeError):
            decoded = ""
        if decoded.startswith("apikey:"):
            return f"Basic {token}"

    raw = f"apikey:{token}".encode()
    encoded = base64.b64encode(raw).decode()
    return f"Basic {encoded}"


class OpenProjectClient:
    def __init__(
        self,
        base_url: str,
        token: str,
        session: Optional[requests.Session] = None,
        version: Optional[str] = None,
    ) -> None:
        self.base_url = base_url.rstrip("/")
        self.session = session or requests.Session()
        
        # Step 1: Detect version FIRST (before authentication)
        # This allows us to handle version-specific authentication
        self.version = version or self._detect_version_without_auth()
        # Expose detected version globally for helpers that don't receive client
        global CURRENT_OP_VERSION
        CURRENT_OP_VERSION = self.version
        
        # Step 2: Set up authentication based on version
        auth_header = self._build_auth_header(token)
        self.session.headers.update(
            {
                "Authorization": auth_header,
                "Content-Type": "application/json",
                "Accept": "application/json",
            }
        )
        
        # Step 3: Verify authentication works
        self._verify_authentication()
    
    def _detect_version_without_auth(self) -> str:
        """Detect OpenProject version without authentication.
        
        Returns 'v13' or 'v16' based on endpoint availability.
        Uses public endpoints or version-specific behavior.
        """
        try:
            # Method 1: Try to get version from status endpoint (if available)
            # Some versions expose version info without auth
            try:
                resp = requests.get(
                    f"{self.base_url}{API_PREFIX}/status",
                    timeout=5
                )
                if resp.status_code == 200:
                    data = resp.json()
                    # Check if version info is in response
                    if "version" in str(data).lower():
                        version_str = str(data).lower()
                        if "13" in version_str or "v13" in version_str:
                            return "v13"
                        if "16" in version_str or "v16" in version_str:
                            return "v16"
            except Exception:
                pass
            
            # Method 2: Check v16-specific endpoint behavior
            # v16 has /time_entries/activities endpoint, v13 doesn't
            # We can check the error response format even without auth
            try:
                resp = requests.get(
                    f"{self.base_url}{API_PREFIX}/time_entries/activities",
                    timeout=5
                )
                # v16 returns 401 (Unauthenticated) with proper error format
                # v13 returns 400 (BadRequest) or 404
                if resp.status_code == 401:
                    # Check error format - v16 has specific error structure
                    try:
                        error_data = resp.json()
                        if "Unauthenticated" in str(error_data):
                            return "v16"  # v16 returns 401 with Unauthenticated
                    except:
                        pass
                elif resp.status_code in (400, 404):
                    return "v13"  # v13 returns 400/404 for non-existent endpoints
            except Exception:
                pass
            
            # Method 3: Default based on common behavior
            # If we can't determine, default to v13 (safer for older versions)
            return "v13"
        except requests.RequestException:
            # If all detection fails, default to v13
            return "v13"
    
    def _build_auth_header(self, token: str) -> str:
        """Build authentication header based on OpenProject version.
        
        Both v13 and v16 use the same Basic Auth format (apikey:token),
        but we handle token validation differently.
        """
        # Both versions use the same authentication method
        return build_basic_auth_header(token)
    
    def _verify_authentication(self) -> None:
        """Verify that authentication is working.
        
        For v13, provides helpful error messages if token authentication fails
        (which might indicate old plain-text tokens from restored database).
        """
        try:
            # Try a simple authenticated request
            resp = self.session.get(
                f"{self.base_url}{API_PREFIX}/users?pageSize=1",
                timeout=5
            )
            
            if resp.status_code == 401:
                error_msg = "Authentication failed"
                try:
                    error_data = resp.json()
                    error_msg = error_data.get("message", error_msg)
                except:
                    pass
                
                if self.version == "v13":
                    raise OpenProjectError(
                        f"{error_msg}. "
                        "For OpenProject v13, if you restored from a database backup, "
                        "old tokens may be stored as plain text instead of hashed. "
                        "Please generate a new API token via the web UI or Rails console."
                    )
                else:
                    raise OpenProjectError(
                        f"{error_msg}. Please check your API token."
                    )
            
            # If we get here, authentication is working
        except OpenProjectError:
            raise
        except requests.RequestException as e:
            # Network errors are handled separately
            raise OpenProjectError(f"Failed to verify authentication: {e}")

    def _request(self, method: str, path: str, **kwargs) -> requests.Response:
        url = f"{self.base_url}{API_PREFIX}{path}"
        # Basic retry for transient connection drops during long operations
        max_attempts = 5
        backoff = 1.0
        last_exc: Optional[Exception] = None
        for attempt in range(1, max_attempts + 1):
            try:
                response = self.session.request(method, url, **kwargs)
                break
            except Exception as e:  # requests exceptions
                last_exc = e
                print(f"  Retry {attempt}/{max_attempts} after connection error: {e}", flush=True)
                time_sleep = min(10.0, backoff)
                try:
                    import time as _t
                    _t.sleep(time_sleep)
                except Exception:
                    pass
                backoff *= 2
        else:
            raise OpenProjectError(f"Failed to call {method} {url}: {last_exc}")
        if response.status_code >= 400:
            message = (
                f"{method} {url} failed with status "
                f"{response.status_code}: {response.text}"
            )
            raise OpenProjectError(message)
        return response

    def _get_collection(self, path: str) -> Iterable[dict]:
        offset = 1
        page_size = 100
        seen_count = 0
        total_count: Optional[int] = None
        while True:
            params = {"offset": offset, "pageSize": page_size}
            resp = self._request("GET", path, params=params)
            payload = resp.json()
            # Determine total for percentage logging (if provided by API)
            if path == "/time_entries" and total_count is None:
                try:
                    total_count = int(payload.get("total")) if payload.get("total") is not None else None
                except Exception:
                    total_count = None
            elements = payload.get("_embedded", {}).get("elements", [])
            # Stream progress for long-running time entries fetches
            if path == "/time_entries":
                if total_count and total_count > 0:
                    seen_count += len(elements or [])
                    pct = max(0.0, min(100.0, (seen_count / float(total_count)) * 100.0))
                    print(f"  Fetching time entries: {pct:.1f}% (page {offset})", flush=True)
                    _write_debug_log("debug", f"fetch_time_entries progress {pct:.1f}% (page {offset}, batch={len(elements or [])})")
                else:
                    print(f"  Fetching time entries: page {offset}", flush=True)
                    _write_debug_log("debug", f"fetch_time_entries page {offset} (batch={len(elements or [])})")
            for elem in elements:
                yield elem
            if not elements:
                break
            next_link = payload.get("_links", {}).get("nextByOffset")
            if not next_link:
                break
            offset += 1

    def list_users(self) -> Dict[str, dict]:
        """Return a mapping of user display name -> API record."""
        users = {}
        for user in self._get_collection("/users"):
            name = user.get("name")
            if name:
                users[name] = user
        return users

    def list_types(self) -> Dict[str, dict]:
        types = {}
        for type_ in self._get_collection("/types"):
            name = type_.get("name")
            if name:
                types[name] = type_
        return types

    def create_type(
        self,
        name: str,
        *,
        color: str = "#cccccc",
        is_milestone: bool = False,
        is_default: bool = False,
        position: int = 0,
        description: str = "",
    ) -> dict:
        payload: Dict[str, object] = {
            "name": name,
            "color": color,
            "isMilestone": is_milestone,
            "isDefault": is_default,
            "position": position,
            "description": {"format": "plain", "raw": description},
        }
        resp = self._request("POST", "/types", data=json.dumps(payload))
        return resp.json()

    def list_projects(self) -> Dict[str, dict]:
        projects = {}
        for project in self._get_collection("/projects"):
            name = project.get("name")
            if name:
                projects[name] = project
        return projects

    def get_project(self, project_id: int) -> dict:
        resp = self._request("GET", f"/projects/{project_id}")
        return resp.json()

    def update_project_types(
        self,
        project_id: int,
        type_ids: Sequence[int],
        lock_version: Optional[int] = None,
    ) -> dict:
        if lock_version is None:
            project = self.get_project(project_id)
            lock_version = project.get("lockVersion")
        payload: Dict[str, object] = {
            "_links": {
                "types": [
                    {"href": f"{API_PREFIX}/types/{type_id}"}
                    for type_id in type_ids
                ]
            }
        }
        if lock_version is not None:
            payload["lockVersion"] = lock_version
        resp = self._request(
            "PATCH",
            f"/projects/{project_id}",
            data=json.dumps(payload),
        )
        return resp.json()

    def create_project(self, name: str) -> dict:
        identifier = slugify(name)
        payload = {
            "name": name,
            "identifier": identifier,
        }
        resp = self._request("POST", "/projects", data=json.dumps(payload))
        return resp.json()

    def enable_project_module(self, project_id: int, module_name: str) -> None:
        project = self.get_project(project_id)
        modules = project.get("modules", [])
        if module_name in modules:
            return
        payload = {
            "modules": sorted(set(modules + [module_name])),
            "lockVersion": project.get("lockVersion"),
        }
        self._request(
            "PATCH",
            f"/projects/{project_id}",
            data=json.dumps(payload),
        )

    def list_roles(self) -> Dict[str, dict]:
        roles = {}
        for role in self._get_collection("/roles"):
            name = role.get("name")
            if name:
                roles[name] = role
        return roles

    def list_project_memberships(self, project_id: int) -> List[dict]:
        memberships: List[dict] = []
        filters = json.dumps(
            [
                {
                    "project": {
                        "operator": "=",
                        "values": [str(project_id)],
                    }
                }
            ]
        )
        offset = 1
        page_size = 100
        while True:
            params = {
                "filters": filters,
                "offset": offset,
                "pageSize": page_size,
            }
            resp = self._request("GET", "/memberships", params=params)
            payload = resp.json()
            elements = payload.get("_embedded", {}).get("elements", [])
            memberships.extend(elements)
            if not elements:
                break
            if not payload.get("_links", {}).get("nextByOffset"):
                break
            offset += 1
        return memberships

    def add_project_membership(
        self,
        project_id: int,
        user_id: int,
        role_id: int,
    ) -> dict:
        """Add a user to a project as a member.
        
        Returns the created membership.
        Raises OpenProjectError if the user cannot be assigned.
        """
        payload = {
            "_links": {
                "project": {"href": f"{API_PREFIX}/projects/{project_id}"},
                "principal": {"href": f"{API_PREFIX}/users/{user_id}"},
                "roles": [{"href": f"{API_PREFIX}/roles/{role_id}"}],
            }
        }
        resp = self._request("POST", "/memberships", data=json.dumps(payload))
        return resp.json()

    def create_work_package(
        self,
        project_id: int,
        type_id: int,
        subject: str,
        assignee_id: int,
        start_date: Optional[str] = None,
        due_date: Optional[str] = None,
        estimated_time: Optional[str] = None,
        description: Optional[str] = None,
    ) -> dict:
        links = {
            "type": {"href": f"{API_PREFIX}/types/{type_id}"},
            "project": {"href": f"{API_PREFIX}/projects/{project_id}"},
            "assignee": {"href": f"{API_PREFIX}/users/{assignee_id}"},
        }
        payload = {
            "subject": subject,
            "_links": links,
        }
        if start_date:
            payload["startDate"] = start_date
        if due_date:
            payload["dueDate"] = due_date
        if estimated_time:
            payload["estimatedTime"] = estimated_time
        if description is not None:
            # Explicitly set description (even if empty string) to ensure it's blank
            payload["description"] = {"format": "markdown", "raw": description}
        else:
            # Explicitly set to empty string to ensure no default description is added
            payload["description"] = {"format": "markdown", "raw": ""}
        resp = self._request(
            "POST",
            "/work_packages",
            data=json.dumps(payload),
        )
        return resp.json()

    def get_work_package(self, work_package_id: int) -> Optional[dict]:
        try:
            resp = self._request("GET", f"/work_packages/{work_package_id}")
        except OpenProjectError as exc:
            if "status 404" in str(exc):
                return None
            raise
        return resp.json()

    def search_work_packages(
        self,
        project_id: int,
        subject: str,
    ) -> List[dict]:
        filters = json.dumps(
            [
                {
                    "subject": {
                        "operator": "~",
                        "values": [subject],
                    }
                }
            ]
        )
        params = {"filters": filters}
        resp = self._request(
            "GET",
            f"/projects/{project_id}/work_packages",
            params=params,
        )
        payload = resp.json()
        return payload.get("_embedded", {}).get("elements", [])

    def create_time_entry(
        self,
        project_id: int,
        work_package_id: int,
        user_id: int,
        spent_on: str,
        hours_iso: str,
        comment: Optional[str] = None,
        activity_id: Optional[int] = None,
    ) -> dict:
        links: Dict[str, Dict[str, str]] = {
            "project": {"href": f"{API_PREFIX}/projects/{project_id}"},
            "workPackage": {
                "href": f"{API_PREFIX}/work_packages/{work_package_id}"
            },
            "user": {"href": f"{API_PREFIX}/users/{user_id}"},
        }
        links["entity"] = {
            "href": f"{API_PREFIX}/work_packages/{work_package_id}"
        }
        payload: Dict[str, object] = {
            "hours": hours_iso,
            "spentOn": spent_on,
            "_links": links,
        }
        if activity_id is not None:
            links["activity"] = {
                "href": f"{API_PREFIX}/time_entries/activities/{activity_id}"
            }
        if comment:
            payload["comment"] = {"format": "plain", "raw": comment}
        resp = self._request("POST", "/time_entries", data=json.dumps(payload))
        return resp.json()

    def update_time_entry_logged_by(
        self,
        time_entry_id: int,
        logged_by_user_id: int,
    ) -> dict:
        """Update the 'logged by' field of a time entry to the specified user."""
        payload = {
            "_links": {
                "loggedBy": {"href": f"{API_PREFIX}/users/{logged_by_user_id}"},
            },
        }
        resp = self._request(
            "PATCH",
            f"/time_entries/{time_entry_id}",
            data=json.dumps(payload),
        )
        return resp.json()

    def list_time_entries(self) -> List[dict]:
        return list(self._get_collection("/time_entries"))

    def create_user(
        self,
        login: str,
        first_name: str,
        last_name: str,
        email: str,
        password: str,
        admin: bool = False,
        status: str = "active",
    ) -> dict:
        payload = {
            "login": login,
            "firstName": first_name,
            "lastName": last_name,
            "email": email,
            "password": password,
            "status": status,
            "admin": admin,
        }
        resp = self._request(
            "POST",
            "/users",
            data=json.dumps(payload),
        )
        return resp.json()

    def list_project_types(self, project_id: int) -> List[dict]:
        resp = self._request("GET", f"/projects/{project_id}/types")
        payload = resp.json()
        elements = payload.get("_embedded", {}).get("elements", []) or []
        return elements

    def list_time_entry_activities(self) -> Dict[str, dict]:
        """List time entry activities with version-specific handling.
        
        v13 may not have this endpoint, so we fall back to database query.
        """
        activities: Dict[str, dict] = {}
        try:
            for activity in self._get_collection("/time_entries/activities"):
                name = activity.get("name")
                if name:
                    activities[name] = activity
        except OpenProjectError as e:
            # v13 may not have this endpoint - fall back to database
            if "404" in str(e) or self.version == "v13":
                # Return empty dict - caller should use database fallback
                return {}
            raise
        return activities

    def list_project_type_ids(self, project_id: int) -> set[int]:
        type_ids: set[int] = set()
        for element in self.list_project_types(project_id):
            href = (
                element.get("_links", {})
                .get("self", {})
                .get("href", "")
            )
            if not href:
                continue
            try:
                type_ids.add(int(href.split("/")[-1]))
            except ValueError:
                continue
        return type_ids


def parse_workbook(path: Path) -> List[WorkbookRow]:
    try:
        from openpyxl import load_workbook  # type: ignore[import]
    except ImportError as exc:  # pragma: no cover - import guard
        message = "openpyxl is required to parse the Excel workbook."
        raise SystemExit(message) from exc

    wb = load_workbook(filename=path, data_only=True)
    if "Work Pakages" not in wb.sheetnames:
        raise ValueError("Expected sheet 'Work Pakages' was not found.")
    ws = wb["Work Pakages"]
    rows: List[WorkbookRow] = []
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = [
        "Date (Spent)",
        "User",
        "Activity",
        "Work package",
        "Project",
        "Units",
    ]
    if header[: len(expected)] != expected:
        raise ValueError(f"Unexpected header row: {header}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        date_spent = iso_date(row[0])
        user_name = re.sub(r"\s+", " ", (row[1] or "").strip())
        activity = (row[2] or "").strip()
        work_package = (row[3] or "").strip()
        project_name = (row[4] or "").strip()
        units_value: Optional[float]
        if row[5] is None:
            units_value = None
        else:
            try:
                units_value = float(row[5])
            except (TypeError, ValueError):
                units_value = None
        if not (user_name and activity and work_package and project_name):
            continue
        rows.append(
            WorkbookRow(
                date_spent=date_spent,
                user_name=user_name,
                activity=activity,
                work_package=work_package,
                project_name=project_name,
                units=units_value,
            )
        )
    return rows


@dataclass
class ValidationIssue:
    row_number: int
    issue_type: str
    message: str
    raw_values: Dict[str, Any]


def validate_workbook(path: Path, ui: ConsoleUI) -> List[ValidationIssue]:
    """Validate the Excel workbook and report all invalid rows.

    Returns a list of validation issues found in the workbook.
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import]
    except ImportError as exc:  # pragma: no cover - import guard
        message = "openpyxl is required to validate the Excel workbook."
        raise SystemExit(message) from exc

    issues: List[ValidationIssue] = []
    wb = load_workbook(filename=path, data_only=True)
    if "Work Pakages" not in wb.sheetnames:
        issues.append(
            ValidationIssue(
                row_number=0,
                issue_type="MISSING_SHEET",
                message="Expected sheet 'Work Pakages' was not found.",
                raw_values={},
            )
        )
        return issues

    ws = wb["Work Pakages"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected = [
        "Date (Spent)",
        "User",
        "Activity",
        "Work package",
        "Project",
        "Units",
    ]
    if header[: len(expected)] != expected:
        issues.append(
            ValidationIssue(
                row_number=1,
                issue_type="INVALID_HEADER",
                message=f"Unexpected header row: {header}",
                raw_values={"header": header},
            )
        )
        return issues

    row_num = 1  # Start at 1 (header is row 1)
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not any(row):
            continue  # Skip completely empty rows

        raw_values = {
            "Date (Spent)": row[0],
            "User": row[1],
            "Activity": row[2],
            "Work package": row[3],
            "Project": row[4],
            "Units": row[5] if len(row) > 5 else None,
        }

        # Check required fields
        user_name = re.sub(r"\s+", " ", (row[1] or "").strip()) if len(row) > 1 else ""
        activity = (row[2] or "").strip() if len(row) > 2 else ""
        work_package = (row[3] or "").strip() if len(row) > 3 else ""
        project_name = (row[4] or "").strip() if len(row) > 4 else ""

        if not user_name:
            issues.append(
                ValidationIssue(
                    row_number=row_num,
                    issue_type="MISSING_USER",
                    message="Missing or empty User field",
                    raw_values=raw_values,
                )
            )
        if not activity:
            issues.append(
                ValidationIssue(
                    row_number=row_num,
                    issue_type="MISSING_ACTIVITY",
                    message="Missing or empty Activity field",
                    raw_values=raw_values,
                )
            )
        if not work_package:
            issues.append(
                ValidationIssue(
                    row_number=row_num,
                    issue_type="MISSING_WORK_PACKAGE",
                    message="Missing or empty Work package field",
                    raw_values=raw_values,
                )
            )
        if not project_name:
            issues.append(
                ValidationIssue(
                    row_number=row_num,
                    issue_type="MISSING_PROJECT",
                    message="Missing or empty Project field",
                    raw_values=raw_values,
                )
            )

        # Check date
        date_spent = iso_date(row[0]) if len(row) > 0 and row[0] else None
        if not date_spent:
            issues.append(
                ValidationIssue(
                    row_number=row_num,
                    issue_type="INVALID_DATE",
                    message=f"Invalid or missing Date (Spent): {row[0]}",
                    raw_values=raw_values,
                )
            )

        # Check units
        if len(row) > 5:
            units_value: Optional[float] = None
            if row[5] is not None:
                try:
                    units_value = float(row[5])
                    if units_value < 0:
                        issues.append(
                            ValidationIssue(
                                row_number=row_num,
                                issue_type="NEGATIVE_UNITS",
                                message=f"Negative units value: {row[5]}",
                                raw_values=raw_values,
                            )
                        )
                except (TypeError, ValueError):
                    issues.append(
                        ValidationIssue(
                            row_number=row_num,
                            issue_type="INVALID_UNITS",
                            message=f"Invalid units value (not a number): {row[5]}",
                            raw_values=raw_values,
                        )
                    )
            else:
                issues.append(
                    ValidationIssue(
                        row_number=row_num,
                        issue_type="MISSING_UNITS",
                        message="Missing Units value",
                        raw_values=raw_values,
                    )
                )

    return issues


def report_validation_issues(issues: List[ValidationIssue], ui: ConsoleUI) -> None:
    """Report validation issues to the user."""
    if not issues:
        ui.info("  ✅ No validation issues found in workbook.")
        return

    ui.info(f"  ⚠️  Found {len(issues)} validation issue(s) in workbook:")
    ui.info("")

    # Group issues by type
    issues_by_type: Dict[str, List[ValidationIssue]] = defaultdict(list)
    for issue in issues:
        issues_by_type[issue.issue_type].append(issue)

    for issue_type, type_issues in sorted(issues_by_type.items()):
        ui.info(f"  {issue_type} ({len(type_issues)} issue(s)):")
        for issue in type_issues[:10]:  # Show first 10 of each type
            ui.info(
                f"    Row {issue.row_number}: {issue.message}"
            )
            if issue.raw_values:
                ui.info(f"      Values: {issue.raw_values}")
        if len(type_issues) > 10:
            ui.info(f"    ... and {len(type_issues) - 10} more")
        ui.info("")

    ui.info(
        "  Note: Invalid data will be handled with defaults during import:"
    )
    ui.info("    - Missing date → today's date")
    ui.info("    - Missing/invalid units → 0.0")
    ui.info("    - Missing activity → first available activity")
    ui.info("")


def summarize_missing(
    required: Iterable[str],
    available: Iterable[str],
) -> List[str]:
    available_set = set(available)
    return sorted({item for item in required if item not in available_set})


def parse_user_name(full_name: str) -> Tuple[str, str]:
    cleaned = re.sub(r"\s+", " ", full_name.strip())
    if not cleaned:
        return "User", "Unknown"
    parts = cleaned.split(" ")
    if len(parts) == 1:
        return parts[0], parts[0]
    first_name = parts[0]
    last_name = " ".join(parts[1:])
    return first_name, last_name


def normalize_person_name(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).lower()


def normalize_activity_name(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).lower()


def build_unique_username(full_name: str, existing_usernames: set[str]) -> str:
    """
    Build a unique username in the format 'first.last' (lowercase, ASCII),
    appending a numeric suffix when needed (e.g., 'first.last2').
    """
    cleaned = unicodedata.normalize("NFKD", full_name).encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"\s+", " ", cleaned.strip())
    if not cleaned:
        base = "user"
    else:
        parts = cleaned.split(" ")
        first = parts[0] if parts else "user"
        last = parts[-1] if len(parts) > 1 else parts[0]
        # Keep only alphanumerics
        first = re.sub(r"[^A-Za-z0-9]", "", first)
        last = re.sub(r"[^A-Za-z0-9]", "", last)
        base = f"{first}.{last}".lower() if first and last else (first or last or "user")
    username = base
    counter = 1
    while username in existing_usernames:
        counter += 1
        username = f"{base}{counter}"
    existing_usernames.add(username)
    return username


def clean_work_package_subject(raw: str) -> str:
    text = raw.strip()
    match = re.match(r"^\s*[^:#]+#\d+\s*:\s*(.+)$", text)
    if match:
        return match.group(1).strip()
    match = re.match(r"^\s*[^:]+:\s*(.+)$", text)
    if match:
        return match.group(1).strip()
    return text


def extract_type_from_work_package(raw: str) -> Optional[str]:
    text = raw.strip()
    match = re.match(r"^\s*([A-Za-z][A-Za-z\s&/-]*)\s*#\d+", text)
    if match:
        return match.group(1).strip()
    return None


def extract_issue_id(raw: str) -> Optional[str]:
    text = raw.strip()
    match = re.search(r"#(\d+)", text)
    if match:
        return match.group(1)
    return None


def parse_task_metadata(raw: str) -> Tuple[str, Optional[str], str]:
    type_name = extract_type_from_work_package(raw) or ""
    issue_id = extract_issue_id(raw)
    subject = clean_work_package_subject(raw)
    if not subject:
        subject = raw.strip()
    return type_name or "Task", issue_id, subject


def compute_due_date(
    start: Optional[dt.date],
    total_hours: float,
) -> Optional[dt.date]:
    if start is None:
        return None
    if total_hours <= 0:
        return start
    days_needed = math.ceil(total_hours / 8.0)
    days = max(0, days_needed - 1)
    delta = dt.timedelta(days=days)
    return start + delta


def aggregate_workbook_rows(rows: List[WorkbookRow]) -> List[AggregatedTask]:
    buckets: Dict[Tuple[str, str, str], AggregationBucket] = {}
    for row in rows:
        raw_title = row.work_package.strip()
        key = (
            row.project_name,
            normalize_person_name(row.user_name),
            raw_title,
        )
        bucket = buckets.get(key)
        if bucket is None:
            type_name, issue_id, subject = parse_task_metadata(
                row.work_package
            )
            if not type_name:
                type_name = row.activity or ""
            bucket = AggregationBucket(
                project_name=row.project_name,
                assignee_name=row.user_name,
                raw_title=raw_title,
                subject=subject,
                type_name=type_name,
                issue_id=issue_id,
            )
            buckets[key] = bucket
        if row.units:
            bucket.total_hours += row.units
        if row.date_spent:
            date_dt = dt.date.fromisoformat(row.date_spent)
            if bucket.start_date is None or date_dt < bucket.start_date:
                bucket.start_date = date_dt
        bucket.logs.append(row)

    aggregated: List[AggregatedTask] = []
    for bucket in buckets.values():
        start_dt = bucket.start_date
        due_dt = compute_due_date(start_dt, bucket.total_hours)
        start_str = start_dt.isoformat() if start_dt else None
        due_str = due_dt.isoformat() if due_dt else None
        aggregated.append(
            AggregatedTask(
                project_name=bucket.project_name,
                assignee_name=bucket.assignee_name,
                raw_title=bucket.raw_title,
                subject=bucket.subject,
                type_name=bucket.type_name,
                issue_id=bucket.issue_id,
                start_date=start_str,
                due_date=due_str,
                total_hours=bucket.total_hours,
                logs=bucket.logs,
            )
        )
    return aggregated


def build_staging(aggregated_tasks: List[AggregatedTask]) -> WorkbookStaging:
    users: Dict[str, StagedUser] = {}
    projects: Dict[str, StagedProject] = {}
    tasks: Dict[str, StagedTask] = {}

    for task in aggregated_tasks:
        normalized_name = normalize_person_name(task.assignee_name)
        if normalized_name not in users:
            first, last = parse_user_name(task.assignee_name)
            users[normalized_name] = StagedUser(
                full_name=task.assignee_name,
                first_name=first,
                last_name=last,
                normalized_name=normalized_name,
            )
        if task.project_name not in projects:
            projects[task.project_name] = StagedProject(
                name=task.project_name,
                slug=slugify(task.project_name),
            )
        if task.issue_id:
            key = f"id:{task.issue_id}:{slugify(task.project_name)}"
        else:
            key = "subject:%s:%s" % (
                slugify(task.subject),
                slugify(task.project_name),
            )
        if key not in tasks:
            tasks[key] = StagedTask(
                key=key,
                project_name=task.project_name,
                type_name=task.type_name,
                subject=task.subject,
                assignee_name=task.assignee_name,
                issue_id=task.issue_id,
                aggregated=task,
            )
        else:
            existing = tasks[key]
            existing.aggregated.logs.extend(task.logs)
            existing.aggregated.total_hours += task.total_hours
            start_dates = [
                existing.aggregated.start_date,
                task.start_date,
            ]
            filtered = [d for d in start_dates if d]
            if filtered:
                existing.aggregated.start_date = min(filtered)
            due_dates = [
                existing.aggregated.due_date,
                task.due_date,
            ]
            filtered_due = [d for d in due_dates if d]
            if filtered_due:
                existing.aggregated.due_date = max(filtered_due)

    for staged in tasks.values():
        logs = staged.aggregated.logs
        total = sum((log.units or 0.0) for log in logs if log.units)
        staged.aggregated.total_hours = total
        iso_dates = [
            iso_date(log.date_spent)
            for log in logs
            if log.date_spent
        ]
        filtered_dates = [d for d in iso_dates if d]
        if filtered_dates:
            earliest = min(filtered_dates)
            staged.aggregated.start_date = earliest
            start_dt = dt.date.fromisoformat(earliest)
            due_dt = compute_due_date(start_dt, total)
            staged.aggregated.due_date = due_dt.isoformat() if due_dt else None
        else:
            staged.aggregated.start_date = None
            staged.aggregated.due_date = None

    return WorkbookStaging(users=users, projects=projects, tasks=tasks)


def scan_and_update_all_project_dates(
    client: OpenProjectClient,
    dry_run: bool,
    ui: ConsoleUI,
    limit_project_ids: Optional[Set[int]] = None,
) -> None:
    """Scan projects and update their creation dates based on earliest work package/time entry.
    
    When limit_project_ids is provided, only those projects are scanned/updated.
    """
    target_scope = "selected" if limit_project_ids else "all"
    ui.info(f"  Scanning {target_scope} projects for earliest work package/time entry dates...")
    
    # Get projects
    ui.info("  Fetching project list from OpenProject...")
    all_projects = client.list_projects()
    if not all_projects:
        ui.info("  No projects found.")
        return
    # Optionally limit to the provided set
    if limit_project_ids:
        projects = {
            name: rec
            for name, rec in all_projects.items()
            if isinstance(rec.get("id"), int) and rec["id"] in limit_project_ids
        }
    else:
        projects = all_projects
    if not projects:
        ui.info("  No target projects in scope.")
        return
    ui.info(f"  Found {len(projects)} project(s).")
    
    project_earliest_dates: Dict[int, datetime.date] = {}
    
    # Query database for earliest dates per project - optimized query
    project_ids = [str(proj.get("id")) for proj in projects.values() if proj.get("id")]
    if not project_ids:
        ui.info("  No project IDs found.")
        return
    
    ui.info(f"  Querying earliest dates for {len(project_ids)} project(s)...")
    sql = """
        SELECT 
            p.id as project_id,
            LEAST(
                COALESCE(MIN(wp.start_date), '9999-12-31'::date),
                COALESCE(MIN(te.spent_on), '9999-12-31'::date)
            ) as earliest_date
        FROM projects p
        LEFT JOIN work_packages wp ON wp.project_id = p.id
        LEFT JOIN time_entries te ON te.project_id = p.id
        WHERE p.id IN ({})
        GROUP BY p.id
        HAVING LEAST(
            COALESCE(MIN(wp.start_date), '9999-12-31'::date),
            COALESCE(MIN(te.spent_on), '9999-12-31'::date)
        ) < '9999-12-31'::date;
    """.format(",".join(project_ids))
    
    # Determine correct container and invocation (embedded v13 uses local postgres user)
    container_name = _get_openproject_container(client.base_url if hasattr(client, "base_url") else None)
    if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
        # Use su - postgres with heredoc to avoid quoting issues
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-c",
            f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F ',' <<'SQL'\n{sql}\nSQL\nEOSQL",
        ]
    else:
        command = [
            "docker",
            "exec",
            "-i",
            container_name,
            "bash",
            "-lc",
            (
                f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                f"-h {OPENPROJECT_DB_HOST} -t -A -F ',' -c \"{sql}\""
            ),
        ]
    
    ui.info("  Executing database query (this may take a moment)...")
    try:
        result = subprocess.run(
            command,
            check=True,
            text=True,
            capture_output=True,
            timeout=180,  # allow a bit more time for large datasets
        )
        ui.info("  Processing query results...")
        for line in result.stdout.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            parts = stripped.split(",")
            if len(parts) >= 2:
                try:
                    project_id = int(parts[0])
                    date_str = parts[1].strip()
                    if date_str and date_str != "9999-12-31":
                        earliest_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        project_earliest_dates[project_id] = earliest_date
                except (ValueError, IndexError):
                    continue
    except subprocess.CalledProcessError as exc:
        ui.info(f"  Error querying database: {exc}")
        return
    
    if not project_earliest_dates:
        ui.info("  No dates found for any projects.")
        return
    
    ui.info(f"  Found earliest dates for {len(project_earliest_dates)} project(s).")

    if dry_run:
        ui.info(
            f"[DRY-RUN] Would update creation dates for "
            f"{len(project_earliest_dates)} project(s)."
        )
        return

    if project_earliest_dates:
        ui.info(
            f"  Updating creation dates for {len(project_earliest_dates)} project(s)..."
        )
        # Update each project separately to avoid exclusion constraint conflicts
        updated_count = 0
        total_projects_to_update = len(project_earliest_dates)
        for idx, (project_id, earliest_date) in enumerate(project_earliest_dates.items(), 1):
            project_name = next(
                (p.get("name", f"Project {project_id}") 
                 for p in projects.values() if p.get("id") == project_id),
                f"Project {project_id}"
            )
            ts = datetime.combine(earliest_date, DEFAULT_CREATION_TIME)
            ts = ts.astimezone(timezone.utc)
            ts_str = ts.isoformat()
            
            # Update project and journal separately to handle exclusion constraint
            # First update the project
            project_statements = [
                f"UPDATE projects SET created_at='{ts_str}' WHERE id={project_id};"
            ]
            try:
                _run_sql_statements(
                    project_statements,
                    ui=ui,
                    description=f"Updating project '{project_name}' creation date",
                )
            except Exception as exc:
                ui.info(
                    f"  [{idx}/{len(project_earliest_dates)}] "
                    f"⚠️  Failed to update project '{project_name}' table: {exc}"
                )
                continue
            
            # Then update the journal - update timestamps and validity_period
            # We set validity_period to start from the new (earlier) timestamp
            journal_statements = [
                (
                    "UPDATE journals SET "
                    "    created_at='{ts}', "
                    "    updated_at='{ts}', "
                    "    validity_period=tstzrange('{ts}', COALESCE(upper(validity_period), NULL)) "
                    "WHERE journable_type='Project' "
                    "  AND journable_id={proj_id} "
                    "  AND version=1;"
                ).format(ts=ts_str, proj_id=project_id)
            ]
            try:
                _run_sql_statements(
                    journal_statements,
                    ui=ui,
                    description=f"Updating project '{project_name}' journal",
                )
                updated_count += 1
                # Percentage progress for Step 19
                ui.progress(idx, total_projects_to_update, "Updating project dates")
                # Reduce console noise; write detail only to debug log
                _write_debug_log(
                    "debug",
                    f"[project_dates] {idx}/{len(project_earliest_dates)} Updated '{project_name}' (earliest: {earliest_date.isoformat()})",
                )
            except Exception as exc:
                _write_debug_log(
                    "debug",
                    f"[project_dates] {idx}/{len(project_earliest_dates)} Failed '{project_name}': {exc}",
                )
        ui.info(
            f"  ✔ Successfully updated {updated_count}/{len(project_earliest_dates)} project creation date(s)."
        )


def update_project_creation_dates(
    imported_packages: List[ImportedWorkPackage],
    dry_run: bool,
    ui: ConsoleUI,
) -> None:
    """Update project creation dates to match the earliest work package date."""
    if not imported_packages:
        return

    # Group packages by project_id and find earliest date for each project
    project_earliest_dates: Dict[int, datetime.date] = {}
    
    for pkg in imported_packages:
        project_id = pkg.project_id
        earliest_date: Optional[datetime.date] = None
        
        # Check work package start_date
        if pkg.task.start_date:
            try:
                earliest_date = datetime.strptime(
                    pkg.task.start_date, "%Y-%m-%d"
                ).date()
            except ValueError:
                pass
        
        # Check time entry dates
        for entry in pkg.time_entries:
            if not entry.spent_on:
                continue
            try:
                entry_date = datetime.strptime(
                    entry.spent_on, "%Y-%m-%d"
                ).date()
            except ValueError:
                continue
            if earliest_date is None or entry_date < earliest_date:
                earliest_date = entry_date
        
        if earliest_date is None:
            continue
        
        # Update project's earliest date if this is earlier
        if project_id not in project_earliest_dates:
            project_earliest_dates[project_id] = earliest_date
        elif earliest_date < project_earliest_dates[project_id]:
            project_earliest_dates[project_id] = earliest_date

    if not project_earliest_dates:
        return

    update_statements: List[str] = []
    for project_id, earliest_date in project_earliest_dates.items():
        ts = datetime.combine(earliest_date, DEFAULT_CREATION_TIME)
        ts = ts.astimezone(timezone.utc)
        ts_str = ts.isoformat()
        
        # Update project created_at
        update_statements.append(
            f"UPDATE projects SET created_at='{ts_str}' WHERE id={project_id};"
        )
        
        # Update project journal (version 1 is the initial creation)
        # Handle exclusion constraint by closing old period first, then setting new one
        # We do this in a single statement using a subquery to avoid conflicts
        update_statements.append(
            (
                "WITH old_period AS ("
                "  SELECT validity_period FROM journals "
                "  WHERE journable_type='Project' "
                "    AND journable_id={proj_id} "
                "    AND version=1"
                "  LIMIT 1"
                ")"
                "UPDATE journals SET "
                "    created_at='{ts}', "
                "    updated_at='{ts}', "
                "    validity_period=tstzrange('{ts}', NULL) "
                "WHERE journable_type='Project' "
                "  AND journable_id={proj_id} "
                "  AND version=1;"
            ).format(ts=ts_str, proj_id=project_id)
        )

    if dry_run:
        ui.info(
            f"[DRY-RUN] Would update creation dates for "
            f"{len(project_earliest_dates)} project(s)."
        )
        return

    if update_statements:
        ui.info(
            f"  Updating creation dates for {len(project_earliest_dates)} project(s)..."
        )
        _run_sql_statements(
            update_statements,
            ui=ui,
            description="Updating project creation dates",
        )
        ui.info(
            f"  ✔ Updated {len(project_earliest_dates)} project creation date(s)."
        )


def apply_history_adjustments(
    imported_packages: List[ImportedWorkPackage],
    dry_run: bool,
    ui: Optional[ConsoleUI] = None,
) -> None:
    if not imported_packages:
        return

    creation_statements: List[str] = []
    work_package_latest: Dict[int, datetime] = {}

    for pkg in imported_packages:
        creation_date: Optional[datetime.date] = None
        if pkg.task.start_date:
            try:
                creation_date = datetime.strptime(
                    pkg.task.start_date, "%Y-%m-%d"
                ).date()
            except ValueError:
                creation_date = None
        if creation_date is None:
            for entry in pkg.time_entries:
                if not entry.spent_on:
                    continue
                try:
                    entry_date = datetime.strptime(
                        entry.spent_on, "%Y-%m-%d"
                    ).date()
                except ValueError:
                    continue
                if creation_date is None or entry_date < creation_date:
                    creation_date = entry_date
        if creation_date is None:
            continue
        ts = datetime.combine(creation_date, DEFAULT_CREATION_TIME)
        ts = ts.astimezone(timezone.utc)
        ts_str = ts.isoformat()
        creation_statements.append(
            "UPDATE work_packages SET created_at='{ts}' WHERE id={wp};".format(
                ts=ts_str,
                wp=pkg.work_package_id,
            )
        )
        creation_statements.append(
            (
                "UPDATE journals SET created_at='{ts}', "
                "    updated_at='{ts}', "
                "    validity_period=tstzrange('{ts}', NULL) "
                "WHERE journable_type='WorkPackage' "
                "  AND journable_id={wp} "
                "  AND version=1;"
            ).format(ts=ts_str, wp=pkg.work_package_id)
        )
        work_package_latest[pkg.work_package_id] = ts

    time_entry_statements: List[str] = []
    for pkg in imported_packages:
        counts: Dict[Tuple[int, dt.date], int] = defaultdict(int)
        for entry in sorted(pkg.time_entries, key=lambda x: x.order_index):
            if not entry.spent_on:
                continue
            try:
                date_value = datetime.strptime(
                    entry.spent_on, "%Y-%m-%d").date()
            except ValueError:
                continue
            offset = counts[(pkg.work_package_id, date_value)]
            counts[(pkg.work_package_id, date_value)] += 1
            timestamp = datetime.combine(
                date_value,
                DEFAULT_LOG_TIME,
            ) + timedelta(minutes=offset)
            timestamp = timestamp.astimezone(timezone.utc)
            ts_str = timestamp.isoformat()
            time_entry_statements.append(
                "UPDATE time_entries "
                f"SET created_at='{ts_str}', updated_at='{ts_str}' "
                f"WHERE id={entry.id};"
            )
            time_entry_statements.append(
                "UPDATE journals "
                f"SET created_at='{ts_str}', "
                f"    updated_at='{ts_str}', "
                f"    validity_period=tstzrange('{ts_str}', NULL) "
                "WHERE journable_type='TimeEntry' "
                f"  AND journable_id={entry.id};"
            )
            current_latest = work_package_latest.get(pkg.work_package_id)
            if current_latest is None or timestamp > current_latest:
                work_package_latest[pkg.work_package_id] = timestamp

    updated_at_statements: List[str] = [
        (
            "UPDATE work_packages SET updated_at='{ts}' WHERE id={wp_id};"
        ).format(ts=ts.isoformat(), wp_id=wp_id)
        for wp_id, ts in work_package_latest.items()
    ]

    if dry_run:
        msg = (
            "[DRY-RUN] Would apply {creates} creation updates, "
            "{logs} logwork updates, {updates} updated_at corrections."
        ).format(
            creates=len(creation_statements),
            logs=len(time_entry_statements),
            updates=len(updated_at_statements),
        )
        if ui:
            ui.info(f"  {msg}")
        else:
            print(msg)
        return

    if creation_statements:
        if ui:
            ui.info(f"  Updating work package creation dates ({len(creation_statements)} statement(s))...")
        _run_sql_statements(
            creation_statements,
            ui=ui,
            description="Updating work package creation dates",
        )
    if time_entry_statements:
        if ui:
            ui.info(f"  Updating time entry timestamps ({len(time_entry_statements)} statement(s))...")
        _run_sql_statements(
            time_entry_statements,
            chunk_size=150,
            ui=ui,
            description="Updating time entry timestamps",
        )
    if updated_at_statements:
        if ui:
            ui.info(f"  Updating work package updated_at timestamps ({len(updated_at_statements)} statement(s))...")
        _run_sql_statements(
            updated_at_statements,
            ui=ui,
            description="Updating work package updated_at timestamps",
        )


def ensure_time_entries_for_task(
    staged_task: StagedTask,
    *,
    client: OpenProjectClient,
    activity_id_lookup: Dict[str, int],
    user_id_map: Dict[str, Optional[int]],
    ui: ConsoleUI,
    dry_run: bool,
    project_id: Optional[int],
    existing_entries: List[dict],
) -> Tuple[List[ImportedTimeEntry], int]:
    work_package_id = staged_task.openproject_id
    if work_package_id is None:
        return [], 0
    if project_id is None:
        ui.info(
            "  [logwork] Skipping '%s'; missing project ID."
            % staged_task.subject
        )
        return [], 0
    # Note: We'll resolve user_id per log entry (from log.user_name) instead of using work package assignee
    # This is because each Excel row can have a different user who logged the time
    # Only process logs with units > 0 to match Excel counting logic
    # (Excel only counts rows with units > 0, so we should only create entries for those)
    logs_with_units = [
        (idx, log)
        for idx, log in enumerate(staged_task.aggregated.logs)
        if log.units is not None and log.units > 0
    ]
    if not logs_with_units:
        return [], 0
    existing_counts: Dict[
        Tuple[str, int, int, int], int
    ] = defaultdict(int)
    for entry in existing_entries:
        spent_on = entry.get("spentOn")
        # Skip entries without a valid date - they can't match workbook entries
        if not spent_on:
            continue
        hours_iso = entry.get("hours")
        minutes_value = minutes_from_iso_duration(hours_iso)
        if minutes_value is None:
            minutes_value = 0
        user_href = (
            entry.get("_links", {}).get("user", {}) or {}
        ).get("href")
        activity_href = (
            entry.get("_links", {}).get("activity", {}) or {}
        ).get("href")
        try:
            user_id = int(user_href.split("/")[-1]) if user_href else None
        except (ValueError, AttributeError):
            user_id = None
        try:
            activity_id = (
                int(activity_href.split("/")[-1]) if activity_href else None
            )
        except (ValueError, AttributeError):
            activity_id = None
        # Skip entries without valid user or activity - they can't match
        if user_id is None or activity_id is None:
            continue
        key = (
            str(spent_on),  # Ensure it's a string
            minutes_value,
            user_id,  # Guaranteed to be int
            activity_id,  # Guaranteed to be int
        )
        existing_counts[key] += 1

    created_entries: List[ImportedTimeEntry] = []
    missing_count = 0
    counts_seen: Dict[
        Tuple[str, int, int, int], int
    ] = defaultdict(int)
    # Track skip reasons for summary
    skipped_already_exists = 0

    for order_index, log in logs_with_units:
        # Resolve user_id from log.user_name (the person who logged the time in Excel)
        # This is different from the work package assignee
        normalized_user = normalize_person_name(log.user_name) if log.user_name else ""
        if not normalized_user:
            # Only show critical errors (missing user) - these prevent import
            ui.error(
                f"Row {order_index + 1} for '{staged_task.subject}': Skipping - missing user name"
            )
            continue
        log_user_id = user_id_map.get(normalized_user)
        if log_user_id is None:
            # Only show critical errors (unresolved user) - these prevent import
            ui.error(
                f"Row {order_index + 1} for '{staged_task.subject}': Skipping - unable to resolve user ID for '{log.user_name}'"
            )
            continue
        
        # Use default activity if missing
        normalized_activity = normalize_activity_name(log.activity) if log.activity else ""
        if not normalized_activity:
            # Try to use a default activity or the first available one
            if activity_id_lookup:
                normalized_activity = next(iter(activity_id_lookup.keys()))
                # Data quality issue - silently handled, no console output
            else:
                raise OpenProjectError(
                    "No time entry activities available in OpenProject."
                )
        activity_id = activity_id_lookup.get(normalized_activity)
        if activity_id is None:
            raise OpenProjectError(
                "Time entry activity '%s' is not available in OpenProject."
                % (log.activity or normalized_activity)
            )
        # Handle invalid units - use 0 if invalid (silently handled)
        try:
            units_value = float(log.units) if log.units is not None else 0.0
        except (TypeError, ValueError):
            units_value = 0.0
        # Ensure we have a valid date - use today if missing (silently handled)
        date_spent = log.date_spent
        if not date_spent:
            from datetime import date
            date_spent = date.today().isoformat()
        hours_iso = iso_duration_from_hours(units_value)
        if not hours_iso:
            # Default to PT0H if conversion fails (silently handled)
            hours_iso = "PT0H"
        minutes_value = int(round(units_value * 60))
        key = (
            date_spent,
            minutes_value,
            log_user_id,
            activity_id,
        )
        counts_seen[key] += 1
        existing_count = existing_counts.get(key, 0)
        # Check if entry already exists in OpenProject
        # Note: We allow duplicate entries from Excel (even if identical) to match Excel exactly
        if existing_count > 0:
            skipped_already_exists += 1
            # Silently skip - will be shown in summary if needed
            continue
        if dry_run:
            # Only show dry-run messages if explicitly requested
            ui.info(
                f"Would log {units_value:.2f} hour(s) on {date_spent} for '{staged_task.subject}'",
                "info"
            )
            missing_count += 1
            continue
        response = client.create_time_entry(
            project_id=int(project_id),
            work_package_id=int(work_package_id),
            user_id=int(log_user_id),
            spent_on=date_spent,
            hours_iso=hours_iso,
            comment=None,
            activity_id=activity_id,
        )
        entry_id = int(response["id"])
        
        # Update the "logged by" field will be handled in batch after all entries are created
        # Store the entry_id and user_id for batch update
        
        # Individual time entry creation messages are too verbose - only log to debug file if needed
        # Remove console logging to reduce noise - summary will be shown instead
        created_entries.append(
            ImportedTimeEntry(
                id=entry_id,
                spent_on=date_spent,
                units=units_value,
                order_index=order_index,
                user_id=int(log_user_id),  # Store user_id for batch update of logged_by
            )
        )
        # Track created entry for potential future use (for idempotency on reruns)
        # Note: We don't increment existing_counts here because we want to allow
        # duplicate entries from Excel (even if identical) to match Excel exactly
        existing_entries.append(
            {
                "spentOn": date_spent,
                "hours": hours_iso,
                "_links": {
                    "user": {"href": f"/api/v3/users/{log_user_id}"},
                    "activity": (
                        {"href": f"/api/v3/time_entries/activities/{activity_id}"}
                        if activity_id is not None
                        else {}
                    ),
                },
            }
        )
        missing_count += 1
    
    # Summary for this task - only show if there were skips or errors
    # Otherwise, the progress bar and final summary are sufficient
    # Only show task summary if there were issues (skipped entries or errors)
    if skipped_already_exists > 0:
        total_created = len(created_entries)
        ui.info(
            f"'{staged_task.subject}': Created={total_created}, Skipped={skipped_already_exists}",
            "warning"
        )
    
    return created_entries, missing_count


def _collect_packages_for_history_adjustment(
    staging: WorkbookStaging,
    existing_packages: Dict[int, ImportedWorkPackage],
    ui: ConsoleUI,
    base_url: Optional[str] = None,
) -> List[ImportedWorkPackage]:
    packages_by_id = dict(existing_packages)
    skip_reasons: Dict[str, int] = defaultdict(int)
    for staged_task in staging.tasks.values():
        work_package_id = staged_task.openproject_id
        if work_package_id is None:
            skip_reasons["no work package id"] += 1
            continue
        if work_package_id in packages_by_id:
            skip_reasons["already processed"] += 1
            continue
        staged_project = staging.projects.get(staged_task.project_name)
        if staged_project is None or staged_project.openproject_id is None:
            skip_reasons["missing project mapping"] += 1
            continue
        logs_with_units = [
            (idx, log)
            for idx, log in enumerate(staged_task.aggregated.logs)
            if log.date_spent and log.units
        ]
        if not logs_with_units:
            skip_reasons["no valid workbook logs"] += 1
            continue
        db_entries = fetch_time_entries_for_work_package_from_db(
            int(work_package_id),
            ui=ui,
            base_url=base_url,
        )
        if not db_entries:
            _write_debug_log("debug", f"[history] skip '{staged_task.subject}': no DB time entries")
            skip_reasons["no db time entries"] += 1
            continue
        if len(db_entries) != len(logs_with_units):
            _write_debug_log("debug", f"[history] skip '{staged_task.subject}': workbook logs={len(logs_with_units)} db entries={len(db_entries)}")
            skip_reasons["log/count mismatch"] += 1
            continue
        imported_entries: List[ImportedTimeEntry] = []
        for order_index, ((_, log), (entry_id, spent_on_db)) in enumerate(
            zip(logs_with_units, db_entries)
        ):
            spent_on_value = log.date_spent or spent_on_db
            imported_entries.append(
                ImportedTimeEntry(
                    id=entry_id,
                    spent_on=spent_on_value,
                    units=log.units,
                    order_index=order_index,
                )
            )
        packages_by_id[int(work_package_id)] = ImportedWorkPackage(
            task=staged_task.aggregated,
            project_id=int(staged_project.openproject_id),
            work_package_id=int(work_package_id),
            time_entries=imported_entries,
        )
    if skip_reasons:
        ui.debug("History adjustment skip summary:")
        for reason, count in sorted(skip_reasons.items()):
            ui.debug(f"  • {count} task(s): {reason}")
    return list(packages_by_id.values())


@dataclass
class ImportVerification:
    """Results of import verification."""
    excel_total_hours: float
    excel_total_entries: int
    openproject_total_hours: float
    openproject_total_entries: int
    work_package_count: int
    project_comparison: Dict[str, Dict[str, float]]


def calculate_excel_totals(
    workbook_path: Path,
    staging: WorkbookStaging,
    tracer: Optional[Any] = None,
) -> Tuple[float, int, Dict[str, float]]:
    """Calculate total hours and entry count from Excel workbook.
    
    Returns:
        (total_hours, total_entries, project_hours_dict)
    """
    total_hours = 0.0
    total_entries = 0
    project_hours: Dict[str, float] = defaultdict(float)
    
    # Parse workbook to get all rows with units
    workbook_rows = parse_workbook(workbook_path)
    
    for row in workbook_rows:
        if row.units is not None and row.units > 0:
            total_hours += row.units
            total_entries += 1
            project_hours[row.project_name] += row.units
            # Log to debug tracer
            if tracer:
                tracer.log_excel_row(
                    project_name=row.project_name,
                    task_name=row.work_package,
                    units=row.units,
                    date=str(row.date_spent) if row.date_spent else "",
                )
    
    return total_hours, total_entries, dict(project_hours)


def calculate_openproject_totals(
    client: OpenProjectClient,
    staging: WorkbookStaging,
    imported_packages: List[ImportedWorkPackage],
    ui: ConsoleUI,
    created_time_entry_ids: Optional[Set[int]] = None,
) -> Tuple[float, int, Dict[str, float]]:
    """Calculate total hours and entry count from OpenProject.
    
    Args:
        created_time_entry_ids: Optional set of time entry IDs to filter by.
            If provided, only time entries with IDs in this set will be counted.
            This ensures we only count entries created in the current import run.
    
    Returns:
        (total_hours, total_entries, project_hours_dict)
    """
    # Choose source for verification totals (DB vs API)
    # Fast path for v13/embedded DB: aggregate directly from database to avoid huge API pagination
    try:
        container_name = _get_openproject_container(client.base_url if hasattr(client, "base_url") else None)
    except Exception:
        container_name = _get_openproject_container()
    # Decide based on flag
    source_decision = VERIFICATION_SOURCE or "auto"
    if source_decision == "auto":
        use_db = ("v13" in container_name) and (created_time_entry_ids is None or len(created_time_entry_ids) == 0)
    elif source_decision == "db":
        # Force DB aggregation regardless of created_time_entry_ids presence
        use_db = True
    else:
        use_db = False
    ui.info(f"Using verification source: {'db' if use_db else 'api'}", "info")
    # Unified fetch method banner for Step 20
    if use_db:
        ui.info("Fetching time entries from OpenProject using Database", "info")
    else:
        ui.info("Fetching time entries from OpenProject using API", "info")
    if use_db:
        # Build project id -> name from staging
        project_id_to_name: Dict[int, str] = {}
        for staged_project in staging.projects.values():
            if staged_project.openproject_id:
                project_id_to_name[int(staged_project.openproject_id)] = staged_project.name
        # Collect relevant work package ids
        wp_ids: Set[int] = set()
        for pkg in imported_packages:
            try:
                wp_ids.add(int(pkg.work_package_id))
            except Exception:
                continue
        # If no new imports this run, include all staged work packages
        if not wp_ids:
            for task in staging.tasks.values():
                if task.openproject_id is not None:
                    try:
                        wp_ids.add(int(task.openproject_id))
                    except Exception:
                        continue
        else:
            ui.info("  Fetching time entries from OpenProject using Database...", "info")
            ids_csv = ",".join(str(i) for i in sorted(wp_ids))
            # Aggregate hours and entries by project for the selected work packages
            if CURRENT_OP_VERSION == "v13":
                where_clause = f"te.work_package_id IN ({ids_csv})"
            else:
                where_clause = (
                    f"(te.work_package_id IN ({ids_csv}) "
                    f"OR (COALESCE(te.entity_type,'') = 'WorkPackage' AND COALESCE(te.entity_id,0) IN ({ids_csv})))"
                )
            sql = f"""
                SELECT
                  te.project_id,
                  COALESCE(SUM(te.hours), 0) AS total_hours,
                  COUNT(*) AS total_entries
                FROM time_entries te
                WHERE {where_clause}
                GROUP BY te.project_id
            """
            # Run query via embedded DB
            if "v13" in container_name or OPENPROJECT_DB_HOST == "127.0.0.1":
                command = [
                    "docker",
                    "exec",
                    "-i",
                    container_name,
                    "bash",
                    "-c",
                    f"su - postgres <<'EOSQL'\npsql -d {OPENPROJECT_DB_NAME} -t -A -F '|' <<'SQL'\n{sql}\nSQL\nEOSQL",
                ]
            else:
                command = [
                    "docker",
                    "exec",
                    "-i",
                    container_name,
                    "bash",
                    "-lc",
                    (
                        f"PGPASSWORD={OPENPROJECT_DB_PASSWORD} "
                        f"psql -U {OPENPROJECT_DB_USER} -d {OPENPROJECT_DB_NAME} "
                        f"-h {OPENPROJECT_DB_HOST} -t -A -F '|' -c \"{sql}\""
                    ),
                ]
            try:
                result = subprocess.run(
                    command,
                    check=True,
                    text=True,
                    capture_output=True,
                    timeout=180,
                )
            except Exception as exc:
                # If forced DB verification, do not fall back to API
                if source_decision == "db":
                    ui.info(f"  DB aggregation failed: {exc}")
                    return 0.0, 0, {}
                ui.info(f"  DB fast path failed, falling back to API: {exc}")
            else:
                total_hours = 0.0
                total_entries = 0
                project_hours: Dict[str, float] = defaultdict(float)
                project_rows = 0
                for line in result.stdout.splitlines():
                    stripped = line.strip()
                    if not stripped:
                        continue
                    parts = stripped.split("|")
                    if len(parts) < 3:
                        continue
                    try:
                        project_id = int(parts[0]) if parts[0] else None
                        hours_val = float(parts[1]) if parts[1] else 0.0
                        entries_val = int(parts[2]) if parts[2] else 0
                    except (ValueError, TypeError):
                        continue
                    name = project_id_to_name.get(project_id, f"Project {project_id}")
                    total_hours += hours_val
                    total_entries += entries_val
                    project_hours[name] += hours_val
                    project_rows += 1
                ui.info(
                    f"[source=DB] Aggregated {total_entries:,} entry(ies) across {project_rows:,} project row(s)",
                    "info",
                )
                return total_hours, total_entries, dict(project_hours)
        # If we intended to use DB and have not returned by now, respect forced DB mode
        if source_decision == "db":
            return 0.0, 0, {}

    total_hours = 0.0
    total_entries = 0
    project_hours: Dict[str, float] = defaultdict(float)
    
    # Get all time entries from OpenProject
    ui.info("  Fetching time entries from OpenProject using API...", "info")
    # Unified progress bar fetch (avoid per-page console spam)
    all_time_entries: List[dict] = []
    try:
        page_size = 100
        offset = 1
        seen_count = 0
        total_from_api: Optional[int] = None
        ui.progress(0, 1, "Fetching time entries")  # show bar start
        while True:
            resp = client._request("GET", "/time_entries", params={"offset": offset, "pageSize": page_size})
            payload = resp.json()
            if total_from_api is None:
                try:
                    total_from_api = int(payload.get("total")) if payload.get("total") is not None else None
                except Exception:
                    total_from_api = None
            elements = payload.get("_embedded", {}).get("elements", []) or []
            all_time_entries.extend(elements)
            seen_count += len(elements)
            if total_from_api and total_from_api > 0:
                ui.progress(min(seen_count, total_from_api), total_from_api, "Fetching time entries")
            else:
                ui.progress(min(offset, 100), 100, "Fetching time entries")
            next_link = payload.get("_links", {}).get("nextByOffset")
            if not elements or not next_link:
                break
            offset += 1
    except OpenProjectError as exc:
        ui.info(f"  Error fetching time entries: {exc}")
        return 0.0, 0, {}
    ui.info(f"[source=API] Loaded {len(all_time_entries):,} time entry record(s) from API for verification", "info")
    
    # Create a mapping from work package ID to project name
    # Use the project from the imported work package, not from the time entry
    wp_id_to_project_name: Dict[int, str] = {}
    project_id_to_name: Dict[int, str] = {}
    
    for staged_project in staging.projects.values():
        if staged_project.openproject_id:
            project_id_to_name[staged_project.openproject_id] = staged_project.name
    
    # Build mapping from imported packages
    # Use the project from imported packages (where work packages were created)
    # This is more accurate than using time entry's project link
    if imported_packages:
        for pkg in imported_packages:
            project_name = project_id_to_name.get(pkg.project_id, f"Project {pkg.project_id}")
            wp_id_to_project_name[pkg.work_package_id] = project_name
    else:
        # Verification-only mode: fetch project from work packages
        ui.info("  Building work package to project mapping from API...")
        all_wp_ids = set()
        for entry in all_time_entries:
            wp_href = (
                entry.get("_links", {}).get("workPackage", {}) or {}
            ).get("href")
            if wp_href:
                try:
                    wp_id = int(wp_href.split("/")[-1])
                    all_wp_ids.add(wp_id)
                except (ValueError, AttributeError):
                    continue
        
        # Fetch work packages to get their projects
        total_wp_to_fetch = len(all_wp_ids)
        if total_wp_to_fetch:
            ui.progress(0, total_wp_to_fetch, "Building WP→Project map")
        wp_progress_interval = max(1, total_wp_to_fetch // 100) if total_wp_to_fetch > 0 else None
        for idx_wp, wp_id in enumerate(all_wp_ids, start=1):
            try:
                wp = client.get_work_package(wp_id)
                if wp:
                    project_href = (
                        wp.get("_links", {}).get("project", {}) or {}
                    ).get("href")
                    if project_href:
                        try:
                            project_id = int(project_href.split("/")[-1])
                            project_name = project_id_to_name.get(
                                project_id, f"Project {project_id}"
                            )
                            wp_id_to_project_name[wp_id] = project_name
                            # reduce console noise; write detail to debug log only
                            _write_debug_log("debug", f"[wp_map] wp={wp_id} -> project_id={project_id} ({project_name})")
                        except (ValueError, AttributeError):
                            continue
            except OpenProjectError:
                continue
            # update progress every 1% (or last item)
            if wp_progress_interval and (idx_wp % wp_progress_interval == 0 or idx_wp == total_wp_to_fetch):
                ui.progress(idx_wp, total_wp_to_fetch, "Building WP→Project map")
    
    imported_wp_ids = set(wp_id_to_project_name.keys())
    
    ui.info(f"  Processing {len(all_time_entries)} time entry(ies)...")
    
    # Progress for processing entries
    total_entries_to_process = len(all_time_entries)
    if total_entries_to_process:
        ui.progress(0, total_entries_to_process, "Processing time entries")
    process_interval = max(1, total_entries_to_process // 100) if total_entries_to_process > 0 else None
    for idx_te, entry in enumerate(all_time_entries, start=1):
        if process_interval and (idx_te % process_interval == 0 or idx_te == total_entries_to_process):
            ui.progress(idx_te, total_entries_to_process, "Processing time entries")
        # If filtering by created time entry IDs, check if this entry is in the set
        if created_time_entry_ids is not None:
            entry_id = entry.get("id")
            if entry_id is None:
                continue
            try:
                entry_id_int = int(entry_id)
            except (ValueError, TypeError):
                continue
            if entry_id_int not in created_time_entry_ids:
                continue
        
        # Get work package ID
        wp_href = (
            entry.get("_links", {}).get("workPackage", {}) or {}
        ).get("href")
        if not wp_href:
            continue
        
        try:
            wp_id = int(wp_href.split("/")[-1])
        except (ValueError, AttributeError):
            continue
        
        # Only count entries for imported work packages
        if wp_id not in imported_wp_ids:
            continue
        
        # Get project name from the work package mapping (not from time entry)
        project_name = wp_id_to_project_name.get(wp_id)
        if not project_name:
            continue
        
        # Get hours from ISO duration
        hours_iso = entry.get("hours")
        if hours_iso:
            minutes = minutes_from_iso_duration(hours_iso)
            if minutes is not None:
                hours = minutes / 60.0
                total_hours += hours
                total_entries += 1
                project_hours[project_name] += hours
    
    return total_hours, total_entries, dict(project_hours)


def verify_import(
    workbook_path: Path,
    client: OpenProjectClient,
    staging: WorkbookStaging,
    imported_packages: List[ImportedWorkPackage],
    ui: ConsoleUI,
    tracer: Optional[Any] = None,
    created_time_entry_ids: Optional[Set[int]] = None,
) -> ImportVerification:
    """Verify import by comparing Excel totals with OpenProject totals.
    
    Args:
        created_time_entry_ids: Optional set of time entry IDs created in this run.
            If provided, only these entries will be counted in verification.
    
    Returns verification results.
    """
    ui.info("  Calculating totals from Excel workbook...")
    excel_hours, excel_entries, excel_project_hours = calculate_excel_totals(
        workbook_path, staging, tracer
    )
    
    ui.info("  Calculating totals from OpenProject...")
    op_hours, op_entries, op_project_hours = calculate_openproject_totals(
        client, staging, imported_packages, ui, created_time_entry_ids
    )
    
    # Combine project hours for comparison
    all_projects = set(excel_project_hours.keys()) | set(op_project_hours.keys())
    project_comparison: Dict[str, Dict[str, float]] = {}
    for project_name in sorted(all_projects):
        project_comparison[project_name] = {
            "excel_hours": excel_project_hours.get(project_name, 0.0),
            "openproject_hours": op_project_hours.get(project_name, 0.0),
        }
    
    verification = ImportVerification(
        excel_total_hours=excel_hours,
        excel_total_entries=excel_entries,
        openproject_total_hours=op_hours,
        openproject_total_entries=op_entries,
        work_package_count=len(imported_packages),
        project_comparison=project_comparison,
    )
    
    return verification


def analyze_missing_entries(
    workbook_path: Path,
    staging: WorkbookStaging,
    imported_packages: List[ImportedWorkPackage],
    client: OpenProjectClient,
    ui: ConsoleUI,
    created_time_entry_ids: Optional[Set[int]] = None,
) -> None:
    """Analyze which Excel rows don't have corresponding time entries."""
    import re
    
    ui.info("  Analyzing missing entries...")
    
    # Parse all Excel rows
    workbook_rows = parse_workbook(workbook_path)
    excel_rows_with_units = [
        row for row in workbook_rows
        if row.units is not None and row.units > 0
    ]
    
    # Build mapping from work package subject+project to OpenProject ID
    # Include all staged tasks that have an OpenProject ID (not just newly imported)
    # Also build a mapping by issue_id when available for more precise matching
    wp_subject_to_id: Dict[Tuple[str, str], int] = {}
    wp_issue_id_to_id: Dict[Tuple[str, str], int] = {}  # (issue_id, project_name) -> wp_id
    staged_wp_ids: Set[int] = set()
    for staged_task in staging.tasks.values():
        if staged_task.openproject_id is not None:
            key = (staged_task.subject.strip(), staged_task.project_name)
            wp_subject_to_id[key] = staged_task.openproject_id
            staged_wp_ids.add(staged_task.openproject_id)
            # Also index by issue_id if available
            if staged_task.issue_id:
                issue_key = (staged_task.issue_id.strip(), staged_task.project_name)
                wp_issue_id_to_id[issue_key] = staged_task.openproject_id
    
    # Get all time entries from OpenProject (respect verification source to avoid slow API pagination)
    all_time_entries: List[dict] = []
    if VERIFICATION_SOURCE == "db":
        ui.info("  Fetching time entries for analysis using Database...")
        for wp_id in sorted(staged_wp_ids):
            try:
                all_time_entries.extend(
                    fetch_time_entries_for_work_package_from_db(
                        wp_id,
                        ui=None,
                        base_url=client.base_url if hasattr(client, "base_url") else None,
                    )
                )
            except Exception as exc:
                ui.info(f"    Failed DB fetch for WP {wp_id}: {exc}")
                continue
    else:
        try:
            all_time_entries = client.list_time_entries()
        except OpenProjectError as exc:
            ui.info(f"  Error fetching time entries: {exc}")
            return
    
    # Filter to only entries for staged work packages
    all_time_entries = [
        entry for entry in all_time_entries
        if entry.get("_links", {}).get("workPackage", {}).get("href")
    ]
    # Extract work package IDs and filter
    filtered_entries = []
    for entry in all_time_entries:
        wp_href = (entry.get("_links", {}).get("workPackage", {}) or {}).get("href")
        if wp_href:
            try:
                wp_id = int(wp_href.split("/")[-1])
                if wp_id in staged_wp_ids:
                    filtered_entries.append(entry)
            except (ValueError, AttributeError):
                continue
    all_time_entries = filtered_entries
    
    # Filter to only created entries if specified
    if created_time_entry_ids is not None:
        all_time_entries = [
            entry for entry in all_time_entries
            if entry.get("id") and int(entry.get("id", 0)) in created_time_entry_ids
        ]
    
    # Build a set of time entry keys (date, hours, user, activity, wp_id)
    # Normalize dates to ISO format (YYYY-MM-DD) for consistent matching
    time_entry_keys: Set[Tuple[str, int, int, int, int]] = set()
    activity_ids_used_in_entries: Set[int] = set()  # Track which activity_ids are actually used
    for entry in all_time_entries:
        wp_href = (entry.get("_links", {}).get("workPackage", {}) or {}).get("href")
        if not wp_href:
            continue
        try:
            wp_id = int(wp_href.split("/")[-1])
        except (ValueError, AttributeError):
            continue
        
        spent_on = entry.get("spentOn")
        if not spent_on:
            continue
        # Normalize date to ISO format (YYYY-MM-DD) - extract just the date part if it includes time
        spent_on_normalized = str(spent_on).split("T")[0] if "T" in str(spent_on) else str(spent_on)
        
        hours_iso = entry.get("hours")
        minutes_value = minutes_from_iso_duration(hours_iso) or 0
        
        user_href = (entry.get("_links", {}).get("user", {}) or {}).get("href")
        try:
            user_id = int(user_href.split("/")[-1]) if user_href else None
        except (ValueError, AttributeError):
            user_id = None
        
        activity_href = (entry.get("_links", {}).get("activity", {}) or {}).get("href")
        try:
            activity_id = int(activity_href.split("/")[-1]) if activity_href else None
        except (ValueError, AttributeError):
            activity_id = None
        
        if user_id is None or activity_id is None:
            continue
        
        activity_ids_used_in_entries.add(activity_id)
        key = (spent_on_normalized, minutes_value, user_id, activity_id, wp_id)
        time_entry_keys.add(key)
    
    # Build user name to ID mapping from staging
    user_name_to_id: Dict[str, int] = {}
    for staged_user in staging.users.values():
        if staged_user.openproject_id is not None:
            user_name_to_id[staged_user.full_name] = staged_user.openproject_id
    
    # Build activity lookup the same way as step 10 (Validate time entry activities)
    # First, extract activities from Excel rows (like workbook_activity_lookup)
    workbook_activity_lookup: Dict[str, str] = {}
    for row in excel_rows_with_units:
        if row.activity:
            normalized_activity = normalize_activity_name(row.activity)
            if normalized_activity:
                workbook_activity_lookup.setdefault(
                    normalized_activity, row.activity.strip()
                )
    
    # Build activity_map from API or database (same as step 10)
    activity_map: Dict[str, Tuple[str, int]] = {}
    try:
        api_activities = client.list_time_entry_activities()
        for name, record in api_activities.items():
            record_id = record.get("id")
            if record_id:
                try:
                    record_id_int = int(record_id)
                    normalized_name = normalize_activity_name(name)
                    if normalized_name:
                        activity_map[normalized_name] = (name, record_id_int)
                except (ValueError, TypeError):
                    continue
    except OpenProjectError as exc:
        if "status 404" in str(exc):
            ui.info("  Fetching activities from database (API returned 404)...")
            db_records = fetch_time_entry_activities_from_db(base_url=getattr(client, "base_url", None))
            for name, activity_id in db_records.items():
                normalized_name = normalize_activity_name(name)
                if normalized_name:
                    activity_map[normalized_name] = (name, activity_id)
        else:
            ui.info(f"  ⚠️  Warning: Failed to fetch activities for analysis: {exc}")
    
    # Build activity_id_lookup from workbook_activity_lookup and activity_map (same as step 10)
    activity_id_lookup: Dict[str, int] = {}
    missing_activities_in_map: List[str] = []
    for normalized, original in workbook_activity_lookup.items():
        entry = activity_map.get(normalized)
        if entry:
            activity_id_lookup[normalized] = entry[1]
        else:
            missing_activities_in_map.append(f"{original} (normalized: {normalized})")
    
    if missing_activities_in_map:
        ui.info(f"  ⚠️  Warning: {len(missing_activities_in_map)} Excel activity(ies) not found in OpenProject activity_map:")
        for act in missing_activities_in_map[:5]:
            ui.info(f"    - {act}")
        if len(missing_activities_in_map) > 5:
            ui.info(f"    ... and {len(missing_activities_in_map) - 5} more")
    
    if not activity_id_lookup:
        ui.info("  ⚠️  Warning: No activity mappings found. Analysis may be inaccurate.")
    else:
        ui.info(f"  ✓ Built activity_id_lookup with {len(activity_id_lookup)} mapping(s)")
    
    # Check each Excel row
    missing_rows: List[dict] = []
    rows_without_wp: List[dict] = []
    rows_without_user: List[dict] = []
    
    for row in excel_rows_with_units:
        # Find matching work package
        # Excel rows may have full names like "User story #56849: Build Common components"
        # but work packages might have just "Build Common components"
        wp_id = None
        row_task_full = row.work_package.strip()
        row_task_subject = None
        
        # First, try to extract issue_id from Excel row and match by issue_id
        # Pattern to match issue ID: "Type #ID: " or "Task #123: "
        pattern = r'^[^:]*#(\d+):\s*'
        match = re.match(pattern, row_task_full)
        extracted_issue_id = None
        if match:
            extracted_issue_id = match.group(1)
            issue_key = (extracted_issue_id, row.project_name)
            if issue_key in wp_issue_id_to_id:
                wp_id = wp_issue_id_to_id[issue_key]
                # Found match by issue_id, skip to time entry check
                # (don't fall through to subject matching)
        
        # If we have an issue_id but didn't find a match, don't fall back to subject matching
        # (different issue_ids = different tasks, even if they have the same subject)
        if wp_id is None and extracted_issue_id:
            # We have an issue_id but no match - this is a missing work package
            # Don't try subject matching, just report it as missing
            pass
        # If no match by issue_id (and no issue_id in Excel row), try exact match by subject
        elif wp_id is None:
            for key, wp_id_val in wp_subject_to_id.items():
                if (key[0].strip() == row_task_full and
                        key[1] == row.project_name):
                    wp_id = wp_id_val
                    break
        
        # If no exact match, try matching by extracting subject from Excel row
        # (remove issue ID prefix like "User story #56849: " or "Task #123: ")
        if wp_id is None:
            # Pattern to match issue ID prefixes: "Type #ID: " or "Type #ID:"
            pattern = r'^[^:]*#\d+:\s*'
            row_task_subject = re.sub(pattern, '', row_task_full, count=1).strip()
            if row_task_subject != row_task_full:  # Only if we found a prefix
                for key, wp_id_val in wp_subject_to_id.items():
                    if (key[0].strip() == row_task_subject and
                            key[1] == row.project_name):
                        wp_id = wp_id_val
                        break
        
        # If still no match, try case-insensitive matching
        if wp_id is None:
            row_task_lower = row_task_full.lower()
            for key, wp_id_val in wp_subject_to_id.items():
                if (key[0].strip().lower() == row_task_lower and
                        key[1] == row.project_name):
                    wp_id = wp_id_val
                    break
        
        # If still no match, try matching subject part (after removing prefix) case-insensitively
        if wp_id is None and row_task_subject is not None:
            row_task_subject_lower = row_task_subject.lower()
            for key, wp_id_val in wp_subject_to_id.items():
                if (key[0].strip().lower() == row_task_subject_lower and
                        key[1] == row.project_name):
                    wp_id = wp_id_val
                    break
        
        if wp_id is None:
            rows_without_wp.append({
                "project": row.project_name,
                "task": row.work_package,
                "date": str(row.date_spent) if row.date_spent else "",
                "hours": row.units,
                "assignee": row.user_name or "",
            })
            continue
        
        # Find user ID
        user_id = user_name_to_id.get(row.user_name or "")
        if user_id is None:
            rows_without_user.append({
                "project": row.project_name,
                "task": row.work_package,
                "date": str(row.date_spent) if row.date_spent else "",
                "hours": row.units,
                "assignee": row.user_name or "",
            })
            continue
        
        # Check if time entry exists
        # Normalize date to ISO format (YYYY-MM-DD) - extract just the date part if it includes time
        date_str = str(row.date_spent) if row.date_spent else ""
        if date_str and "T" in date_str:
            date_str = date_str.split("T")[0]
        minutes_value = int(round((row.units or 0.0) * 60))
        
        # Try to find activity ID using the same logic as ensure_time_entries_for_task
        activity_id = None
        
        if not activity_id_lookup:
            missing_rows.append({
                "project": row.project_name,
                "task": row.work_package,
                "date": date_str,
                "hours": row.units,
                "assignee": row.user_name or "",
                "reason": "No activities found in OpenProject",
            })
            continue
        
        normalized_activity = normalize_activity_name(row.activity) if row.activity else ""
        if not normalized_activity:
            # Use default activity (first available) - same as ensure_time_entries_for_task
            if activity_id_lookup:
                normalized_activity = next(iter(activity_id_lookup.keys()))
        
        activity_id = activity_id_lookup.get(normalized_activity)
        
        if activity_id is None:
            # Debug: show what we're looking for vs what's available
            available_activities = list(activity_id_lookup.keys())
            reason = f"Activity '{row.activity or normalized_activity}' (normalized: '{normalized_activity}') not found in activity_id_lookup"
            if available_activities:
                reason += f" [Available: {', '.join(available_activities[:3])}{'...' if len(available_activities) > 3 else ''}]"
            missing_rows.append({
                "project": row.project_name,
                "task": row.work_package,
                "date": date_str,
                "hours": row.units,
                "assignee": row.user_name or "",
                "reason": reason,
            })
            continue
        
        key = (date_str, minutes_value, user_id, activity_id, wp_id)
        if key not in time_entry_keys:
            # Check if there are similar entries (same date, wp_id, user_id but different activity or minutes)
            similar_entries = []
            for existing_key in time_entry_keys:
                existing_date, existing_minutes, existing_user_id, existing_activity_id, existing_wp_id = existing_key
                if (existing_date == date_str and 
                    existing_wp_id == wp_id and 
                    existing_user_id == user_id):
                    similar_entries.append({
                        "minutes": existing_minutes,
                        "hours": existing_minutes / 60.0,
                        "activity_id": existing_activity_id,
                    })
            
            reason = "Time entry not found"
            if similar_entries:
                # Build a more detailed reason
                similar_strs = []
                for sim in similar_entries[:3]:  # Show up to 3 similar entries
                    similar_strs.append(
                        f"{sim['hours']:.2f}h (activity_id={sim['activity_id']}, minutes={sim['minutes']})"
                    )
                reason = f"Time entry not found (similar entries exist: {', '.join(similar_strs)})"
            else:
                # Check if there are entries with same date, wp_id, user_id but different activity
                # or entries with same date, wp_id, activity_id, user_id but different minutes
                entries_same_date_wp_user = [
                    (existing_date, existing_minutes, existing_user_id, existing_activity_id, existing_wp_id)
                    for existing_date, existing_minutes, existing_user_id, existing_activity_id, existing_wp_id in time_entry_keys
                    if existing_date == date_str and existing_wp_id == wp_id and existing_user_id == user_id
                ]
                if entries_same_date_wp_user:
                    reason += f" [Found {len(entries_same_date_wp_user)} entry(ies) with same date/wp/user but different activity/minutes]"
                else:
                    # Check if there are entries with same date, wp_id, activity_id, user_id but different minutes
                    entries_same_date_wp_activity_user = [
                        (existing_date, existing_minutes, existing_user_id, existing_activity_id, existing_wp_id)
                        for existing_date, existing_minutes, existing_user_id, existing_activity_id, existing_wp_id in time_entry_keys
                        if existing_date == date_str and existing_wp_id == wp_id and existing_activity_id == activity_id and existing_user_id == user_id
                    ]
                    if entries_same_date_wp_activity_user:
                        minutes_list = [m for _, m, _, _, _ in entries_same_date_wp_activity_user]
                        reason += f" [Found {len(entries_same_date_wp_activity_user)} entry(ies) with same date/wp/activity/user but different minutes: {minutes_list} vs {minutes_value}]"
            
            missing_rows.append({
                "project": row.project_name,
                "task": row.work_package,
                "date": date_str,
                "hours": row.units,
                "assignee": row.user_name or "",
                "reason": reason,
                "expected_key": {
                    "date": date_str,
                    "minutes": minutes_value,
                    "hours": row.units,
                    "user_id": user_id,
                    "activity_id": activity_id,
                    "wp_id": wp_id,
                },
                "similar_entries": similar_entries[:3] if similar_entries else [],
            })
    
    # Report findings
    if rows_without_wp:
        ui.warning(
            f"Found {len(rows_without_wp)} Excel row(s) without matching work package:"
        )
        for row in rows_without_wp[:10]:  # Show first 10
            ui.info(
                f"  • {row['project']} / {row['task']} / "
                f"{row['date']} / {row['hours']}h / {row['assignee']}",
                "warning"
            )
        if len(rows_without_wp) > 10:
            ui.info(f"  ... and {len(rows_without_wp) - 10} more", "warning")
    
    if rows_without_user:
        ui.warning(
            f"Found {len(rows_without_user)} Excel row(s) without matching user:"
        )
        for row in rows_without_user[:10]:  # Show first 10
            ui.info(
                f"  • {row['project']} / {row['task']} / "
                f"{row['date']} / {row['hours']}h / {row['assignee']}",
                "warning"
            )
        if len(rows_without_user) > 10:
            ui.info(f"  ... and {len(rows_without_user) - 10} more", "warning")
    
    if missing_rows:
        ui.warning(
            f"Found {len(missing_rows)} Excel row(s) without time entries:"
        )
        for row in missing_rows[:10]:  # Show first 10
            reason = row['reason']
            # Add detailed info if available
            if 'expected_key' in row and row['expected_key']:
                exp = row['expected_key']
                reason += (
                    f" [Looking for: {exp['hours']}h, "
                    f"user_id={exp['user_id']}, "
                    f"activity_id={exp['activity_id']}, "
                    f"wp_id={exp['wp_id']}]"
                )
            if 'similar_entries' in row and row['similar_entries']:
                sim_strs = [
                    f"{s['hours']:.2f}h(act={s['activity_id']})"
                    for s in row['similar_entries']
                ]
                reason += f" [Similar entries exist: {', '.join(sim_strs)}]"
            ui.info(
                f"  • {row['project']} / {row['task']} / "
                f"{row['date']} / {row['hours']}h / {row['assignee']} ({reason})",
                "warning"
            )
        if len(missing_rows) > 10:
            ui.info(f"  ... and {len(missing_rows) - 10} more", "warning")
    
    # Show activity_id usage summary
    if activity_ids_used_in_entries:
        activity_ids_in_lookup = set(activity_id_lookup.values())
        ui.section("Activity ID Analysis")
        ui.info(
            f"Activity IDs used in time entries: {sorted(activity_ids_used_in_entries)}",
            "info"
        )
        ui.info(
            f"Activity IDs in activity_id_lookup: {sorted(activity_ids_in_lookup)}",
            "info"
        )
        missing_activity_ids = activity_ids_used_in_entries - activity_ids_in_lookup
        if missing_activity_ids:
            ui.warning(
                f"Activity IDs used in entries but NOT in lookup: "
                f"{sorted(missing_activity_ids)}"
            )
        extra_activity_ids = activity_ids_in_lookup - activity_ids_used_in_entries
        if extra_activity_ids:
            ui.info(
                f"Activity IDs in lookup but NOT used in entries: "
                f"{sorted(extra_activity_ids)}",
                "info"
            )
    
    if not rows_without_wp and not rows_without_user and not missing_rows:
        ui.success("🎉 All Excel rows have corresponding time entries!")


def report_verification(verification: ImportVerification, ui: ConsoleUI) -> None:
    """Display beautiful verification report."""
    ui.section("📊 Import Verification Summary")
    
    excel_hours = verification.excel_total_hours
    excel_entries = verification.excel_total_entries
    op_hours = verification.openproject_total_hours
    op_entries = verification.openproject_total_entries
    
    ui.info("Overall Totals:", "info")
    ui.info(f"  Excel workbook:  {excel_entries:>6,} entries,  {excel_hours:>12,.2f} hours")
    ui.info(f"  OpenProject:     {op_entries:>6,} entries,  {op_hours:>12,.2f} hours")
    ui.info("")
    
    entry_diff = excel_entries - op_entries
    hours_diff = excel_hours - op_hours
    entry_diff_pct = (entry_diff / excel_entries * 100) if excel_entries > 0 else 0.0
    hours_diff_pct = (hours_diff / excel_hours * 100) if excel_hours > 0 else 0.0
    
    ui.info("Differences:", "info")
    diff_color = "success" if abs(entry_diff) < 1 and abs(hours_diff) < 0.01 else "warning"
    ui.info(f"  Entries:  {entry_diff:>+6,} ({entry_diff_pct:+.2f}%)", diff_color)
    ui.info(f"  Hours:    {hours_diff:>+12,.2f} hours ({hours_diff_pct:+.2f}%)", diff_color)
    ui.info("")
    
    if abs(entry_diff) < 1 and abs(hours_diff) < 0.01:
        ui.success("🎉 Import verification: PERFECT MATCH!")
    else:
        ui.warning("⚠️  Import verification: Differences detected")
    ui.info("")
    
    ui.info(f"Work Packages Imported: {verification.work_package_count:,}", "info")
    ui.info("")
    
    if verification.project_comparison:
        ui.section("Per-Project Breakdown")
        ui.table_header(
            "Project Name".ljust(50),
            "Excel Hours".rjust(12),
            "OP Hours".rjust(12),
            "Diff".rjust(12)
        )
        
        for project_name in sorted(verification.project_comparison.keys()):
            comp = verification.project_comparison[project_name]
            excel_h = comp["excel_hours"]
            op_h = comp["openproject_hours"]
            diff = excel_h - op_h
            diff_str = f"{diff:>+12.2f}"
            ui.table_row(
                project_name[:50].ljust(50),
                f"{excel_h:>12.2f}",
                f"{op_h:>12.2f}",
                diff_str
            )
        ui.info("")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import work packages into OpenProject."
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("assets/openproject/very short Openproject Data.xlsx"),
        help="Path to the Excel workbook to import.",
    )
    parser.add_argument(
        "--time-entry-cache",
        type=Path,
        default=DEFAULT_TIME_ENTRY_CACHE,
        help="Path to a JSON cache of created time entries to allow resume without duplicates.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help=(
            "Run validations and show planned actions without creating "
            "anything."
        ),
    )
    parser.add_argument(
        "--skip-project-creation",
        action="store_true",
        help="Do not create missing projects (fails if project absent).",
    )
    parser.add_argument(
        "--default-type",
        default="Task",
        help=(
            "Work package type to use when an activity from the workbook does "
            "not exist in OpenProject."
        ),
    )
    parser.add_argument(
        "--member-role",
        default="Member",
        help=(
            "Role name to assign when adding users to projects as members."
        ),
    )
    parser.add_argument(
        "--user-email-domain",
        default="galaxytechnology.vn",
        help=(
            "Email domain to use when creating placeholder accounts "
            "for missing users."
        ),
    )
    parser.add_argument(
        "--server",
        help="OpenProject server URL; if omitted you will be prompted.",
    )
    parser.add_argument(
        "--token",
        help="OpenProject API token; if omitted you will be prompted.",
    )
    parser.add_argument(
        "--staging-cache",
        type=Path,
        help="Debug only: path to read/write staging cache JSON.",
    )
    parser.add_argument(
        "--yes",
        action="store_true",
        help="Automatically confirm all interactive prompts.",
    )
    parser.add_argument(
        "--auto-assign-memberships",
        action="store_true",
        help=(
            "Automatically add missing project memberships when required "
            "for task assignments."
        ),
    )
    parser.add_argument(
        "--update-project-dates",
        action="store_true",
        help=(
            "Scan all projects and update their creation dates based on "
            "earliest work package/time entry dates. Skips full import."
        ),
    )
    parser.add_argument(
        "--analyze-only",
        action="store_true",
        help=(
            "Run only the analysis step to check which Excel rows are missing "
            "time entries. Skips full import."
        ),
    )
    parser.add_argument(
        "--update-logged-by",
        action="store_true",
        help=(
            "Update logged_by field for time entries based on Excel workbook. "
            "Requires --workbook. Skips full import."
        ),
    )
    parser.add_argument(
        "--op-version",
        choices=["auto", "v13", "v16"],
        default="auto",
        help="OpenProject version (auto-detect if not specified, default: auto)",
    )
    parser.add_argument(
        "--time-entry-source",
        choices=["auto", "db", "api"],
        default="auto",
        help=(
            "Source to read existing time entries when logging (Step 17). "
            "'db' = read per work-package directly from database; "
            "'api' = list via API and filter; "
            "'auto' = choose 'db' on v13, otherwise 'api'."
        ),
    )
    parser.add_argument(
        "--verification-source",
        choices=["auto", "db", "api"],
        default="auto",
        help=(
            "Source to read time entries for verification (Step 20). "
            "'db' = aggregate directly from database; "
            "'api' = list via API; "
            "'auto' = choose 'db' on v13 (without created-ID filter), otherwise 'api'."
        ),
    )
    parser.add_argument(
        "--logged-by-source",
        choices=["auto", "db", "api"],
        default="auto",
        help=(
            "Source to read time entries for Step 16 (logged_by updates). "
            "'db' = query per work package from database; "
            "'api' = list all via API; "
            "'auto' = choose 'db' on v13 with API fallback, otherwise 'api'."
        ),
    )
    parser.add_argument(
        "--clean-cache-and-logs",
        action="store_true",
        help=(
            "Delete local staging cache and import logs before running. "
            "Removes /tmp/staging-cache.json (or --staging-cache) and /tmp/op_import_run*.log"
        ),
    )
    parser.add_argument(
        "--debug-log",
        type=Path,
        default=Path("/tmp/op_import_debug.log"),
        help="Path to write a detailed debug log for this run.",
    )

    args = parser.parse_args(argv)

    server = args.server.strip().rstrip("/") if args.server else ""
    if not server:
        server_prompt = (
            "OpenProject server URL (e.g., https://openproject.example.com): "
        )
        server = input(server_prompt).strip().rstrip("/")
    if not server:
        print("Server URL is required.", file=sys.stderr)
        return 1

    token = (
        args.token.strip()
        if args.token
        else getpass("OpenProject API token: ").strip()
    )
    if not token:
        print("API token is required.", file=sys.stderr)
        return 1

    # Configure debug logging path
    global DEBUG_LOG_PATH
    DEBUG_LOG_PATH = args.debug_log
    ui = ConsoleUI(auto_confirm=args.yes)
    # Start wall-clock timer for whole process
    _process_start_ts = time_module.time()
    # Configure verification source for Step 20
    global VERIFICATION_SOURCE
    VERIFICATION_SOURCE = args.verification_source
    
    # Set default cache path if not provided
    if args.staging_cache is None:
        args.staging_cache = Path("/tmp/staging-cache.json")

    # Optional cleanup of cache and logs
    if args.clean_cache_and_logs:
        ui.start_step("Clean caches and logs")
        removed_any = False
        # Remove staging cache
        try:
            if args.staging_cache and args.staging_cache.exists():
                args.staging_cache.unlink()
                ui.info(f"Removed cache: {args.staging_cache}")
                removed_any = True
        except Exception as exc:
            ui.warning(f"Could not remove cache '{args.staging_cache}': {exc}")
        # Remove time-entry cache
        try:
            te_cache = args.time_entry_cache or DEFAULT_TIME_ENTRY_CACHE
            if te_cache and Path(te_cache).exists():
                Path(te_cache).unlink()
                ui.info(f"Removed cache: {te_cache}")
                removed_any = True
        except Exception as exc:
            ui.warning(f"Could not remove cache '{te_cache}': {exc}")
        # Remove import logs
        try:
            import glob as _glob
            for log_path in _glob.glob("/tmp/op_import_run*.log"):
                try:
                    Path(log_path).unlink()
                    ui.info(f"Removed log: {log_path}")
                    removed_any = True
                except Exception as exc:
                    ui.warning(f"Could not remove log '{log_path}': {exc}")
        except Exception as exc:
            ui.warning(f"Could not enumerate logs: {exc}")
        # Remove debug log
        try:
            if DEBUG_LOG_PATH and DEBUG_LOG_PATH.exists():
                DEBUG_LOG_PATH.unlink()
                ui.info(f"Removed log: {DEBUG_LOG_PATH}")
                removed_any = True
        except Exception as exc:
            ui.warning(f"Could not remove debug log '{DEBUG_LOG_PATH}': {exc}")
        if not removed_any:
            ui.skip_step("Nothing to remove.")
        else:
            ui.complete_step("Cache and logs cleaned.")

    # If --update-project-dates is set, run only that and exit
    if args.update_project_dates:
        client = OpenProjectClient(server, token)
        ui.start_step("Update project creation dates")
        scan_and_update_all_project_dates(
            client,
            dry_run=args.dry_run,
            ui=ui,
        )
        ui.complete_step("Project date update completed.")
        _print_total_duration(ui, _process_start_ts)
        return 0

    # If --analyze-only is set, run only analysis and exit
    if args.analyze_only:
        if not args.workbook.exists():
            print(
                "Workbook not found: %s" % args.workbook,
                file=sys.stderr,
            )
            return 1
        
        # Validate and load workbook
        ui.start_step("Validate workbook")
        validation_issues = validate_workbook(args.workbook, ui)
        report_validation_issues(validation_issues, ui)
        if validation_issues:
            ui.info(
                f"  Found {len(validation_issues)} validation issue(s). "
                "Analysis will continue with defaults for invalid data."
            )
        ui.complete_step("Workbook validation completed.")

        ui.start_step("Load workbook data")
        workbook_rows = parse_workbook(args.workbook)
        aggregated_tasks = aggregate_workbook_rows(workbook_rows)
        for task in aggregated_tasks:
            if not task.type_name:
                task.type_name = args.default_type
        if not aggregated_tasks:
            ui.complete_step("No valid rows found in workbook. Nothing to do.")
            return 0
        staging = build_staging(aggregated_tasks)
        ui.complete_step(
            "Prepared %d unique tasks across %d projects and %d users."
            % (
                len(staging.tasks),
                len(staging.projects),
                len(staging.users),
            )
        )

        # Connect to OpenProject
        op_version = args.op_version if args.op_version != "auto" else None
        client = OpenProjectClient(server, token, version=op_version)
        if client.version:
            ui.info(
                f"OpenProject version: {client.version} (auto-detected) "
                f"(auto-detected)" if op_version is None else f"(specified)",
                "info"
            )
        
        # Fetch projects and users to populate IDs
        ui.start_step("Fetch projects and users")
        project_records = client.list_projects()
        for staged_project in staging.projects.values():
            record = project_records.get(staged_project.name)
            if record:
                href = record["_links"]["self"]["href"]
                staged_project.openproject_id = int(href.split("/")[-1])
        user_records = client.list_users()
        for staged_user in staging.users.values():
            record = user_records.get(staged_user.full_name)
            if record:
                href = record["_links"]["self"]["href"]
                staged_user.openproject_id = int(href.split("/")[-1])
        resolved_projects = sum(
            1 for project in staging.projects.values() if project.openproject_id
        )
        resolved_users = sum(
            1 for user in staging.users.values() if user.openproject_id
        )
        ui.complete_step(
            f"Resolved {resolved_projects}/{len(staging.projects)} projects "
            f"and {resolved_users}/{len(staging.users)} users."
        )
        
        # Load staging cache if available (for better work package matching)
        staging_cache_path: Optional[Path] = args.staging_cache
        if staging_cache_path and staging_cache_path.exists():
            try:
                with staging_cache_path.open("r", encoding="utf-8") as handle:
                    cache_payload = json.load(handle)
                cache_assignments = 0
                for key, entry in cache_payload.items():
                    staged_task = staging.tasks.get(key)
                    if staged_task:
                        wp_id = entry.get("openproject_id")
                        if wp_id:
                            staged_task.openproject_id = int(wp_id)
                            cache_assignments += 1
                if cache_assignments > 0:
                    ui.info(
                        f"  Loaded {cache_assignments} cached work package mapping(s) from '{staging_cache_path}'."
                    )
            except (OSError, json.JSONDecodeError) as exc:
                ui.info(f"  Failed to load cache '{staging_cache_path}': {exc}")
        
        # Check existing work packages (simplified version for analysis)
        ui.start_step("Check existing work packages")
        tasks_to_check = list(staging.tasks.values())
        total_tasks = len(tasks_to_check)
        progress_interval = max(1, total_tasks // 100) if total_tasks > 0 else None
        matched_count = 0
        
        for idx, staged_task in enumerate(tasks_to_check, start=1):
            staged_project = staging.projects.get(staged_task.project_name)
            if not staged_project or staged_project.openproject_id is None:
                continue
            
            project_id = staged_project.openproject_id
            matched = False
            
            # Try matching by issue ID first
            if staged_task.issue_id and staged_task.issue_id.isdigit():
                try:
                    existing_wp = client.get_work_package(int(staged_task.issue_id))
                    if existing_wp:
                        project_link = (
                            existing_wp.get("_links", {})
                            .get("project", {})
                            .get("href", "")
                        )
                        try:
                            existing_project_id = int(project_link.split("/")[-1])
                        except (ValueError, AttributeError):
                            existing_project_id = None
                        if existing_project_id == project_id:
                            subject = (existing_wp.get("subject") or "").strip()
                            if subject == staged_task.subject.strip():
                                staged_task.openproject_id = int(existing_wp["id"])
                                matched = True
                                matched_count += 1
                        if matched:
                            if progress_interval and (
                                idx % progress_interval == 0 or idx == total_tasks
                            ):
                                ui.progress(
                                    idx,
                                    total_tasks,
                                    "Checking work packages",
                                )
                            continue
                except Exception:
                    pass
            
            # Try matching by subject (but only if we don't have an issue_id or if there's a unique match)
            if not matched:
                try:
                    matches = client.search_work_packages(
                        project_id=project_id,
                        subject=staged_task.subject,
                    )
                    matching_subjects = [
                        m for m in matches
                        if (m.get("subject", "").strip() == staged_task.subject.strip())
                    ]
                    # If we have an issue_id and multiple subject matches, don't match
                    # (different issue_ids = different tasks, even if they have the same subject)
                    if staged_task.issue_id and len(matching_subjects) > 1:
                        # Can't disambiguate - skip matching
                        pass
                    elif matching_subjects:
                        # Single match or no issue_id - use first match
                        staged_task.openproject_id = int(matching_subjects[0]["id"])
                        matched = True
                        matched_count += 1
                except Exception:
                    pass
            
            if progress_interval and (
                idx % progress_interval == 0 or idx == total_tasks
            ):
                ui.progress(
                    idx,
                    total_tasks,
                    "Checking work packages",
                )
        
        ui.complete_step(f"Found {matched_count} existing work package(s).")
        
        # Run verification and analysis
        ui.start_step("Verify import")
        verification = verify_import(
            args.workbook,
            client,
            staging,
            [],  # No imported packages in analyze-only mode
            ui,
            tracer=None,
            created_time_entry_ids=None,  # Check all entries
        )
        report_verification(verification, ui)

        # Always run analysis
        ui.info("")
        analyze_missing_entries(
            args.workbook,
            staging,
            [],  # No imported packages in analyze-only mode
            client,
            ui,
            created_time_entry_ids=None,  # Check all entries
        )

        ui.complete_step("Analysis completed.")
        _print_total_duration(ui, _process_start_ts)
        return 0

    # If --update-logged-by is set, update logged_by field and exit
    if args.update_logged_by:
        if not args.workbook.exists():
            print(
                "Workbook not found: %s" % args.workbook,
                file=sys.stderr,
            )
            return 1
        
        # Validate and load workbook
        ui.start_step("Validate workbook")
        validation_issues = validate_workbook(args.workbook, ui)
        report_validation_issues(validation_issues, ui)
        if validation_issues:
            ui.info(
                f"  Found {len(validation_issues)} validation issue(s). "
                "Update will continue with defaults for invalid data."
            )
        ui.complete_step("Workbook validation completed.")

        ui.start_step("Load workbook data")
        workbook_rows = parse_workbook(args.workbook)
        aggregated_tasks = aggregate_workbook_rows(workbook_rows)
        for task in aggregated_tasks:
            if not task.type_name:
                task.type_name = args.default_type
        if not aggregated_tasks:
            ui.complete_step("No valid rows found in workbook. Nothing to do.")
            return 0
        staging = build_staging(aggregated_tasks)
        ui.complete_step(
            "Prepared %d unique tasks across %d projects and %d users."
            % (
                len(staging.tasks),
                len(staging.projects),
                len(staging.users),
            )
        )

        # Connect to OpenProject
        op_version = args.op_version if args.op_version != "auto" else None
        client = OpenProjectClient(server, token, version=op_version)
        if client.version:
            ui.info(
                f"OpenProject version: {client.version} (auto-detected) "
                f"(auto-detected)" if op_version is None else f"(specified)",
                "info"
            )
        
        # Load staging cache if available (for work package matching)
        staging_cache_path: Optional[Path] = args.staging_cache or Path("/tmp/staging-cache.json")
        if staging_cache_path and staging_cache_path.exists():
            try:
                with staging_cache_path.open("r", encoding="utf-8") as handle:
                    cache_payload = json.load(handle)
                cache_assignments = 0
                for key, entry in cache_payload.items():
                    staged_task = staging.tasks.get(key)
                    if staged_task:
                        wp_id = entry.get("openproject_id")
                        if wp_id:
                            staged_task.openproject_id = int(wp_id)
                            cache_assignments += 1
                if cache_assignments > 0:
                    ui.info(
                        f"  Loaded {cache_assignments} cached work package mapping(s)."
                    )
            except (OSError, json.JSONDecodeError) as exc:
                ui.info(f"  Failed to load cache: {exc}")
        
        # Fetch projects and match work packages
        ui.start_step("Fetch projects and match work packages")
        project_records = client.list_projects()
        for staged_project in staging.projects.values():
            record = project_records.get(staged_project.name)
            if record:
                href = record["_links"]["self"]["href"]
                staged_project.openproject_id = int(href.split("/")[-1])
        
        # Match work packages by subject
        matched_count = 0
        for staged_task in staging.tasks.values():
            if staged_task.openproject_id:
                matched_count += 1
                continue
            staged_project = staging.projects.get(staged_task.project_name)
            if not staged_project or not staged_project.openproject_id:
                continue
            try:
                matches = client.search_work_packages(
                    project_id=staged_project.openproject_id,
                    subject=staged_task.subject,
                )
                matching_subjects = [
                    m for m in matches
                    if (m.get("subject", "").strip() == staged_task.subject.strip())
                ]
                if matching_subjects:
                    staged_task.openproject_id = int(matching_subjects[0]["id"])
                    matched_count += 1
            except Exception:
                pass
        ui.complete_step(f"Matched {matched_count} work package(s).")
        
        # Fetch users to build user_id_map
        ui.start_step("Fetch users")
        user_records = client.list_users()
        user_id_map: Dict[str, Optional[int]] = {}
        for user_name, record in user_records.items():
            href = record["_links"]["self"]["href"]
            user_id_map[normalize_person_name(user_name)] = int(href.split("/")[-1])
        ui.complete_step(f"Fetched {len(user_id_map)} user(s).")
        
        # Update logged_by field using the reusable function
        updated_count = update_time_entry_logged_by_from_excel(
            workbook_path=args.workbook,
            client=client,
            staging=staging,
            ui=ui,
            dry_run=args.dry_run,
        )
        
        if updated_count == 0:
            ui.complete_step("No time entries need logged_by update.")
        else:
            ui.complete_step(
                f"Updated logged_by for {updated_count} time entry(ies)."
            )
        _print_total_duration(ui, _process_start_ts)
        return 0

    if not args.workbook.exists():
        print(
            "Workbook not found: %s" % args.workbook,
            file=sys.stderr,
        )
        return 1

    ui.start_step("Validate workbook")
    validation_issues = validate_workbook(args.workbook, ui)
    report_validation_issues(validation_issues, ui)
    if validation_issues:
        ui.info(
            f"  Found {len(validation_issues)} validation issue(s). "
            "Import will continue with defaults for invalid data."
        )
    ui.complete_step("Workbook validation completed.")

    ui.start_step("Load workbook data")
    workbook_rows = parse_workbook(args.workbook)
    aggregated_tasks = aggregate_workbook_rows(workbook_rows)
    for task in aggregated_tasks:
        if not task.type_name:
            task.type_name = args.default_type
    if not aggregated_tasks:
        ui.complete_step("No valid rows found in workbook. Nothing to do.")
        return 0
    staging = build_staging(aggregated_tasks)
    
    # DEBUG: Limit removed - process all projects
    # if len(staging.projects) > 1:
    #     project_names = sorted(staging.projects.keys())[:1]
    #     ui.info(f"  [DEBUG] Limiting to first 1 project: {', '.join(project_names)}")
    #     # Filter projects
    #     staging.projects = {name: staging.projects[name] for name in project_names}
    #     # Filter tasks to only include those from the first 1 project
    #     staging.tasks = {
    #         key: task
    #         for key, task in staging.tasks.items()
    #         if task.project_name in project_names
    #     }
    #     # Filter users to only include those referenced by remaining tasks
    #     task_assignees = {task.assignee_name for task in staging.tasks.values() if task.assignee_name}
    #     staging.users = {
    #         key: user
    #         for key, user in staging.users.items()
    #         if user.full_name in task_assignees
    #     }
    #     ui.info(f"  [DEBUG] After filtering: {len(staging.tasks)} tasks, {len(staging.projects)} projects, {len(staging.users)} users")
    #     
    #     # Count total logs with units > 0
    #     total_logs_with_units = sum(
    #         sum(1 for log in task.aggregated.logs if log.units and log.units > 0)
    #         for task in staging.tasks.values()
    #     )
    #     ui.info(f"  [DEBUG] Total logs with units > 0: {total_logs_with_units}")
    
    ui.complete_step(
        f"Prepared {len(staging.tasks)} unique tasks across "
        f"{len(staging.projects)} projects and {len(staging.users)} users."
    )

    staging_cache_path: Optional[Path] = args.staging_cache
    workbook_activity_lookup: Dict[str, str] = {}
    for row in workbook_rows:
        if row.activity:
            normalized_activity = normalize_activity_name(row.activity)
            if normalized_activity:
                workbook_activity_lookup.setdefault(
                    normalized_activity, row.activity.strip()
                )
    activity_id_lookup: Dict[str, int] = {}
    ui.start_step("Load staging cache")
    cache_assignments = 0
    if staging_cache_path is None:
        ui.skip_step("No staging cache file specified.")
    else:
        if staging_cache_path.exists():
            try:
                with staging_cache_path.open("r", encoding="utf-8") as handle:
                    cache_payload = json.load(handle)
            except (OSError, json.JSONDecodeError) as exc:
                ui.skip_step(
                    "Failed to load cache '%s': %s"
                    % (staging_cache_path, exc)
                )
            else:
                for key, entry in cache_payload.items():
                    staged_task = staging.tasks.get(key)
                    if not staged_task:
                        continue
                    openproject_id = entry.get("openproject_id")
                    if openproject_id is None:
                        continue
                    staged_task.openproject_id = int(openproject_id)
                    staged_task.existing_match_details = entry.get(
                        "existing_match_details",
                        "Loaded from cache.",
                    )
                    cache_assignments += 1
                ui.complete_step(
                    "Loaded %s cached task mapping(s) from '%s'."
                    % (cache_assignments, staging_cache_path)
                )
        else:
            ui.skip_step(
                "Staging cache '%s' does not exist." % staging_cache_path
            )

    # Support version detection or manual specification
    op_version = args.op_version if args.op_version != "auto" else None
    client = OpenProjectClient(server, token, version=op_version)
    if client.version:
        ui.info(
            f"OpenProject version: {client.version} "
            f"(auto-detected)" if op_version is None else f"(specified)",
            "info"
        )
    # Baseline counts before import (captured for final verification summary)
    before_counts = get_openproject_counts(client, ui)

    ui.start_step("Fetch users from OpenProject")
    user_records = client.list_users()
    normalized_user_records: Dict[str, dict] = {
        normalize_person_name(name): record
        for name, record in user_records.items()
    }
    for staged_user in staging.users.values():
        record = normalized_user_records.get(staged_user.normalized_name)
        if record:
            staged_user.openproject_id = int(record["id"])
    resolved_users = sum(
        1 for user in staging.users.values() if user.openproject_id
    )
    missing_users = [
        staged_user
        for staged_user in staging.users.values()
        if staged_user.openproject_id is None
    ]
    ui.progress_summary(
        existing=resolved_users,
        missing=len(missing_users),
        entity_name="users"
    )
    ui.complete_step(
        f"Resolved {resolved_users}/{len(staging.users)} users; "
        f"{len(missing_users)} missing."
    )

    ui.start_step("Create missing users")
    if not missing_users:
        ui.skip_step("No user creation required.")
    elif args.dry_run:
        ui.skip_step(
            f"Dry-run: would create {len(missing_users)} user(s)."
        )
        return 0
    else:
        if not ui.ask_confirmation(
            f"Create {len(missing_users)} missing user(s) now?"
        ):
            ui.skip_step("User creation cancelled by operator.")
            print("Import aborted.")
            return 0
        email_domain = args.user_email_domain.lstrip("@")
        existing_logins: set[str] = {
            login
            for login in (
                record.get("login") for record in user_records.values()
            )
            if isinstance(login, str) and login
        }
        existing_emails: set[str] = {
            email
            for email in (
                record.get("email") for record in user_records.values()
            )
            if isinstance(email, str) and email
        }
        created_count = 0
        sorted_missing = sorted(
            missing_users,
            key=lambda user: user.full_name.lower(),
        )
        # 1% progress bar for user creation
        total_to_create_users = len(sorted_missing)
        if total_to_create_users:
            ui.progress(0, total_to_create_users, "Creating users")
        users_progress_interval = max(1, total_to_create_users // 100) if total_to_create_users > 0 else None
        for idx_u, staged_user in enumerate(sorted_missing, start=1):
            if users_progress_interval and (idx_u % users_progress_interval == 0 or idx_u == total_to_create_users):
                ui.progress(idx_u, total_to_create_users, "Creating users")
            login = build_unique_username(staged_user.full_name, existing_logins)
            # Ensure email uniqueness alongside username
            base_email_local = login
            email_local = base_email_local
            suffix = 1
            while f"{email_local}@{email_domain}" in existing_emails:
                suffix += 1
                email_local = f"{base_email_local}{suffix}"
            email = f"{email_local}@{email_domain}"
            password = secrets.token_urlsafe(12)
            try:
                created_user = client.create_user(
                    login=login,
                    first_name=staged_user.first_name,
                    last_name=staged_user.last_name,
                    email=email,
                    password=password,
                )
            except OpenProjectError as exc:
                ui.skip_step(
                    f"Failed to create user '{staged_user.full_name}': {exc}"
                )
                print("Import aborted due to user creation failure.",
                      file=sys.stderr)
                return 1
            display_name = created_user.get("name") or staged_user.full_name
            staged_user.openproject_id = int(created_user["id"])
            normalized_display = normalize_person_name(display_name)
            normalized_user_records[normalized_display] = created_user
            normalized_user_records[staged_user.normalized_name] = created_user
            user_records[display_name] = created_user
            existing_logins.add(login)
            existing_emails.add(email)
            created_count += 1
            ui.info(
                f"Created user '{display_name}' (login: {login})."
            )
        ui.progress_summary(
            existing=resolved_users,
            missing=len(missing_users) - created_count,
            created=created_count,
            entity_name="users"
        )
        ui.complete_step(f"Created {created_count} user(s).")

    ui.start_step("Fetch projects from OpenProject")
    project_records = client.list_projects()
    for staged_project in staging.projects.values():
        record = project_records.get(staged_project.name)
        if record:
            href = record["_links"]["self"]["href"]
            staged_project.openproject_id = int(href.split("/")[-1])
    resolved_projects = sum(
        1 for project in staging.projects.values() if project.openproject_id
    )
    missing_projects = [
        project
        for project in staging.projects.values()
        if project.openproject_id is None
    ]
    ui.progress_summary(
        existing=resolved_projects,
        missing=len(missing_projects),
        entity_name="projects"
    )
    ui.complete_step(
        f"Resolved {resolved_projects}/{len(staging.projects)} projects; "
        f"{len(missing_projects)} missing."
    )

    ui.start_step("Create missing projects")
    if not missing_projects:
        ui.skip_step("No project creation required.")
    elif args.dry_run:
        ui.skip_step(
            f"Dry-run: would create {len(missing_projects)} project(s)."
        )
        return 0
    else:
        if args.skip_project_creation:
            ui.skip_step("Project creation is disabled via CLI flag.")
            print("Import aborted because required projects are missing.")
            return 1
        if not ui.ask_confirmation(
            f"Create {len(missing_projects)} missing project(s) now?"
        ):
            ui.skip_step("Project creation cancelled by operator.")
            print("Import aborted.")
            return 0
        # 1% progress bar for project creation
        total_to_create_projects = len(missing_projects)
        if total_to_create_projects:
            ui.progress(0, total_to_create_projects, "Creating projects")
        proj_progress_interval = max(1, total_to_create_projects // 100) if total_to_create_projects > 0 else None
        for idx_p, staged_project in enumerate(missing_projects, start=1):
            if proj_progress_interval and (idx_p % proj_progress_interval == 0 or idx_p == total_to_create_projects):
                ui.progress(idx_p, total_to_create_projects, "Creating projects")
            created = client.create_project(staged_project.name)
            project_records[staged_project.name] = created
            staged_project.openproject_id = int(created["id"])
            ui.info(
                "Created project '%s' (ID %s)."
                % (staged_project.name, staged_project.openproject_id)
            )
            try:
                client.enable_project_module(staged_project.openproject_id, "backlogs")
            except OpenProjectError as exc:
                ui.info(
                    "  [modules] Failed to enable Backlogs module for '%s': %s"
                    % (staged_project.name, exc)
                )
        ui.complete_step(f"Created {len(missing_projects)} project(s).")

    ui.start_step("Fetch work package types")
    type_records = client.list_types()
    type_id_map = {name: int(record["id"])
                   for name, record in type_records.items()}
    workbook_type_names = {
        task.type_name for task in staging.tasks.values() if task.type_name
    }
    missing_type_names = sorted(
        name for name in workbook_type_names if name not in type_id_map
    )
    for missing_name in missing_type_names:
        ui.info(f"  Creating missing work package type '{missing_name}'...")
        try:
            created = client.create_type(
                missing_name,
                color="#f9f9f9",
                position=len(type_id_map) + 1,
                description="",
            )
        except OpenProjectError as exc:
            fallback_id: Optional[int] = None
            if "status 404" in str(exc):
                try:
                    fallback_id = create_type_in_db(missing_name)
                except OpenProjectError as db_exc:
                    ui.skip_step(
                        "Unable to create work package type '%s' via API "
                        "or database: %s" % (missing_name, db_exc)
                    )
                    return 1
                else:
                    ui.info(
                        "  Created work package type '%s' directly in "
                        "database (id %s)." % (missing_name, fallback_id)
                    )
                    # Refresh type records from API so subsequent steps see
                    # the newly inserted type.
                    type_records = client.list_types()
                    type_id_map = {
                        type_name: int(record["id"])
                        for type_name, record in type_records.items()
                    }
                    if missing_name not in type_id_map:
                        ui.skip_step(
                            "Work package type '%s' was created in the "
                            "database but is not visible via the API."
                            % missing_name
                        )
                        return 1
                    continue
            ui.skip_step(
                "Unable to create work package type '%s': %s"
                % (missing_name, exc)
            )
            return 1
        created_id = created.get("id")
        try:
            type_id_map[missing_name] = int(created_id)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            ui.skip_step(
                "Created work package type '%s' returned an invalid id: %r"
                % (missing_name, created_id)
            )
            return 1
        type_records[missing_name] = created
    unresolved_types = {
        task.type_name
        for task in staging.tasks.values()
        if task.type_name and task.type_name not in type_id_map
    }
    if unresolved_types:
        ui.skip_step(
            "Unable to resolve work package type(s): "
            + ", ".join(sorted(unresolved_types))
        )
        return 1
    for staged_task in staging.tasks.values():
        if staged_task.type_name:
            type_id = type_id_map.get(staged_task.type_name)
        else:
            type_id = type_id_map.get(args.default_type)
        if type_id is None:
            ui.skip_step(
                "Unable to determine type id for task subject '%s'."
                % staged_task.subject
            )
            return 1
        staged_task.resolved_type_id = type_id
    ui.complete_step(
        f"Loaded {len(type_records)} work package type(s)."
    )

    ui.start_step("Inspect project configuration")
    for staged_project in staging.projects.values():
        if staged_project.openproject_id is None:
            ui.skip_step(
                f"Project '{staged_project.name}' missing ID; cannot continue."
            )
            print(
                "Import aborted due to unresolved project IDs.",
                file=sys.stderr,
            )
            return 1
        project_id_value = int(staged_project.openproject_id)
        try:
            client.enable_project_module(project_id_value, "backlogs")
        except OpenProjectError as exc:
            ui.info(
                "  [modules] Failed to enable Backlogs module for '%s': %s"
                % (staged_project.name, exc)
            )
        project_info = client.get_project(project_id_value)
        allowed_ids = client.list_project_type_ids(project_id_value)
        staged_project.allowed_type_ids = allowed_ids
        staged_project.lock_version = project_info.get("lockVersion")
    ui.complete_step(
        f"Fetched configuration for {len(staging.projects)} project(s)."
    )

    ui.start_step("Validate time entry activities")
    if not workbook_activity_lookup:
        ui.skip_step("No time entry activity values found in workbook.")
    else:
        activity_map: Dict[str, Tuple[str, int]] = {}
        activity_source = "api"
        try:
            api_activities = client.list_time_entry_activities()
        except OpenProjectError as exc:
            if "status 404" in str(exc):
                activity_source = "database"
                db_records = fetch_time_entry_activities_from_db(base_url=server)
                for name, activity_id in db_records.items():
                    normalized_name = normalize_activity_name(name)
                    if not normalized_name:
                        continue
                    activity_map[normalized_name] = (name, activity_id)
                if not activity_map:
                    ui.skip_step(
                        "Unable to retrieve time entry activities via API "
                        "and database query returned no records."
                    )
                    return 1
            else:
                ui.skip_step(
                    "Unable to retrieve time entry activities via API: %s"
                    % exc
                )
                return 1
        else:
            for name, record in api_activities.items():
                normalized_name = normalize_activity_name(name)
                if not normalized_name:
                    continue
                record_id = record.get("id")
                if isinstance(record_id, int):
                    record_id_int = record_id
                elif isinstance(record_id, str):
                    try:
                        record_id_int = int(record_id)
                    except ValueError:
                        continue
                else:
                    continue
                activity_map[normalized_name] = (name, record_id_int)
            # v13 may return empty activities via API; fallback to DB if empty
            if client.version == "v13" and not activity_map:
                try:
                    db_records = fetch_time_entry_activities_from_db(base_url=server)
                    for dname, did in db_records.items():
                        n = normalize_activity_name(dname)
                        if n:
                            activity_map[n] = (dname, did)
                    activity_source = "database"
                except Exception:
                    pass
        missing_activity_names: List[str] = []
        for normalized, original in workbook_activity_lookup.items():
            entry = activity_map.get(normalized)
            if entry:
                activity_id_lookup[normalized] = entry[1]
            else:
                missing_activity_names.append(original)
        if missing_activity_names:
            unique_missing = sorted(
                {name.strip() for name in missing_activity_names}
            )
            ui.info(
                "  Creating missing time entry activities via database:"
            )
            for name in unique_missing:
                ui.info(f"    - {name}")
            try:
                create_time_entry_activities_in_db(unique_missing, base_url=server)
            except OpenProjectError as exc:
                ui.skip_step(
                    "Unable to create time entry activities via database: %s"
                    % exc
                )
                return 1
            db_records = fetch_time_entry_activities_from_db(base_url=server)
            for name, activity_id in db_records.items():
                normalized_name = normalize_activity_name(name)
                if not normalized_name:
                    continue
                activity_map[normalized_name] = (name, activity_id)
            still_missing: List[str] = []
            for normalized, original in workbook_activity_lookup.items():
                entry = activity_map.get(normalized)
                if entry:
                    activity_id_lookup[normalized] = entry[1]
                else:
                    still_missing.append(original)
            if still_missing:
                missing_lines = "\n".join(
                    f"  - {name}" for name in sorted(set(still_missing))
                )
                ui.skip_step(
                    "Unable to resolve time entry activities even after "
                    "database creation:\n" + missing_lines
                )
                return 1
            ui.info(
                "  Created %s time entry activit(ies)."
                % len(unique_missing)
            )
        else:
            for normalized, entry in activity_map.items():
                activity_id_lookup[normalized] = entry[1]
        validation_msg = (
            "Validated {count} time entry activity mapping(s)."
        ).format(count=len(activity_id_lookup))
        if activity_source == "database":
            validation_msg += " (Verified via database query.)"
        ui.complete_step(validation_msg)

    roles = client.list_roles()
    member_role_record = roles.get(args.member_role)
    if not member_role_record:
        ui.skip_step(
            f"Role '{args.member_role}' was not found in OpenProject."
        )
        return 1
    member_role_id = int(member_role_record["id"])

    project_required_type_ids: Dict[str, set[int]] = defaultdict(set)
    project_user_ids_required: Dict[str, set[int]] = defaultdict(set)
    user_id_map = {
        name: staged_user.openproject_id
        for name, staged_user in staging.users.items()
        if staged_user.openproject_id is not None
    }
    user_id_to_name: Dict[int, str] = {
        int(staged_user.openproject_id): staged_user.full_name
        for staged_user in staging.users.values()
        if staged_user.openproject_id is not None
    }
    for staged_task in staging.tasks.values():
        project_required_type_ids[staged_task.project_name].add(
            staged_task.resolved_type_id or fallback_type_id
        )
        assignee_normalized = normalize_person_name(staged_task.assignee_name)
        assignee_id = user_id_map.get(assignee_normalized)
        if assignee_id is not None:
            project_user_ids_required[staged_task.project_name].add(
                assignee_id)

    project_type_updates: List[Tuple[StagedProject, List[int]]] = []
    for project_name, required_ids in project_required_type_ids.items():
        staged_project = staging.projects[project_name]
        project_id_required: Optional[int] = staged_project.openproject_id
        if project_id_required is None:
            continue
        missing_type_ids = required_ids - staged_project.allowed_type_ids
        if not missing_type_ids:
            continue
        new_type_ids = sorted(staged_project.allowed_type_ids | required_ids)
        project_type_updates.append((staged_project, new_type_ids))

    ui.start_step("Ensure project type permissions")
    if not project_type_updates:
        ui.skip_step("All required types already enabled for each project.")
    elif args.dry_run:
        for staged_project, new_ids in project_type_updates:
            ui.info(
                "Would update project '%s' allowed types to %s."
                % (staged_project.name, new_ids)
            )
        ui.skip_step("Dry-run: skipping project type updates.")
    else:
        for staged_project, new_ids in project_type_updates:
            project_id_for_update: Optional[int] = (
                staged_project.openproject_id
            )
            if project_id_for_update is None:
                continue
            updated = client.update_project_types(
                int(project_id_for_update),
                new_ids,
                lock_version=staged_project.lock_version,
            )
            staged_project.lock_version = updated.get("lockVersion")
            refreshed_ids = client.list_project_type_ids(
                int(project_id_for_update)
            )
            staged_project.allowed_type_ids = refreshed_ids
            missing_post_update = set(new_ids) - refreshed_ids
            if missing_post_update:
                missing_ids_display = ", ".join(
                    str(tid) for tid in sorted(missing_post_update)
                )
                ui.info(
                    (
                        "  [types] Falling back to database assignment for "
                        "project '%s' (missing type id(s): %s)."
                    )
                    % (staged_project.name, missing_ids_display)
                )
                for type_id_value in sorted(missing_post_update):
                    assign_type_to_project_in_db(
                        int(project_id_for_update), int(type_id_value)
                    )
                refreshed_ids = client.list_project_type_ids(
                    int(project_id_for_update)
                )
                staged_project.allowed_type_ids = refreshed_ids
        ui.complete_step(f"Updated {len(project_type_updates)} project(s).")

    project_membership_additions: List[
        Tuple[StagedProject, List[int]]
    ] = []
    for project_name, required_user_ids in project_user_ids_required.items():
        staged_project = staging.projects[project_name]
        project_id_for_membership: Optional[int] = (
            staged_project.openproject_id
        )
        if project_id_for_membership is None:
            continue
        memberships = client.list_project_memberships(
            int(project_id_for_membership)
        )
        existing_user_ids: set[int] = set()
        for membership in memberships:
            principal = membership.get("_links", {}).get("principal", {})
            href = principal.get("href")
            if not href:
                continue
            try:
                existing_user_ids.add(int(href.split("/")[-1]))
            except ValueError:
                continue
        missing_member_ids = sorted(required_user_ids - existing_user_ids)
        if missing_member_ids:
            project_membership_additions.append(
                (staged_project, missing_member_ids)
            )

    ui.start_step("Ensure project memberships")
    if not project_membership_additions:
        ui.skip_step("All users already belong to their required projects.")
    else:
        ui.info("  Pending membership assignments by project:")
        for staged_project, member_ids in project_membership_additions:
            names = [
                user_id_to_name.get(user_id, f"User {user_id}")
                for user_id in member_ids
            ]
            ui.info(
                "    - %s: %s user(s) pending -> %s"
                % (
                    staged_project.name,
                    len(member_ids),
                    ", ".join(sorted(names)) or "None",
                )
            )
        if args.dry_run:
            ui.skip_step("Dry-run: membership updates not applied.")
        elif not args.auto_assign_memberships:
            # Membership assignment is required before creating work packages
            # to ensure assignees are valid project members
            if project_membership_additions:
                ui.info(
                    "  ⚠️  Membership assignment is required before creating work packages."
                )
                ui.info(
                    "  Please run with --auto-assign-memberships to assign users to projects."
                )
                raise OpenProjectError(
                    "Cannot proceed: users must be assigned to projects before "
                    "creating work packages. Use --auto-assign-memberships flag."
                )
            ui.skip_step("No membership updates required.")
        else:
            total_memberships = sum(
                len(member_ids)
                for _, member_ids in project_membership_additions
            )
            membership_prompt = (
                "Add {count} membership(s) across {projects} project(s)?"
            ).format(
                count=total_memberships,
                projects=len(project_membership_additions),
            )
            if not ui.ask_confirmation(membership_prompt):
                ui.skip_step("Membership updates cancelled by operator.")
                return 0
            added_count = 0
            # 1% progress across all membership additions
            if total_memberships:
                ui.progress(0, total_memberships, "Assigning memberships")
            mem_progress_interval = max(1, total_memberships // 100) if total_memberships > 0 else None
            for staged_project, member_ids in project_membership_additions:
                project_id_membership_update: Optional[int] = (
                    staged_project.openproject_id
                )
                if project_id_membership_update is None:
                    continue
                for user_id in member_ids:
                    try:
                        result = client.add_project_membership(
                            project_id=int(project_id_membership_update),
                            user_id=user_id,
                            role_id=member_role_id,
                        )
                        added_count += 1
                        if mem_progress_interval and (added_count % mem_progress_interval == 0 or added_count == total_memberships):
                            ui.progress(added_count, total_memberships, "Assigning memberships")
                    except OpenProjectError as e:
                        error_msg = str(e)
                        user_name = user_id_to_name.get(user_id, f"User {user_id}")
                        project_name = staged_project.name
                        
                        # Check if user is already a member (should not happen, but handle it)
                        if "already been taken" in error_msg.lower():
                            ui.warning(
                                f"  User '{user_name}' is already a member of "
                                f"project '{project_name}' (skipping)"
                            )
                            # User is already a member - this is fine, just skip
                            continue
                        
                        # User cannot be assigned - must fix before continuing
                        ui.error(
                            f"Failed to assign user '{user_name}' (ID: {user_id}) "
                            f"to project '{project_name}' (ID: {project_id_membership_update})"
                        )
                        ui.error(f"Error: {error_msg}")
                        
                        # Try to fix via API first
                        ui.info("  Attempting to fix via API...")
                        fixed = _try_fix_user_assignment_via_api(
                            client, user_id, project_id_membership_update, ui
                        )
                        
                        if not fixed:
                            # Try to fix via database
                            ui.info("  Attempting to fix via database...")
                            fixed = _try_fix_user_assignment_via_db(
                                user_id, project_id_membership_update, ui, base_url=getattr(client, "base_url", None)
                            )
                        
                        if not fixed:
                            # Cannot fix - provide instructions and STOP
                            ui.error("")
                            ui.error("=" * 70)
                            ui.error("CANNOT PROCEED: User assignment failed")
                            ui.error("=" * 70)
                            ui.error("")
                            ui.error(f"User: {user_name} (ID: {user_id})")
                            ui.error(f"Project: {project_name} (ID: {project_id_membership_update})")
                            ui.error("")
                            ui.error("Please fix this issue in OpenProject Web UI:")
                            base_url = client.base_url if hasattr(client, 'base_url') else "http://localhost:8081"
                            ui.error(f"  1. Go to: {base_url}/projects/{project_id_membership_update}/members")
                            ui.error(f"  2. Add user '{user_name}' to the project")
                            ui.error("  3. Ensure the user is active and unlocked")
                            ui.error("  4. Run the import script again")
                            ui.error("")
                            raise OpenProjectError(
                                f"Cannot assign user '{user_name}' to project '{project_name}'. "
                                "Please fix in OpenProject Web UI and try again."
                            )
                        
                        # Retry after fix
                        result = client.add_project_membership(
                            project_id=int(project_id_membership_update),
                            user_id=user_id,
                            role_id=member_role_id,
                        )
                        added_count += 1
                        if mem_progress_interval and (added_count % mem_progress_interval == 0 or added_count == total_memberships):
                            ui.progress(added_count, total_memberships, "Assigning memberships")
                        ui.success(f"  ✅ Fixed and assigned user '{user_name}'")
            
            ui.complete_step(
                f"Added {added_count} membership(s) across "
                f"{len(project_membership_additions)} project(s)."
            )

    ui.start_step("Check existing work packages")
    pending_tasks: List[StagedTask]
    cached_with_ids = [
        task for task in staging.tasks.values() if task.openproject_id is not None
    ]
    total_tasks = len(staging.tasks)
    stale_assignments = 0
    if cached_with_ids:
        ui.info(
            "  [cache] Verifying %s cached work package id(s)."
            % len(cached_with_ids)
        )
        verify_interval = max(1, len(cached_with_ids) // 100)
        for index, staged_task in enumerate(cached_with_ids, start=1):
            staged_project = staging.projects.get(staged_task.project_name)
            if staged_project is None or staged_project.openproject_id is None:
                staged_task.openproject_id = None
                staged_task.existing_match_details = None
                stale_assignments += 1
                continue
            work_package_id = staged_task.openproject_id
            existing_wp = client.get_work_package(int(work_package_id))
            if not existing_wp:
                stale_assignments += 1
                staged_task.openproject_id = None
                staged_task.existing_match_details = None
            else:
                project_link = (
                    existing_wp.get("_links", {})
                    .get("project", {})
                    .get("href", "")
                )
                try:
                    existing_project_id = int(project_link.split("/")[-1])
                except (ValueError, AttributeError):
                    existing_project_id = None
                if existing_project_id != int(staged_project.openproject_id):
                    stale_assignments += 1
                    staged_task.openproject_id = None
                    staged_task.existing_match_details = None
            if verify_interval and (
                index % verify_interval == 0 or index == len(cached_with_ids)
            ):
                ui.progress(
                    index,
                    len(cached_with_ids),
                    "Verifying cached IDs",
                )
        if stale_assignments:
            ui.info(
                "  [cache] Detected %s stale cached assignment(s); will recreate."
                % stale_assignments
            )
    tasks_to_check = [
        task for task in staging.tasks.values() if task.openproject_id is None
    ]
    if not tasks_to_check:
        pending_tasks = tasks_to_check
        ui.skip_step("All tasks already resolved from cache.")
    else:
        matched_by_id = 0
        matched_by_subject = 0
        total_tasks = len(tasks_to_check)
        progress_interval = max(1, total_tasks // 100)
        # Build reverse cache mapping once: openproject_id -> staging_key
        # This helps us match by issue_id when multiple work packages have the same subject
        reverse_cache: Dict[int, str] = {}
        if staging_cache_path and staging_cache_path.exists():
            try:
                with staging_cache_path.open("r", encoding="utf-8") as handle:
                    cache_payload = json.load(handle)
                    for cache_key, cache_entry in cache_payload.items():
                        wp_id = cache_entry.get("openproject_id")
                        if wp_id:
                            reverse_cache[int(wp_id)] = cache_key
            except (OSError, json.JSONDecodeError, (ValueError, TypeError)):
                pass  # Ignore cache errors, just proceed without it
        for idx, staged_task in enumerate(tasks_to_check, start=1):
            staged_project = staging.projects[staged_task.project_name]
            project_id = staged_project.openproject_id
            if project_id is None:
                continue
            matched = False
            # Note: Excel issue_id is NOT the OpenProject work package ID
            # We can't match by issue_id directly since OpenProject doesn't store it
            # The staging cache maps issue_id -> work_package_id for future runs
            # For now, we only match by subject (or use cache if available)
            if matched:
                if idx % progress_interval == 0 or idx == total_tasks:
                    ui.progress(
                        idx,
                        total_tasks,
                        "Checking work packages",
                    )
                continue
            matches = client.search_work_packages(
                project_id, staged_task.subject)
            # Try to find a match that has the same staging key (which includes issue_id)
            best_match = None
            matching_subjects = []
            for match in matches:
                subject = (match.get("subject") or "").strip()
                if subject == staged_task.subject.strip():
                    matching_subjects.append(match)
                    match_wp_id = int(match["id"])
                    # If we have a staging key (which includes issue_id), prefer matches
                    # that were created from a StagedTask with the same key
                    if staged_task.key and match_wp_id in reverse_cache:
                        cache_key = reverse_cache[match_wp_id]
                        if cache_key == staged_task.key:
                            # Perfect match - same staging key (includes issue_id)
                            best_match = match
                            break
                    elif best_match is None:
                        # Fallback: first subject match
                        best_match = match
            
            # If we have an issue_id and multiple subject matches but no cache match,
            # don't match to avoid incorrect associations
            # (different issue_ids = different tasks, even if they have the same subject)
            if (staged_task.issue_id and 
                len(matching_subjects) > 1 and 
                best_match and 
                (not staged_task.key or best_match["id"] not in reverse_cache or
                 reverse_cache.get(int(best_match["id"])) != staged_task.key)):
                # We have an issue_id but couldn't find a cache match
                # This means the work package might not have been created from this StagedTask
                # Don't match - let it create a new work package instead
                best_match = None
            
            if best_match:
                staged_task.openproject_id = int(best_match["id"])
                staged_task.existing_match_details = (
                    "Matched existing work package by subject."
                )
                matched_by_subject += 1
                matched = True
            if idx % progress_interval == 0 or idx == total_tasks:
                ui.progress(
                    idx,
                    total_tasks,
                    "Checking work packages",
                )
        pending_tasks = [
            task
            for task in staging.tasks.values()
            if task.openproject_id is None
        ]
        existing_wps = matched_by_id + matched_by_subject
        missing_wps = len(pending_tasks)
        ui.progress_summary(
            existing=existing_wps,
            missing=missing_wps,
            entity_name="work packages"
        )
        summary_msg = (
            "Found {matched} existing work package(s); "
            "{remaining} remaining to create."
        ).format(
            matched=existing_wps,
            remaining=missing_wps,
        )
        ui.complete_step(summary_msg)

    imported_packages_map: Dict[int, ImportedWorkPackage] = {}

    ui.start_step("Create new work packages")
    if not pending_tasks:
        ui.skip_step("No new work packages required.")
    elif args.dry_run:
        ui.skip_step(
            f"Dry-run: would create {len(pending_tasks)} work package(s)."
        )
    else:
        creation_prompt = (
            "Create {count} work package(s)?"
        ).format(count=len(pending_tasks))
        if not ui.ask_confirmation(creation_prompt):
            ui.skip_step("Work package creation cancelled by operator.")
            return 0

        project_type_cache: Dict[int, set[int]] = {}
        project_lock_versions: Dict[int, Optional[int]] = {}

        def refresh_project_types(project: StagedProject) -> set[int]:
            project_id_local = project.openproject_id
            if project_id_local is None:
                return set()
            project_id_int = int(project_id_local)
            type_ids = client.list_project_type_ids(project_id_int)
            project_type_cache[project_id_int] = type_ids
            info = client.get_project(project_id_int)
            project_lock_versions[project_id_int] = info.get("lockVersion")
            project.allowed_type_ids = type_ids
            project.lock_version = info.get("lockVersion")
            return type_ids

        created_count = 0
        total_tasks = len(pending_tasks)
        # Calculate interval for 1% progress updates
        progress_interval = max(1, total_tasks // 100) if total_tasks > 0 else None
        for index, staged_task in enumerate(pending_tasks, start=1):
            # Show progress every 1%
            if progress_interval and (
                index % progress_interval == 0 or index == total_tasks
            ):
                ui.progress(index, total_tasks, "Creating work packages")
            
            staged_project = staging.projects[staged_task.project_name]
            project_id_for_task: Optional[int] = staged_project.openproject_id
            assignee_name_raw = staged_task.assignee_name
            assignee_normalized = normalize_person_name(assignee_name_raw)
            assignee_id = user_id_map.get(assignee_normalized)
            if project_id_for_task is None or assignee_id is None:
                ui.info(
                    "Skipping task '%s' due to missing project/user mapping."
                    % staged_task.subject
                )
                continue
            project_id_int = int(project_id_for_task)
            required_type_id = int(
                staged_task.resolved_type_id or fallback_type_id
            )

            allowed_types = project_type_cache.get(project_id_int)
            if allowed_types is None:
                allowed_types = refresh_project_types(staged_project)

            if required_type_id not in allowed_types:
                ui.info(
                    "Project '%s' missing type ID %s; updating configuration."
                    % (staged_project.name, required_type_id)
                )
                lock_version = project_lock_versions.get(project_id_int)
                if lock_version is None:
                    lock_version = staged_project.lock_version
                new_type_ids = sorted(allowed_types | {required_type_id})
                client.update_project_types(
                    project_id_int,
                    new_type_ids,
                    lock_version=lock_version,
                )
                allowed_types = refresh_project_types(staged_project)
                if required_type_id not in allowed_types:
                    ui.info(
                        (
                            "  [types] Falling back to database assignment for "
                            "project '%s' and type %s."
                        )
                        % (staged_project.name, required_type_id)
                    )
                    assign_type_to_project_in_db(
                        project_id_int,
                        required_type_id,
                    )
                    allowed_types = refresh_project_types(staged_project)
                    if required_type_id not in allowed_types:
                        ui.skip_step(
                            "Project '%s' still lacks type %s after API and "
                            "database updates."
                            % (staged_project.name, required_type_id)
                        )
                        raise OpenProjectError(
                            "Configure project '%s' to allow type ID %s and rerun."
                            % (staged_project.name, required_type_id)
                        )

            # Create new work package (each StagedTask with different issue_id is a different task)
            estimated_time = iso_duration_from_hours(
                staged_task.aggregated.total_hours
            )
            try:
                created_wp = client.create_work_package(
                    project_id=project_id_int,
                    type_id=required_type_id,
                    subject=staged_task.subject,
                    assignee_id=int(assignee_id),
                    start_date=staged_task.aggregated.start_date,
                    due_date=staged_task.aggregated.due_date,
                    estimated_time=estimated_time,
                    description=None,
                )
            except OpenProjectError as exc:
                error_msg = str(exc)
                ui.error(
                    f"Failed to create work package '{staged_task.subject}' "
                    f"in project '{staged_project.name}'"
                )
                ui.error(f"Error: {error_msg}")
                
                # Check if it's a required custom field error
                if "can't be blank" in error_msg.lower() or "required" in error_msg.lower():
                    # Try to fix via API first
                    ui.info("  Attempting to fix required custom fields via API...")
                    fixed = _try_fix_required_custom_fields_via_api(
                        client, project_id_int, required_type_id, error_msg, ui
                    )
                    
                    if not fixed:
                        # Try to fix via database
                        ui.info("  Attempting to fix required custom fields via database...")
                        fixed = _try_fix_required_custom_fields_via_db(
                            project_id_int, error_msg, ui
                        )
                    
                    if not fixed:
                        # Cannot fix - provide instructions and STOP
                        ui.error("")
                        ui.error("=" * 70)
                        ui.error("CANNOT PROCEED: Required custom fields missing")
                        ui.error("=" * 70)
                        ui.error("")
                        ui.error(f"Work Package: {staged_task.subject}")
                        ui.error(f"Project: {staged_project.name} (ID: {project_id_int})")
                        ui.error("")
                        ui.error("The following custom fields are required but missing:")
                        if "customField15" in error_msg:
                            ui.error("  - customField15: QC Activity")
                        if "customField38" in error_msg:
                            ui.error("  - customField38: Finished Date")
                        ui.error("")
                        ui.error("Please fix this issue in OpenProject Web UI:")
                        base_url = client.base_url if hasattr(client, 'base_url') else "http://localhost:8081"
                        ui.error(f"  1. Go to: {base_url}/projects/{project_id_int}/settings/custom_fields")
                        ui.error("  2. Make these custom fields optional (not required)")
                        ui.error("  3. Or provide default values for these fields")
                        ui.error("  4. Run the import script again")
                        ui.error("")
                        raise OpenProjectError(
                            f"Cannot create work package '{staged_task.subject}' due to "
                            "required custom fields. Please fix in OpenProject Web UI and try again."
                        )
                    
                    # Retry after fix
                    ui.info("  Retrying work package creation after fix...")
                    created_wp = client.create_work_package(
                        project_id=project_id_int,
                        type_id=required_type_id,
                        subject=staged_task.subject,
                        assignee_id=int(assignee_id),
                        start_date=staged_task.aggregated.start_date,
                        due_date=staged_task.aggregated.due_date,
                        estimated_time=estimated_time,
                        description=None,
                    )
                    ui.success(f"  ✅ Fixed and created work package '{staged_task.subject}'")
                else:
                    # Other error - cannot fix, stop
                    ui.error("")
                    ui.error("=" * 70)
                    ui.error("CANNOT PROCEED: Work package creation failed")
                    ui.error("=" * 70)
                    ui.error("")
                    ui.error(f"Work Package: {staged_task.subject}")
                    ui.error(f"Project: {staged_project.name} (ID: {project_id_int})")
                    ui.error(f"Error: {error_msg}")
                    ui.error("")
                    raise
            work_package_id = int(created_wp["id"])
            staged_task.openproject_id = work_package_id
            imported_packages_map[work_package_id] = ImportedWorkPackage(
                task=staged_task.aggregated,
                project_id=project_id_int,
                work_package_id=work_package_id,
                time_entries=[],
            )
            created_count += 1
        ui.progress_summary(
            existing=existing_wps,
            missing=missing_wps - created_count,
            created=created_count,
            entity_name="work packages"
        )
        ui.complete_step(f"Created {created_count} work package(s).")
    
    # Save staging cache after all work packages are created (both matched and newly created)
    if staging_cache_path is not None:
        cache_payload = {
            task.key: {
                "openproject_id": task.openproject_id,
                "existing_match_details": task.existing_match_details,
            }
            for task in staging.tasks.values()
            if task.openproject_id is not None
        }
        try:
            staging_cache_path.parent.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
        try:
            with staging_cache_path.open("w", encoding="utf-8") as handle:
                json.dump(cache_payload, handle, indent=2)
        except OSError as exc:
            ui.info(
                "Failed to write staging cache '%s': %s"
                % (staging_cache_path, exc)
            )
        else:
            ui.info(
                "Saved staging cache with {count} entries to '{path}'.".format(
                    count=len(cache_payload),
                    path=staging_cache_path,
                )
            )

    ui.start_step("Log time entries")
    total_created_entries = 0
    total_missing_entries = 0
    # Load time-entry cache (resume support)
    time_entry_cache_path: Path = args.time_entry_cache or DEFAULT_TIME_ENTRY_CACHE
    cached_entries_by_wp: Dict[int, List[dict]] = {}
    if time_entry_cache_path.exists():
        try:
            with time_entry_cache_path.open("r", encoding="utf-8") as _h:
                raw_cache = json.load(_h)
            # keys are strings in JSON; convert to int
            for k, v in raw_cache.items():
                try:
                    cached_entries_by_wp[int(k)] = list(v) if isinstance(v, list) else []
                except Exception:
                    continue
            if cached_entries_by_wp:
                ui.info(f"Loaded cached time entries for {len(cached_entries_by_wp)} work package(s) (resume mode).", "info")
        except Exception as _exc:
            ui.warning(f"Failed to read time-entry cache '{time_entry_cache_path}': {_exc}")
    
    # Determine source for existing time entries
    source_decision = args.time_entry_source
    if source_decision == "auto":
        # Prefer DB on v13 where embedded DB access is available
        source_decision = "db" if client.version == "v13" else "api"
    ui.info(f"Using time entry source: {source_decision}", "info")
    # Unified fetch method banner for Step 17
    if source_decision == "db":
        ui.info("Fetching time entries from OpenProject using Database", "info")
    else:
        ui.info("Fetching time entries from OpenProject using API", "info")

    # Option A: DB fast-path (per work package)
    # Option B: API (single bulk list, then group by work package)
    time_entries_by_wp: Dict[int, List[dict]] = defaultdict(list)
    total_api_entries_fetched = 0
    total_db_entries_fetched = 0
    total_db_wps = 0
    if source_decision == "api":
        ui.info("Fetching time entries from OpenProject using API (may take time)...", "info")
        try:
            all_entries_api = client.list_time_entries()
        except OpenProjectError as exc:
            ui.warning(f"Failed to fetch time entries via API: {exc}")
            all_entries_api = []
        # Group by work package id
        for entry in all_entries_api:
            wp_href = (entry.get("_links", {}).get("workPackage", {}) or {}).get("href")
            if not wp_href:
                continue
            try:
                wp_id_val = int(wp_href.split("/")[-1])
            except (ValueError, AttributeError):
                continue
            time_entries_by_wp[wp_id_val].append(entry)
        total_api_entries_fetched = len(all_entries_api)
        ui.info(f"[source=API] Loaded {total_api_entries_fetched:,} time entry record(s) from API", "info")

    # Process all tasks with work package IDs, regardless of log validity
    loggable_tasks = [
        task
        for task in staging.tasks.values()
        if task.openproject_id is not None
    ]
    total_log_tasks = len(loggable_tasks)
    if total_log_tasks:
        ui.info(f"Verifying time entries for {total_log_tasks:,} task(s)", "info")
        # Show initial progress
        ui.progress(0, total_log_tasks, "Verifying time entries")
    # Calculate interval for 1% progress updates
    progress_interval = (
        max(1, total_log_tasks // 100) if total_log_tasks > 0 else None
    )

    # Track which work packages have been processed to avoid duplicates
    processed_wp_ids: Set[int] = set()
    
    for index, staged_task in enumerate(loggable_tasks, start=1):
        # Show progress every 1%
        if progress_interval and (
            index % progress_interval == 0 or index == total_log_tasks
        ):
            ui.progress(index, total_log_tasks, "Logging time entries")

        staged_project = staging.projects.get(staged_task.project_name)
        project_id_for_task = (
            staged_project.openproject_id if staged_project else None
        )
        wp_id_int = int(staged_task.openproject_id)
        
        # Skip if this work package has already been processed
        # (multiple staged_task objects can map to the same work package)
        if wp_id_int in processed_wp_ids:
            continue
        # Preload from cache first (resume support)
        if wp_id_int in cached_entries_by_wp and cached_entries_by_wp[wp_id_int]:
            time_entries_by_wp[wp_id_int].extend(cached_entries_by_wp[wp_id_int])
        
        # Populate existing time entries for this work package per source selection
        if source_decision == "db":
            if wp_id_int not in time_entries_by_wp:
                try:
                    # Suppress verbose DB query messages during progress
                    db_entries = fetch_time_entries_for_work_package_from_db(
                        wp_id_int,
                        ui=None,  # Don't show individual DB query messages
                        base_url=client.base_url if hasattr(client, "base_url") else None,
                    )
                    time_entries_by_wp[wp_id_int] = db_entries
                    total_db_entries_fetched += len(db_entries)
                    total_db_wps += 1
                except Exception as exc:
                    ui.warning(
                        f"  [logwork] Failed to fetch time entries from DB for WP {wp_id_int}: {exc}"
                    )
                    time_entries_by_wp[wp_id_int] = []
        else:
            # API path already grouped above; ensure key exists
            time_entries_by_wp.setdefault(wp_id_int, [])

        # Create missing entries for this work package
        existing_entries = time_entries_by_wp[wp_id_int]
        created_entries, missing_count = ensure_time_entries_for_task(
            staged_task,
            client=client,
            activity_id_lookup=activity_id_lookup,
            user_id_map=user_id_map,
            ui=ui,
            dry_run=args.dry_run,
            project_id=project_id_for_task,
            existing_entries=existing_entries,
        )
        total_missing_entries += missing_count
        # Mark this work package as processed even if no entries were created
        # to ensure we only process each work package once
        processed_wp_ids.add(wp_id_int)
        if not created_entries:
            continue
        if project_id_for_task is None:
            continue
        pkg = imported_packages_map.get(wp_id_int)
        if pkg is None:
            pkg = ImportedWorkPackage(
                task=staged_task.aggregated,
                project_id=int(project_id_for_task),
                work_package_id=wp_id_int,
                time_entries=[],
            )
            imported_packages_map[wp_id_int] = pkg
        pkg.time_entries.extend(created_entries)
        total_created_entries += len(created_entries)

    # Summarize the source used (post-processing)
    if source_decision == "db":
        ui.info(f"[source=DB] Loaded {total_db_entries_fetched:,} time entry record(s) across {total_db_wps:,} work package(s)", "info")
    
    # Calculate existing entries (total - missing - created)
    total_expected_entries = sum(
        len(task.aggregated.logs)
        for task in loggable_tasks
        if task.aggregated.logs
    )
    existing_entries = total_expected_entries - total_missing_entries - total_created_entries
    if total_expected_entries > 0:
        ui.progress_summary(
            existing=max(0, existing_entries),
            missing=total_missing_entries,
            created=total_created_entries,
            entity_name="time entries"
        )
    ui.info(f"Total logs processed: {total_log_tasks:,} tasks", "info")
    ui.success(f"Total entries created: {total_created_entries:,}")
    if total_missing_entries > 0:
        ui.info(f"Total entries missing (would create): {total_missing_entries:,}", "warning")
    if args.dry_run:
        if total_missing_entries:
            ui.skip_step(
                f"Dry-run: would create {total_missing_entries} time entry(ies)."
            )
        else:
            ui.skip_step("Dry-run: no time entries required.")
    else:
        if total_created_entries == 0:
            ui.skip_step("No time entries required.")
        else:
            ui.complete_step(
                f"Created {total_created_entries} time entry(ies)."
            )

    imported_packages = list(imported_packages_map.values())
    # Save time-entry cache after logging step (serialize API-like dicts only)
    try:
        if time_entries_by_wp:
            # Convert keys to str for JSON and ensure values are plain dicts/lists
            serializable_cache = {str(k): v for k, v in time_entries_by_wp.items()}
            time_entry_cache_path.parent.mkdir(parents=True, exist_ok=True)
            with time_entry_cache_path.open("w", encoding="utf-8") as _h:
                json.dump(serializable_cache, _h, indent=2)
            ui.info(f"Saved time-entry cache to '{time_entry_cache_path}'.", "info")
    except Exception as _exc:
        ui.warning(f"Failed to write time-entry cache '{time_entry_cache_path}': {_exc}")

    # Step 15.5: Update logged_by field for time entries
    ui.start_step("Update time entry logged_by field")
    if args.dry_run:
        ui.skip_step("Dry-run: would update logged_by field for time entries.")
    else:
        updated_count = update_time_entry_logged_by_from_excel(
            workbook_path=args.workbook,
            client=client,
            staging=staging,
            ui=ui,
            dry_run=False,
            source=args.logged_by_source,
        )
        if updated_count == 0:
            ui.skip_step("No time entries need logged_by update.")
        else:
            ui.complete_step(
                f"Updated logged_by for {updated_count} time entry(ies)."
            )

    ui.start_step("Adjust activity history")
    packages_dict = {pkg.work_package_id: pkg for pkg in imported_packages}
    history_packages = _collect_packages_for_history_adjustment(
        staging,
        packages_dict,
        ui,
        base_url=client.base_url if hasattr(client, "base_url") else None,
    )
    if not history_packages:
        ui.skip_step("No history adjustments required.")
    elif args.dry_run:
        ui.skip_step("Dry-run: would adjust journal timestamps for new items.")
    else:
        if not ui.ask_confirmation(
            "Apply history adjustments to align creation and log timestamps?"
        ):
            ui.skip_step("History adjustments skipped by operator.")
        else:
            apply_history_adjustments(history_packages, dry_run=False, ui=ui)
            ui.complete_step(
                "Updated journals and time entries in the database.")

    ui.start_step("Update project creation dates")
    if args.dry_run:
        ui.skip_step("Dry-run: would update project creation dates.")
    else:
        # Scan all projects and update their creation dates based on earliest
        # work package/time entry dates from the database
        # Limit scope to projects touched in this import when possible
        target_project_ids: Set[int] = set()
        if imported_packages:
            target_project_ids = {int(pkg.project_id) for pkg in imported_packages if pkg.project_id is not None}
        elif staging and staging.projects:
            # fallback to projects referenced in the workbook
            for sp in staging.projects.values():
                if sp.openproject_id is not None:
                    target_project_ids.add(int(sp.openproject_id))
        scan_and_update_all_project_dates(
            client,
            dry_run=False,
            ui=ui,
            limit_project_ids=target_project_ids if target_project_ids else None,
        )
        ui.complete_step("Updated project creation dates.")

    ui.start_step("Verify import")
    if args.dry_run:
        ui.skip_step("Dry-run: skipping import verification.")
    else:
        # Collect all time entry IDs created in this run
        created_time_entry_ids: Set[int] = set()
        for pkg in imported_packages:
            for te in pkg.time_entries:
                created_time_entry_ids.add(te.id)
        
        tracer = None  # DebugTracer not implemented yet, set to None
        verification = verify_import(
            args.workbook,
            client,
            staging,
            imported_packages,
            ui,
            tracer,
            created_time_entry_ids=created_time_entry_ids if created_time_entry_ids else None,
        )
        report_verification(verification, ui)
        
        # If there are differences, analyze what's missing
        entry_diff = verification.excel_total_entries - verification.openproject_total_entries
        hours_diff = verification.excel_total_hours - verification.openproject_total_hours
        if abs(entry_diff) >= 1 or abs(hours_diff) >= 0.01:
            ui.info("")
            analyze_missing_entries(
                args.workbook,
                staging,
                imported_packages,
                client,
                ui,
                created_time_entry_ids=created_time_entry_ids if created_time_entry_ids else None,
            )
        
        # Merge before/after instance counts into the verification section
        after_counts = get_openproject_counts(client, ui)
        ui.section("Instance Totals (Before → After)")
        # Use fixed-width columns for alignment
        metric_w = 18
        num_w = 8
        delta_w = 6
        ui.table_header(
            "Metric".ljust(metric_w),
            "Before".rjust(num_w),
            "After".rjust(num_w),
            "Δ".rjust(delta_w),
        )
        proj_b = before_counts.get("projects", 0); proj_a = after_counts.get("projects", 0)
        user_b = before_counts.get("users", 0); user_a = after_counts.get("users", 0)
        wp_b = before_counts.get("work_packages", 0); wp_a = after_counts.get("work_packages", 0)
        ui.table_row(
            "Projects".ljust(metric_w),
            f"{proj_b:>{num_w}d}",
            f"{proj_a:>{num_w}d}",
            f"{(proj_a - proj_b):>{delta_w}d}",
        )
        ui.table_row(
            "Users".ljust(metric_w),
            f"{user_b:>{num_w}d}",
            f"{user_a:>{num_w}d}",
            f"{(user_a - user_b):>{delta_w}d}",
        )
        ui.table_row(
            "Work packages".ljust(metric_w),
            f"{wp_b:>{num_w}d}",
            f"{wp_a:>{num_w}d}",
            f"{(wp_b and wp_a - wp_b or 0):>{delta_w}d}",
        )
        print()
        ui.complete_step("Import verification completed.")

    ui.complete_step("Import completed successfully.")
    _print_total_duration(ui, _process_start_ts)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except OpenProjectError as exc:
        print(f"OpenProject API error: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
